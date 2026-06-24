from pathlib import Path

from openpyxl import load_workbook

from cad_budget.export_excel import HEADERS, export_quantity_result
from cad_budget.models import (
    ConstructionKind,
    ConstructionQuantityDetail,
    DataStatus,
    ExteriorQuantityRow,
    HeightMode,
    Point,
    ProjectInput,
    QuantityRow,
    QuantityResult,
    RoomBoundary,
    SpaceType,
)
from cad_budget.quantity import calculate_quantities


def test_export_quantity_result_creates_workbook(tmp_path: Path):
    project_name = "Export Demo"
    project = ProjectInput(
        project_name=project_name,
        rooms=[
            RoomBoundary(
                id="room",
                name="书房",
                points=[
                    Point(x=0, y=0),
                    Point(x=3, y=0),
                    Point(x=3, y=3),
                    Point(x=0, y=3),
                    Point(x=0, y=0),
                ],
            )
        ],
    )
    result = calculate_quantities(project)
    output = tmp_path / "takeoff.xlsx"

    export_quantity_result(result, output)

    workbook = load_workbook(output)
    sheet = workbook["算量表"]
    assert sheet["A1"].value == "项目名称"
    assert sheet["B1"].value == project_name
    assert sheet["A3"].value == "楼层"
    assert sheet["A3"].value == HEADERS[0]
    assert sheet["B4"].value == "书房"
    assert [cell.value for cell in sheet[3][: len(HEADERS)]] == HEADERS
    assert sheet.freeze_panes == "A4"
    assert sheet.auto_filter.ref == f"A3:U{sheet.max_row}"
    assert sheet.column_dimensions["B"].width >= 12
    assert sheet["T3"].value == "空间ID"
    assert sheet["T4"].value == "room"
    assert sheet.column_dimensions["T"].hidden is True
    assert sheet["U3"].value == "报价明细JSON"
    assert sheet.column_dimensions["U"].hidden is True


def test_export_quantity_result_writes_building_area_metadata(tmp_path: Path):
    result = QuantityResult(project_name="Area Demo", rows=[], building_area=88.6, exceptions=[])
    output = tmp_path / "takeoff.xlsx"

    export_quantity_result(result, output)

    sheet = load_workbook(output).active
    assert sheet["A2"].value == "建筑面积"
    assert sheet["B2"].value == 88.6
    assert [cell.value for cell in sheet[3][: len(HEADERS)]] == HEADERS


def test_export_quantity_result_uses_formulas_and_editable_styles(tmp_path: Path):
    result = QuantityResult(
        project_name="Formula Demo",
        rows=[
            QuantityRow(
                room_id="room",
                floor="1F",
                room_name="书房",
                space_type=SpaceType.NORMAL,
                height=2.8,
                height_mode=HeightMode.PROJECT_DEFAULT,
                floor_area=9.0,
                floor_perimeter=12.0,
                wall_measure_perimeter=11.5,
                open_boundary_length=0.5,
                gross_wall_area=32.2,
                window_count=1,
                window_area=1.8,
                door_opening_count=1,
                door_opening_area=0.0,
                net_wall_area=30.4,
                is_outdoor=False,
                include_in_floor_quantity=True,
                include_in_wall_paint_quantity=True,
                status=DataStatus.DEFAULT_INFERRED,
                exception_notes=[],
            )
        ],
        exceptions=[],
    )
    output = tmp_path / "formula_takeoff.xlsx"

    export_quantity_result(result, output)

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["算量表"]
    assert sheet["I4"].value == "=G4*D4"
    assert sheet["N4"].value == "=I4-K4"
    assert sheet["D4"].number_format == "0.###"
    assert sheet["I4"].number_format == "0.###"
    assert sheet["D4"].fill.fill_type == "solid"
    assert sheet["I4"].fill.fill_type == "solid"
    assert sheet["D4"].fill.fgColor.rgb != sheet["I4"].fill.fgColor.rgb


def test_export_quantity_result_joins_exception_notes(tmp_path: Path):
    result = QuantityResult(
        project_name="Exception Demo",
        rows=[
            QuantityRow(
                room_id="room",
                floor="1F",
                room_name="书房",
                space_type=SpaceType.NORMAL,
                height=2.8,
                height_mode=HeightMode.PROJECT_DEFAULT,
                floor_area=9.0,
                floor_perimeter=12.0,
                wall_measure_perimeter=12.0,
                open_boundary_length=0.0,
                gross_wall_area=33.6,
                window_count=0,
                window_area=0.0,
                door_opening_count=0,
                door_opening_area=0.0,
                net_wall_area=33.6,
                is_outdoor=False,
                include_in_floor_quantity=True,
                include_in_wall_paint_quantity=True,
                status=DataStatus.CONFIRMED,
                exception_notes=["异常编码A", "编码描述B"],
            )
        ],
        exceptions=[],
    )
    output = tmp_path / "takeoff_with_exception.xlsx"

    export_quantity_result(result, output)

    workbook = load_workbook(output)
    sheet = workbook["算量表"]
    assert sheet["S4"].value == "异常编码A；编码描述B"


def test_export_quantity_result_creates_exterior_sheet(tmp_path: Path):
    result = QuantityResult(
        project_name="Exterior Demo",
        rows=[],
        exterior_rows=[
            ExteriorQuantityRow(
                exterior_wall_id="ext-wall-1",
                floor="1F",
                height=3.0,
                measure_length=4.0,
                opening_length=1.0,
                gross_area=12.0,
                net_area=9.0,
            )
        ],
        exceptions=[],
    )
    output = tmp_path / "exterior_takeoff.xlsx"

    export_quantity_result(result, output)

    workbook = load_workbook(output)
    sheet = workbook["外墙表"]
    assert sheet["A1"].value == "项目名称"
    assert sheet["B1"].value == "Exterior Demo"
    assert sheet["A3"].value == "楼层"
    assert sheet["B3"].value == "外墙编号"
    assert sheet["B4"].value == "ext-wall-1"
    assert sheet["D4"].value == 4.0
    assert sheet["F4"].value == 12.0
    assert sheet["G4"].value == 9.0
    assert sheet.freeze_panes == "A4"
    assert sheet.auto_filter.ref == "A3:G4"


def test_export_quantity_result_creates_construction_sheet(tmp_path: Path):
    result = QuantityResult(
        project_name="Construction Demo",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="demo-1",
                kind=ConstructionKind.DEMO_WALL,
                floor="1F",
                length=3.0,
                effective_height=2.8,
                height_defaulted=True,
                area=8.4,
                count=1,
            )
        ],
        exceptions=[],
    )
    output = tmp_path / "construction_takeoff.xlsx"

    export_quantity_result(result, output)

    workbook = load_workbook(output)
    sheet = workbook["施工标识表"]
    assert sheet["A1"].value == "项目名称"
    assert sheet["B1"].value == "Construction Demo"
    assert sheet["B3"].value == "标识编号"
    assert sheet["B4"].value == "demo-1"
    assert sheet["C4"].value == ConstructionKind.DEMO_WALL.value
    assert sheet["D4"].value == 3.0
    assert sheet["F4"].value == 2.8
    assert sheet["G4"].value is True
    assert sheet["I4"].value == 8.4
    assert sheet.freeze_panes == "A4"
    assert sheet.auto_filter.ref == "A3:J4"
