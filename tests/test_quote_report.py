import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cad_budget.models import (
    DataStatus,
    DoorQuantityDetail,
    FixtureKind,
    FixturePricingMode,
    FixtureQuantityDetail,
    HeightMode,
    QuantityResult,
    QuantityRow,
    SpaceType,
    WindowQuantityDetail,
)
from cad_budget.quote_report import generate_quote_review_report


def test_generate_quote_review_report_groups_review_rows(tmp_path: Path):
    quote_path = tmp_path / "quote.xlsx"
    report_path = tmp_path / "report.md"
    _write_quote_workbook(quote_path)

    report_text = generate_quote_review_report(quote_path, report_path)

    assert report_path.read_text(encoding="utf-8") == report_text
    assert "# 报价复核报告" in report_text
    assert "- 自动生成-默认推断：6 行" in report_text
    assert "- 自动生成-异常提示：1 行" in report_text
    assert "- 按模板生成：1 行" in report_text
    assert "## 复核行动建议" in report_text
    assert "- 补窗高：影响 2 个报价行，涉及项目：卧室墙面项目、墙面乳胶漆；Excel 行 5、6" in report_text
    assert "- 补新砌墙高度/厚度：影响 1 个报价行，涉及项目：砌240厚砖墙；Excel 行 7" in report_text
    assert "- 补门洞/推拉门高度：影响 1 个报价行，涉及项目：厨房推拉门；Excel 行 9" in report_text
    assert "- 补全屋定制高度/类型：影响 1 个报价行，涉及项目：全屋定制；Excel 行 10" in report_text
    assert "- 补管道/包管标识：影响 1 个报价行，涉及项目：包上/下水管道(单管)；Excel 行 11" in report_text
    assert "- 复核外墙修补范围：影响 1 个报价行，涉及项目：外墙修补；Excel 行 8" in report_text
    assert "| 5 | 101 | 卧室墙面项目 | 31.5 | 墙面净面积 | 自动生成-默认推断 | 窗高缺失 1 个 |" in report_text
    assert "| 6 | 102 | 墙面乳胶漆 | 31.5 | 墙面净面积 | 自动生成-默认推断 | 窗高缺失 1 个 |" in report_text
    assert "| 7 | 103 | 砌240厚砖墙 | 6 | 新砌240mm砖墙面积汇总 | 自动生成-默认推断 | 墙体标识缺少高度 |" in report_text
    assert "| 8 | 104 | 外墙修补 | 0 | 外墙修补范围面积汇总 | 自动生成-异常提示 | 需要确认修补范围 |" in report_text
    assert "| 12 | 107 | 主材包 | 1 |  | 按模板生成 |  |" in report_text
    assert "普通自动行" not in report_text


def test_generate_quote_review_report_uses_quantity_result_for_room_object_summary(tmp_path: Path):
    quote_path = tmp_path / "quote.xlsx"
    report_path = tmp_path / "report.md"
    _write_quote_workbook(quote_path)
    quantity_result = QuantityResult(
        project_name="对象摘要",
        rows=[
            _quantity_row(
                "bed",
                "卧室",
                window_details=[
                    WindowQuantityDetail(id="w1", width=1.2, height=1.8, area=2.16, height_defaulted=True),
                    WindowQuantityDetail(id="w2", width=1.0, height=1.8, area=1.8, height_defaulted=True),
                ],
            ),
            _quantity_row(
                "kitchen",
                "厨房",
                door_details=[
                    DoorQuantityDetail(
                        id="door-kitchen",
                        room_id="kitchen",
                        width=1.8,
                        height=2.2,
                        effective_height=2.2,
                        height_defaulted=True,
                        area=3.96,
                    )
                ],
            ),
            _quantity_row(
                "bath-a",
                "卫生间",
                window_details=[
                    WindowQuantityDetail(id="bath-a-window", width=0.8, height=1.8, area=1.44, height_defaulted=True),
                ],
                door_details=[
                    DoorQuantityDetail(
                        id="bath-a-door",
                        room_id="bath-a",
                        width=0.8,
                        height=2.2,
                        effective_height=2.2,
                        height_defaulted=True,
                        area=1.76,
                    )
                ],
            ),
            _quantity_row(
                "bath-b",
                "卫生间",
                window_details=[
                    WindowQuantityDetail(id="bath-b-window", width=0.8, height=1.8, area=1.44, height_defaulted=True),
                ],
                door_details=[
                    DoorQuantityDetail(
                        id="bath-b-door",
                        room_id="bath-b",
                        width=0.8,
                        height=2.2,
                        effective_height=2.2,
                        height_defaulted=True,
                        area=1.76,
                    )
                ],
            ),
            _quantity_row(
                "master",
                "主卧",
                custom_details=[
                    FixtureQuantityDetail(
                        id="custom-wardrobe",
                        room_id="master",
                        room_name="主卧",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        height=None,
                        effective_height=2.6,
                        height_defaulted=True,
                        projected_area=5.2,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                        fixture_type=None,
                    )
                ],
            ),
        ],
        construction_details=[],
        exceptions=[],
    )

    report_text = generate_quote_review_report(quote_path, report_path, quantity_result=quantity_result)

    assert "涉及对象：卧室窗高 2 个、卫生间窗高 2 个" in report_text
    assert "涉及对象：厨房门洞/推拉门高度 1 个、卫生间门洞/推拉门高度 2 个" in report_text
    assert "涉及对象：主卧全屋定制高度/类型 1 处" in report_text
    assert "涉及对象：厨房、卫生间湿区空间" in report_text


def test_generate_quote_review_report_can_write_structured_json(tmp_path: Path):
    quote_path = tmp_path / "quote.xlsx"
    report_path = tmp_path / "report.md"
    json_path = tmp_path / "report.json"
    _write_quote_workbook(quote_path)
    quantity_result = QuantityResult(
        project_name="结构化复核",
        rows=[
            _quantity_row(
                "bed",
                "卧室",
                window_details=[
                    WindowQuantityDetail(id="w1", width=1.2, height=1.8, area=2.16, height_defaulted=True),
                ],
            )
        ],
        construction_details=[],
        exceptions=[],
    )

    report_text = generate_quote_review_report(
        quote_path,
        report_path,
        quantity_result=quantity_result,
        json_output=json_path,
    )

    assert report_path.read_text(encoding="utf-8") == report_text
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert data["status_counts"] == {
        "自动生成-默认推断": 6,
        "自动生成-异常提示": 1,
        "按模板生成": 1,
    }
    assert data["source_counts"] == {
        "模板默认": 1,
        "自动汇总": 5,
        "自动算量": 2,
    }
    window_action = next(action for action in data["actions"] if action["label"] == "补窗高")
    assert window_action == {
        "label": "补窗高",
        "priority": "high",
        "suggested_action": "在 QUOTE_WINDOW 窗块属性或窗洞轮廓 XDATA 中补充 HEIGHT；也可由预算员在报价 Excel 中复核默认窗高。",
        "quote_row_count": 2,
        "item_names": ["卧室墙面项目", "墙面乳胶漆"],
        "excel_rows": [5, 6],
        "objects": ["卧室窗高 1 个"],
    }
    pipe_action = next(action for action in data["actions"] if action["label"] == "补管道/包管标识")
    assert pipe_action["priority"] == "medium"
    assert "QUOTE_PIPE_INSULATION" in pipe_action["suggested_action"]
    exterior_row = next(row for row in data["rows"] if row["excel_row"] == 8)
    assert exterior_row == {
        "excel_row": 8,
        "number": 104,
        "item_name": "外墙修补",
        "quantity": 0,
        "quantity_source": "自动汇总",
        "source_room": "全屋",
        "basis": "外墙修补范围面积汇总",
        "review_status": "自动生成-异常提示",
        "review_note": "需要确认修补范围",
    }


def test_generate_quote_review_report_can_write_action_checklist_excel(tmp_path: Path):
    quote_path = tmp_path / "quote.xlsx"
    report_path = tmp_path / "report.md"
    checklist_path = tmp_path / "quote-review-checklist.xlsx"
    _write_quote_workbook(quote_path)
    quantity_result = QuantityResult(
        project_name="清单复核",
        rows=[
            _quantity_row(
                "bed",
                "卧室",
                window_details=[
                    WindowQuantityDetail(id="w1", width=1.2, height=1.8, area=2.16, height_defaulted=True),
                ],
            )
        ],
        construction_details=[],
        exceptions=[],
    )

    generate_quote_review_report(
        quote_path,
        report_path,
        quantity_result=quantity_result,
        checklist_output=checklist_path,
    )

    workbook = load_workbook(checklist_path)
    sheet = workbook["复核清单"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["优先级", "行动类型", "建议动作", "影响报价行数", "涉及项目", "Excel行", "涉及对象", "处理状态", "备注"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:I7"
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[1] for row in rows[:3]] == ["补窗高", "补新砌墙高度/厚度", "复核外墙修补范围"]
    first_row = rows[0]
    assert first_row[:8] == (
        "high",
        "补窗高",
        "在 QUOTE_WINDOW 窗块属性或窗洞轮廓 XDATA 中补充 HEIGHT；也可由预算员在报价 Excel 中复核默认窗高。",
        2,
        "卧室墙面项目、墙面乳胶漆",
        "5、6",
        "卧室窗高 1 个",
        "待处理",
    )


def _write_quote_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品房整装报价"
    headers = [
        "编号",
        "项目名称",
        "单位",
        "数量",
        "主材单价",
        "辅材单价",
        "人工单价",
        "合价",
        "工艺说明",
        "数量来源",
        "来源空间",
        "空间ID",
        "计量口径",
        "复核状态",
        "复核备注",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column).value = header
    rows = [
        [1, "普通自动行", "m2", 10, None, None, None, None, None, "自动算量", "客厅", "living", "地面面积", "自动生成", None],
        [101, "卧室墙面项目", "m2", 31.5, None, None, None, None, None, "自动算量", "卧室", "bed", "墙面净面积", "自动生成-默认推断", "窗高缺失 1 个"],
        [102, "墙面乳胶漆", "m2", 31.5, None, None, None, None, None, "自动算量", "卧室", "bed", "墙面净面积", "自动生成-默认推断", "窗高缺失 1 个"],
        [103, "砌240厚砖墙", "m2", 6, None, None, None, None, None, "自动汇总", "全屋", None, "新砌240mm砖墙面积汇总", "自动生成-默认推断", "墙体标识缺少高度"],
        [104, "外墙修补", "m2", 0, None, None, None, None, None, "自动汇总", "全屋", None, "外墙修补范围面积汇总", "自动生成-异常提示", "需要确认修补范围"],
        [105, "厨房推拉门", "m2", 3.96, None, None, None, None, None, "自动汇总", "全屋", None, "宽度>=1.4m门洞面积汇总", "自动生成-默认推断", "推拉门门高缺少时默认推拉门高度2.4m"],
        [106, "全屋定制", "m2", 5.2, None, None, None, None, None, "自动汇总", "全屋", None, "全屋定制投影面积汇总", "自动生成-默认推断", "缺少高度的定制项按默认2.6m计算；部分定制项缺少类型，需复核"],
        [108, "包上/下水管道(单管)", "M", 3.0, None, None, None, None, None, "自动汇总", "全屋", None, "厨房/卫生间层高合计*1.5默认长度", "自动生成-默认推断", "未识别QUOTE_PIPE_WRAP包管标识，默认0"],
        [107, "主材包", "项", 1, None, None, None, None, None, "模板默认", None, None, None, "按模板生成", None],
    ]
    for row_index, values in enumerate(rows, start=4):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column).value = value
    workbook.save(path)


def _quantity_row(
    room_id: str,
    room_name: str,
    *,
    window_details: list[WindowQuantityDetail] | None = None,
    door_details: list[DoorQuantityDetail] | None = None,
    custom_details: list[FixtureQuantityDetail] | None = None,
) -> QuantityRow:
    return QuantityRow(
        room_id=room_id,
        floor=None,
        room_name=room_name,
        space_type=SpaceType.NORMAL,
        height=2.8,
        height_mode=HeightMode.PROJECT_DEFAULT,
        floor_area=10.0,
        floor_perimeter=0,
        wall_measure_perimeter=0,
        open_boundary_length=0,
        gross_wall_area=20.0,
        window_count=len(window_details or []),
        window_area=sum(window.area for window in window_details or []),
        window_details=window_details or [],
        door_opening_count=len(door_details or []),
        door_opening_area=sum(door.area for door in door_details or []),
        door_details=door_details or [],
        custom_details=custom_details or [],
        cabinet_details=[],
        net_wall_area=20.0,
        is_outdoor=False,
        include_in_floor_quantity=True,
        include_in_wall_paint_quantity=True,
        status=DataStatus.CONFIRMED,
        exception_notes=[],
    )
