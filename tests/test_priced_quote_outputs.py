from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from scripts.check_priced_quote_outputs import (
    PricedQuoteOutputError,
    check_priced_quote_outputs,
)


def test_check_priced_quote_outputs_accepts_complete_priced_package(tmp_path: Path) -> None:
    output_dir = tmp_path / "priced"
    output_dir.mkdir()
    _write_quote_workbook(output_dir / "quote-priced.xlsx")
    _write_review_json(output_dir / "quote-priced-review.json")
    (output_dir / "quote-priced-review.md").write_text("# review\n", encoding="utf-8")
    _write_placeholder_workbook(output_dir / "quote-priced-review-checklist.xlsx")
    unit_prices = tmp_path / "quote-unit-prices.xlsx"
    _write_unit_prices(unit_prices)

    summary = check_priced_quote_outputs(
        output_dir,
        unit_prices_path=unit_prices,
        expected_automation_counts={"自动算量": 2, "自动汇总": 1, "模板默认": 0},
        expected_status_counts={"自动生成-默认推断": 1, "自动生成-异常提示": 0, "按模板生成": 0},
    )

    assert summary["files"] == 4
    assert summary["matched_unit_price_rows"] == 2
    assert summary["automation_counts"] == {"自动算量": 2, "自动汇总": 1, "模板默认": 0}
    assert summary["status_counts"] == {"自动生成-默认推断": 1, "自动生成-异常提示": 0, "按模板生成": 0}


def test_check_priced_quote_outputs_reports_missing_file(tmp_path: Path) -> None:
    output_dir = tmp_path / "priced"
    output_dir.mkdir()
    _write_quote_workbook(output_dir / "quote-priced.xlsx")
    _write_review_json(output_dir / "quote-priced-review.json")

    with pytest.raises(PricedQuoteOutputError, match="Missing priced quote output"):
        check_priced_quote_outputs(output_dir)


def test_check_priced_quote_outputs_reports_unapplied_unit_price(tmp_path: Path) -> None:
    output_dir = tmp_path / "priced"
    output_dir.mkdir()
    _write_quote_workbook(output_dir / "quote-priced.xlsx", main_price=9)
    _write_review_json(output_dir / "quote-priced-review.json")
    (output_dir / "quote-priced-review.md").write_text("# review\n", encoding="utf-8")
    _write_placeholder_workbook(output_dir / "quote-priced-review-checklist.xlsx")
    unit_prices = tmp_path / "quote-unit-prices.xlsx"
    _write_unit_prices(unit_prices)

    with pytest.raises(PricedQuoteOutputError, match="Unit price mismatch"):
        check_priced_quote_outputs(output_dir, unit_prices_path=unit_prices)


def _write_quote_workbook(path: Path, *, main_price: float = 1) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品房整装报价"
    sheet.append(["编号", "项目名称", "单位", "数量", "主材单价", "辅材单价", "人工单价"])
    sheet.append([1, "地面砖铺贴(750X1500)", "M2", 10, main_price, 2, 3])
    sheet.append([2, "地面砖铺贴(750X1500)", "M2", 5, main_price, 2, 3])
    sheet.cell(row=1, column=17, value="报价自动化统计")
    sheet.cell(row=2, column=17, value="数量来源")
    sheet.cell(row=2, column=18, value="行数")
    sheet.cell(row=3, column=17, value="自动算量")
    sheet.cell(row=3, column=18, value=2)
    sheet.cell(row=4, column=17, value="自动汇总")
    sheet.cell(row=4, column=18, value=1)
    sheet.cell(row=5, column=17, value="模板默认")
    sheet.cell(row=5, column=18, value=0)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_review_json(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "status_counts": {
                    "自动生成-默认推断": 1,
                    "自动生成-异常提示": 0,
                    "按模板生成": 0,
                },
                "source_counts": {},
                "actions": [],
                "rows": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _write_unit_prices(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["项目名称", "单位", "主材单价", "辅材单价", "人工单价"])
    sheet.append(["地面砖铺贴(750X1500)", "M2", 1, 2, 3])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_placeholder_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.save(path)
