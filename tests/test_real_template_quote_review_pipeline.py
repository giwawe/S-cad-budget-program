from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from cad_budget.cad_adapter_models import CadImportResult
from cad_budget.models import ProjectInput, QuantityResult
from scripts import run_real_template_quote_review as pipeline


def test_run_quote_review_pipeline_can_generate_and_check_priced_outputs(tmp_path: Path, monkeypatch) -> None:
    dxf_path = tmp_path / "plan.dxf"
    template_path = tmp_path / "template.xlsx"
    unit_prices_path = tmp_path / "quote-unit-prices.xlsx"
    output_dir = tmp_path / "current"
    priced_output_dir = tmp_path / "priced"
    dxf_path.write_text("0\nEOF\n", encoding="utf-8")
    _write_placeholder_workbook(template_path)
    _write_unit_prices(unit_prices_path)
    project = ProjectInput(project_name="sample")
    quantity_result = QuantityResult(project_name="sample", rows=[], construction_details=[], exceptions=[])

    monkeypatch.setattr(
        pipeline,
        "import_dxf",
        lambda options: CadImportResult(project=project, source_path=options.source_path, dxf_path=options.source_path),
    )
    monkeypatch.setattr(pipeline, "calculate_quantities", lambda imported_project: quantity_result)
    monkeypatch.setattr(pipeline, "export_quantity_result", lambda result, path: _write_placeholder_workbook(path))
    monkeypatch.setattr(pipeline, "export_residential_quote", _write_quote_workbook)
    monkeypatch.setattr(pipeline, "generate_quote_review_report", _write_review_outputs)
    monkeypatch.setattr(
        pipeline,
        "build_quote_review_data",
        lambda quote_path, quantity_result=None: {
            "status_counts": {"自动生成-默认推断": 1, "自动生成-异常提示": 0, "按模板生成": 0},
            "actions": [],
        },
    )

    summary = pipeline.run_quote_review_pipeline(
        dxf_path=dxf_path,
        template_path=template_path,
        output_dir=output_dir,
        unit_prices_path=unit_prices_path,
        priced_output_dir=priced_output_dir,
        check_priced_output=True,
    )

    assert (priced_output_dir / "quote-priced.xlsx").exists()
    assert (priced_output_dir / "quote-priced-review.md").exists()
    assert (priced_output_dir / "quote-priced-review.json").exists()
    assert (priced_output_dir / "quote-priced-review-checklist.xlsx").exists()
    assert summary.priced_output_dir == priced_output_dir
    assert summary.priced_output_check is not None
    assert summary.priced_output_check["matched_unit_price_rows"] == 1
    summary_json = json.loads((output_dir / "summary.json").read_text(encoding="utf-8"))
    assert summary_json["priced_output_dir"] == str(priced_output_dir)
    assert summary_json["priced_output_check"]["matched_unit_price_rows"] == 1


def _write_quote_workbook(result, template_path: Path, output_path: Path, *, rules_path=None, unit_prices_path=None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品房整装报价"
    sheet.append(["编号", "项目名称", "单位", "数量", "主材单价", "辅材单价", "人工单价"])
    sheet.append([1, "地面砖铺贴(750X1500)", "M2", 10, 1, 2, 3])
    sheet.cell(row=1, column=17, value="报价自动化统计")
    sheet.cell(row=2, column=17, value="数量来源")
    sheet.cell(row=2, column=18, value="行数")
    sheet.cell(row=3, column=17, value="自动算量")
    sheet.cell(row=3, column=18, value=2)
    sheet.cell(row=4, column=17, value="自动汇总")
    sheet.cell(row=4, column=18, value=1)
    sheet.cell(row=5, column=17, value="模板默认")
    sheet.cell(row=5, column=18, value=0)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_review_outputs(input_excel: Path, markdown_output: Path, *, quantity_result=None, json_output=None, checklist_output=None):
    markdown_output.parent.mkdir(parents=True, exist_ok=True)
    markdown_output.write_text("# review\n", encoding="utf-8")
    if json_output is not None:
        json_output.write_text(
            json.dumps(
                {
                    "status_counts": {"自动生成-默认推断": 1, "自动生成-异常提示": 0, "按模板生成": 0},
                    "source_counts": {},
                    "actions": [],
                    "rows": [],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
    if checklist_output is not None:
        _write_placeholder_workbook(checklist_output)
    return "# review\n"


def _write_unit_prices(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["项目名称", "单位", "主材单价", "辅材单价", "人工单价"])
    sheet.append(["地面砖铺贴(750X1500)", "M2", 1, 2, 3])
    workbook.save(path)


def _write_placeholder_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.save(path)
