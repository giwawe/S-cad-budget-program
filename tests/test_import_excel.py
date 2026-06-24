from pathlib import Path

from openpyxl import load_workbook

from cad_budget.export_excel import export_quantity_result
from cad_budget.import_excel import import_quantity_result
from cad_budget.models import (
    ConstructionKind,
    ConstructionQuantityDetail,
    DataStatus,
    DoorQuantityDetail,
    FixtureKind,
    FixturePricingMode,
    FixtureQuantityDetail,
    HeightMode,
    QuantityRow,
    QuantityResult,
    SpaceType,
    WindowQuantityDetail,
)


def test_import_quantity_result_reads_edited_room_rows(tmp_path: Path):
    source = QuantityResult(
        project_name="Editable Demo",
        rows=[
            QuantityRow(
                room_id="room-1",
                floor="1F",
                room_name="Study",
                space_type=SpaceType.NORMAL,
                height=2.8,
                height_mode=HeightMode.PROJECT_DEFAULT,
                floor_area=9.0,
                floor_perimeter=12.0,
                wall_measure_perimeter=11.5,
                open_boundary_length=0.5,
                gross_wall_area=32.2,
                window_count=1,
                window_area=1.8,
                door_opening_count=1,
                door_opening_area=0.0,
                net_wall_area=30.4,
                is_outdoor=False,
                include_in_floor_quantity=True,
                include_in_wall_paint_quantity=True,
                status=DataStatus.DEFAULT_INFERRED,
                exception_notes=["window_height_defaulted"],
            )
        ],
        exceptions=[],
    )
    workbook_path = tmp_path / "editable.xlsx"
    export_quantity_result(source, workbook_path)

    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    sheet["B4"] = "Edited Study"
    sheet["D4"] = 3.0
    sheet["G4"] = 10.0
    sheet["K4"] = 2.5
    sheet["R4"] = DataStatus.MANUALLY_EDITED.value
    sheet["S4"] = "manual_review"
    workbook.save(workbook_path)

    result = import_quantity_result(workbook_path)

    row = result.rows[0]
    assert result.project_name == "Editable Demo"
    assert row.room_id == "room-1"
    assert row.room_name == "Edited Study"
    assert row.height == 3.0
    assert row.wall_measure_perimeter == 10.0
    assert row.gross_wall_area == 30.0
    assert row.window_area == 2.5
    assert row.net_wall_area == 27.5
    assert row.status is DataStatus.MANUALLY_EDITED
    assert row.exception_notes == ["manual_review"]


def test_import_quantity_result_preserves_hidden_quote_details(tmp_path: Path):
    source = QuantityResult(
        project_name="Detail Round Trip",
        rows=[
            QuantityRow(
                room_id="room-1",
                floor=None,
                room_name="Bedroom",
                space_type=SpaceType.NORMAL,
                height=2.8,
                height_mode=HeightMode.PROJECT_DEFAULT,
                floor_area=9.0,
                floor_perimeter=12.0,
                wall_measure_perimeter=11.5,
                open_boundary_length=0.0,
                gross_wall_area=32.2,
                window_count=1,
                window_area=1.8,
                window_details=[
                    WindowQuantityDetail(
                        id="w1",
                        room_id="room-1",
                        width=1.2,
                        height=1.5,
                        area=1.8,
                        height_defaulted=False,
                        wall_segment_key="north",
                        wall_segment_length=3.0,
                    )
                ],
                door_opening_count=1,
                door_opening_area=0.0,
                door_details=[
                    DoorQuantityDetail(
                        id="d1",
                        room_id="room-1",
                        width=0.9,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=1.89,
                    )
                ],
                net_wall_area=30.4,
                custom_details=[
                    FixtureQuantityDetail(
                        id="custom-1",
                        room_id="room-1",
                        room_name="Bedroom",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        effective_height=2.6,
                        height_defaulted=True,
                        projected_area=5.2,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                    )
                ],
                cabinet_details=[
                    FixtureQuantityDetail(
                        id="cabinet-1",
                        room_id="room-1",
                        room_name="Bedroom",
                        kind=FixtureKind.CABINET,
                        length=3.0,
                        pricing_mode=FixturePricingMode.LENGTH,
                    )
                ],
                is_outdoor=False,
                include_in_floor_quantity=True,
                include_in_wall_paint_quantity=True,
                status=DataStatus.CONFIRMED,
            )
        ],
        exceptions=[],
    )
    workbook_path = tmp_path / "details.xlsx"
    export_quantity_result(source, workbook_path)

    result = import_quantity_result(workbook_path)

    row = result.rows[0]
    assert row.window_details[0].wall_segment_length == 3.0
    assert row.door_details[0].effective_height == 2.1
    assert row.door_details[0].height_defaulted is True
    assert row.custom_details[0].projected_area == 5.2
    assert row.custom_details[0].pricing_mode is FixturePricingMode.PROJECTED_AREA
    assert row.cabinet_details[0].length == 3.0


def test_import_quantity_result_reads_exterior_sheet(tmp_path: Path):
    source = QuantityResult(
        project_name="Exterior Editable",
        rows=[],
        exterior_rows=[],
        exceptions=[],
    )
    workbook_path = tmp_path / "editable.xlsx"
    export_quantity_result(source, workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook.create_sheet("\u5916\u5899\u8868")
    sheet["A1"] = "\u9879\u76ee\u540d\u79f0"
    sheet["B1"] = "Exterior Editable"
    sheet.append([])
    sheet.append([
        "\u697c\u5c42",
        "\u5916\u5899\u7f16\u53f7",
        "\u9ad8\u5ea6",
        "\u5916\u5899\u8ba1\u91cf\u957f\u5ea6",
        "\u5916\u5899\u6d1e\u53e3\u957f\u5ea6",
        "\u5916\u5899\u6bdb\u9762\u79ef",
        "\u5916\u5899\u51c0\u9762\u79ef",
    ])
    sheet.append(["1F", "ext-1", 3.0, 6.0, 1.2, "=D4*C4", "=F4-E4*C4"])
    workbook.save(workbook_path)

    result = import_quantity_result(workbook_path)

    assert len(result.exterior_rows) == 1
    exterior = result.exterior_rows[0]
    assert exterior.exterior_wall_id == "ext-1"
    assert exterior.gross_area == 18.0
    assert exterior.net_area == 14.4


def test_import_quantity_result_preserves_construction_sheet(tmp_path: Path):
    source = QuantityResult(
        project_name="Construction Round Trip",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="new-120",
                kind=ConstructionKind.NEW_WALL,
                floor="1F",
                length=2.0,
                height=3.0,
                effective_height=3.0,
                thickness=0.12,
                area=6.0,
                count=1,
            ),
            ConstructionQuantityDetail(
                id="hole-1",
                kind=ConstructionKind.LINTEL_HOLE,
                count=1,
            ),
        ],
        exceptions=[],
    )
    workbook_path = tmp_path / "construction.xlsx"
    export_quantity_result(source, workbook_path)

    result = import_quantity_result(workbook_path)

    assert len(result.construction_details) == 2
    new_wall = result.construction_details[0]
    assert new_wall.id == "new-120"
    assert new_wall.kind is ConstructionKind.NEW_WALL
    assert new_wall.length == 2.0
    assert new_wall.thickness == 0.12
    assert new_wall.area == 6.0
    hole = result.construction_details[1]
    assert hole.kind is ConstructionKind.LINTEL_HOLE
    assert hole.count == 1
