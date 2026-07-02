import json
from pathlib import Path

import ezdxf
from openpyxl import Workbook, load_workbook
from typer.testing import CliRunner

from cad_budget.cli import app
from cad_budget.quote_excel import load_quote_rules


def test_cli_calculates_json_output(tmp_path: Path):
    runner = CliRunner()
    fixture = Path(__file__).parent / "fixtures" / "simple_apartment.json"
    output = tmp_path / "result.json"

    result = runner.invoke(
        app,
        [
            "calculate",
            str(fixture),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 0
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["project_name"] == "Simple Apartment"
    assert data["rows"][0]["room_name"] == "卧室"
    assert data["rows"][0]["floor_area"] == 12


def test_cli_creates_nested_output_directory(tmp_path: Path):
    runner = CliRunner()
    fixture = Path(__file__).parent / "fixtures" / "simple_apartment.json"
    output = tmp_path / "nested" / "dir" / "result.json"

    result = runner.invoke(
        app,
        [
            "calculate",
            str(fixture),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 0
    assert output.exists()
    assert output.parent.is_dir()
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["rows"][0]["room_name"] == "卧室"


def test_cli_writes_nested_excel_output(tmp_path: Path):
    runner = CliRunner()
    fixture = Path(__file__).parent / "fixtures" / "simple_apartment.json"
    json_output = tmp_path / "nested" / "json" / "result.json"
    excel_output = tmp_path / "nested" / "excel" / "result.xlsx"

    result = runner.invoke(
        app,
        [
            "calculate",
            str(fixture),
            "--json-output",
            str(json_output),
            "--excel-output",
            str(excel_output),
        ],
    )

    assert result.exit_code == 0
    assert json_output.exists()
    assert excel_output.exists()
    workbook = load_workbook(excel_output)
    assert workbook.sheetnames == ["算量表"]
    assert "Wrote" in result.output


def test_cli_reports_excel_output_failure_when_parent_is_file(tmp_path: Path):
    runner = CliRunner()
    fixture = Path(__file__).parent / "fixtures" / "simple_apartment.json"
    output = tmp_path / "result.json"
    blocked_dir = tmp_path / "not_a_dir"
    blocked_dir.write_text("blocked", encoding="utf-8")
    excel_output = blocked_dir / "result.xlsx"

    result = runner.invoke(
        app,
        [
            "calculate",
            str(fixture),
            "--json-output",
            str(output),
            "--excel-output",
            str(excel_output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to write Excel output" in error_text
    assert "Traceback" not in error_text
    assert "Wrote" not in error_text


def test_cli_reports_missing_input_file(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "output" / "result.json"

    result = runner.invoke(
        app,
        [
            "calculate",
            str(tmp_path / "missing.json"),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to read input JSON" in error_text
    assert "missing.json" in error_text
    assert "Traceback" not in error_text


def test_cli_reports_invalid_json_input(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "output" / "result.json"
    bad_input = tmp_path / "bad.json"
    bad_input.write_text("{invalid}", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "calculate",
            str(bad_input),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Invalid project JSON" in error_text
    assert "Traceback" not in error_text


def test_cli_reports_non_utf8_input(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "output" / "result.json"
    bad_input = tmp_path / "non_utf8.bin"
    bad_input.write_bytes(bytes([0xFF, 0xFE, 0xFD]))

    result = runner.invoke(
        app,
        [
            "calculate",
            str(bad_input),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to read input JSON" in error_text
    assert "non_utf8.bin" in error_text
    assert "Traceback" not in error_text


def test_cli_reports_invalid_room_geometry_without_traceback(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "result.json"
    bad_input = tmp_path / "bad_room.json"
    bad_input.write_text(
        json.dumps(
            {
                "project_name": "Bad Geometry",
                "default_height": 2.8,
                "rooms": [
                    {
                        "id": "bow-tie",
                        "points": [
                            {"x": 0, "y": 0},
                            {"x": 2, "y": 2},
                            {"x": 0, "y": 2},
                            {"x": 2, "y": 0},
                            {"x": 0, "y": 0},
                        ],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "calculate",
            str(bad_input),
            "--json-output",
            str(output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to calculate quantities" in error_text
    assert "Traceback" not in error_text


def test_cli_import_cad_dxf_writes_project_json(tmp_path: Path):
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4
    modelspace = doc.modelspace()
    doc.layers.add("QUOTE_ROOM")
    modelspace.add_lwpolyline(
        [(0, 0), (4000, 0), (4000, 3000), (0, 3000), (0, 0)],
        dxfattribs={"layer": "QUOTE_ROOM", "closed": True},
    )
    dxf_path = tmp_path / "plan.dxf"
    doc.saveas(dxf_path)
    output = tmp_path / "project.json"
    runner = CliRunner()

    result = runner.invoke(
        app,
        [
            "import-cad",
            str(dxf_path),
            "--json-output",
            str(output),
            "--unit",
            "mm",
        ],
    )

    assert result.exit_code == 0
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["project_name"] == "plan"
    assert len(data["rooms"]) == 1


def test_cli_import_cad_dwg_requires_converter(tmp_path: Path):
    dwg_path = tmp_path / "plan.dwg"
    dwg_path.write_bytes(b"fake")
    output = tmp_path / "project.json"
    runner = CliRunner()

    result = runner.invoke(app, ["import-cad", str(dwg_path), "--json-output", str(output)])

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "DWG input requires a configured DWG-to-DXF converter" in error_text
    assert not output.exists()


def test_cli_import_cad_dwg_reports_conversion_output_path_error(tmp_path: Path):
    dwg_path = tmp_path / "plan.dwg"
    dwg_path.write_bytes(b"fake")
    blocked_parent = tmp_path / "not_a_dir"
    blocked_parent.write_text("blocked", encoding="utf-8")
    output = blocked_parent / "project.json"
    runner = CliRunner()

    result = runner.invoke(
        app,
        [
            "import-cad",
            str(dwg_path),
            "--json-output",
            str(output),
            "--dwg-converter",
            "dummy-converter",
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to prepare DWG conversion output" in error_text
    assert "Traceback" not in error_text
    assert not output.exists()


def test_cli_import_cad_unsupported_extension_exits_1(tmp_path: Path):
    cad_path = tmp_path / "plan.txt"
    cad_path.write_text("fake", encoding="utf-8")
    output = tmp_path / "project.json"
    runner = CliRunner()

    result = runner.invoke(app, ["import-cad", str(cad_path), "--json-output", str(output)])

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Unsupported CAD file extension" in error_text
    assert ".txt" in error_text
    assert not output.exists()


def test_cli_import_cad_dxf_blocker_exits_1_without_writing(tmp_path: Path):
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4
    dxf_path = tmp_path / "empty.dxf"
    doc.saveas(dxf_path)
    output = tmp_path / "project.json"
    runner = CliRunner()

    result = runner.invoke(
        app,
        [
            "import-cad",
            str(dxf_path),
            "--json-output",
            str(output),
            "--unit",
            "mm",
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "blocker: QUOTE_ROOM_MISSING" in error_text
    assert not output.exists()


def test_cli_calculate_writes_exterior_rows_to_json(tmp_path: Path):
    input_json = tmp_path / "exterior_project.json"
    output = tmp_path / "result.json"
    input_json.write_text(
        json.dumps(
            {
                "project_name": "Exterior CLI",
                "default_height": 2.8,
                "floor_heights": {"1F": 3.0},
                "exterior_walls": [
                    {
                        "id": "ext-wall",
                        "layer": "QUOTE_EXT_WALL",
                        "floor": "1F",
                        "points": [{"x": 0, "y": 0}, {"x": 4, "y": 0}],
                    }
                ],
                "exterior_openings": [
                    {
                        "id": "ext-open",
                        "layer": "QUOTE_EXT_OPENING",
                        "floor": "1F",
                        "points": [{"x": 1, "y": 0}, {"x": 2, "y": 0}],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    runner = CliRunner()

    result = runner.invoke(app, ["calculate", str(input_json), "--json-output", str(output)])

    assert result.exit_code == 0
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["rows"] == []
    assert data["exterior_rows"][0]["gross_area"] == 12.0
    assert data["exterior_rows"][0]["net_area"] == 9.0


def test_cli_import_excel_writes_edited_quantity_json(tmp_path: Path):
    runner = CliRunner()
    fixture = Path(__file__).parent / "fixtures" / "simple_apartment.json"
    calculated_json = tmp_path / "result.json"
    excel_output = tmp_path / "result.xlsx"

    calculate_result = runner.invoke(
        app,
        [
            "calculate",
            str(fixture),
            "--json-output",
            str(calculated_json),
            "--excel-output",
            str(excel_output),
        ],
    )
    assert calculate_result.exit_code == 0

    workbook = load_workbook(excel_output)
    sheet = workbook.active
    sheet["D4"] = 3.0
    sheet["G4"] = 10.0
    sheet["K4"] = 2.0
    workbook.save(excel_output)

    imported_json = tmp_path / "edited-result.json"
    import_result = runner.invoke(
        app,
        [
            "import-excel",
            str(excel_output),
            "--json-output",
            str(imported_json),
        ],
    )

    assert import_result.exit_code == 0
    data = json.loads(imported_json.read_text(encoding="utf-8"))
    assert data["rows"][0]["gross_wall_area"] == 30.0
    assert data["rows"][0]["net_wall_area"] == 28.0


def test_cli_quote_writes_residential_fitout_excel(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)

    result = runner.invoke(app, ["quote", str(input_json), "--template", str(template), "--excel-output", str(output)])

    assert result.exit_code == 0
    assert output.exists()
    workbook = load_workbook(output, data_only=False)
    sheet = workbook.active
    assert workbook.sheetnames == ["商品房整装报价"]
    assert sheet["B5"].value == "客厅工程"
    assert sheet["J3"].value == "数量来源"
    assert sheet["N6"].value == "自动生成"
    assert sheet["O3"].value == "复核备注"
    assert sheet["Q1"].value == "报价自动化统计"
    assert "Wrote" in result.output


def test_cli_export_prices_writes_deduplicated_unit_price_table(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    quote_output = tmp_path / "quote.xlsx"
    prices_output = tmp_path / "unit-prices.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)
    quote_result = runner.invoke(app, ["quote", str(input_json), "--template", str(template), "--excel-output", str(quote_output)])
    assert quote_result.exit_code == 0

    result = runner.invoke(app, ["export-prices", str(quote_output), "--excel-output", str(prices_output)])

    assert result.exit_code == 0
    rows = list(load_workbook(prices_output).active.iter_rows(values_only=True))
    floor_tile_rows = [row for row in rows[1:] if row[0] == "地面砖铺贴(750X1500)" and row[1] == "M2"]
    assert len(floor_tile_rows) == 1
    assert "Wrote" in result.output


def test_cli_quote_applies_global_unit_price_table(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    prices = tmp_path / "unit-prices.xlsx"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload(include_kitchen=True)), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)
    _write_unit_price_cli_table(prices, [("地面砖铺贴(750X1500)", "M2", 1, 2, 3)])

    result = runner.invoke(
        app,
        ["quote", str(input_json), "--template", str(template), "--unit-prices", str(prices), "--excel-output", str(output)],
    )

    assert result.exit_code == 0
    rows = list(load_workbook(output, data_only=False).active.iter_rows(values_only=True))
    living_floor_tile = _row_containing_after(rows, "客厅工程", "地面砖铺贴(750X1500)")
    kitchen_floor_tile = _row_containing_after(rows, "厨房工程", "地面砖铺贴(750X1500)")
    assert living_floor_tile[4:7] == (1, 2, 3)
    assert kitchen_floor_tile[4:7] == (1, 2, 3)


def test_cli_quote_uses_default_unit_price_workbook(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "quote.xlsx"
    default_prices = tmp_path / "config" / "quote-unit-prices.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload(include_kitchen=True)), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)
    default_prices.parent.mkdir()
    _write_unit_price_cli_table(default_prices, [("地面砖铺贴(750X1500)", "M2", 1, 2, 3)])

    result = runner.invoke(app, ["quote", str(input_json), "--template", str(template), "--excel-output", str(output)])

    assert result.exit_code == 0
    rows = list(load_workbook(output, data_only=False).active.iter_rows(values_only=True))
    living_floor_tile = _row_containing_after(rows, "客厅工程", "地面砖铺贴(750X1500)")
    kitchen_floor_tile = _row_containing_after(rows, "厨房工程", "地面砖铺贴(750X1500)")
    assert living_floor_tile[4:7] == (1, 2, 3)
    assert kitchen_floor_tile[4:7] == (1, 2, 3)


def test_cli_init_prices_writes_default_unit_price_table_and_refuses_overwrite(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    runner = CliRunner()
    template = tmp_path / "template.xlsx"
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)
    default_prices = tmp_path / "config" / "quote-unit-prices.xlsx"

    result = runner.invoke(app, ["init-prices", str(template)])
    second_result = runner.invoke(app, ["init-prices", str(template)])

    assert result.exit_code == 0
    assert default_prices.exists()
    rows = list(load_workbook(default_prices).active.iter_rows(values_only=True))
    assert rows[0][:5] == ("项目名称", "单位", "主材单价", "辅材单价", "人工单价")
    floor_tile_rows = [row for row in rows[1:] if row[0] == "地面砖铺贴(750X1500)" and row[1] == "M2"]
    assert len(floor_tile_rows) == 1
    assert second_result.exit_code == 1
    assert "already exists" in second_result.output


def test_cli_init_prices_writes_custom_output_without_default_path(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    runner = CliRunner()
    template = tmp_path / "template.xlsx"
    custom_prices = tmp_path / "prices" / "custom-unit-prices.xlsx"
    default_prices = tmp_path / "config" / "quote-unit-prices.xlsx"
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)

    result = runner.invoke(app, ["init-prices", str(template), "--output", str(custom_prices)])

    assert result.exit_code == 0
    assert custom_prices.exists()
    assert not default_prices.exists()
    rows = list(load_workbook(custom_prices).active.iter_rows(values_only=True))
    floor_tile_rows = [row for row in rows[1:] if row[0] == "地面砖铺贴(750X1500)" and row[1] == "M2"]
    assert len(floor_tile_rows) == 1


def test_cli_priced_quote_generates_quote_and_review_outputs(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output_dir = tmp_path / "priced"
    default_prices = tmp_path / "config" / "quote-unit-prices.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload(include_kitchen=True)), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)
    default_prices.parent.mkdir()
    _write_unit_price_cli_table(default_prices, [("地面砖铺贴(750X1500)", "M2", 1, 2, 3)])

    result = runner.invoke(app, ["priced-quote", str(input_json), "--template", str(template), "--output-dir", str(output_dir)])

    assert result.exit_code == 0
    quote_output = output_dir / "quote-priced.xlsx"
    assert quote_output.exists()
    assert (output_dir / "quote-priced-review.md").exists()
    assert (output_dir / "quote-priced-review.json").exists()
    assert (output_dir / "quote-priced-review-checklist.xlsx").exists()
    rows = list(load_workbook(quote_output, data_only=False).active.iter_rows(values_only=True))
    living_floor_tile = _row_containing_after(rows, "客厅工程", "地面砖铺贴(750X1500)")
    assert living_floor_tile[4:7] == (1, 2, 3)
    assert "Wrote" in result.output


def test_cli_priced_quote_requires_unit_price_workbook(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output_dir = tmp_path / "priced"
    input_json.write_text(json.dumps(_quantity_result_payload(include_kitchen=True)), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)

    result = runner.invoke(app, ["priced-quote", str(input_json), "--template", str(template), "--output-dir", str(output_dir)])

    assert result.exit_code == 1
    assert "Unit price workbook is required for priced-quote" in result.output
    assert not output_dir.exists()


def test_cli_check_prices_reports_invalid_unit_price_table(tmp_path: Path):
    runner = CliRunner()
    prices = tmp_path / "unit-prices.xlsx"
    _write_unit_price_cli_table(
        prices,
        [
            ("墙面乳胶漆", "M2", 10, None, 5),
            ("顶面批嵌", "M2", 1, 2, 3),
            ("顶面批嵌", "M2", 1, 2, 4),
        ],
    )

    result = runner.invoke(app, ["check-prices", str(prices)])

    assert result.exit_code == 1
    assert "MISSING_PRICE" in result.output
    assert "DUPLICATE_CONFLICT" in result.output
    assert "墙面乳胶漆" in result.output


def test_cli_check_prices_accepts_valid_unit_price_table(tmp_path: Path):
    runner = CliRunner()
    prices = tmp_path / "unit-prices.xlsx"
    _write_unit_price_cli_table(prices, [("墙面乳胶漆", "M2", 10, 0, 5)])

    result = runner.invoke(app, ["check-prices", str(prices)])

    assert result.exit_code == 0
    assert "Unit price table OK" in result.output


def test_cli_quote_report_writes_markdown_review_report(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    quote_output = tmp_path / "quote.xlsx"
    report_output = tmp_path / "quote-review.md"
    json_output = tmp_path / "quote-review.json"
    checklist_output = tmp_path / "quote-review-checklist.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)
    quote_result = runner.invoke(
        app,
        ["quote", str(input_json), "--template", str(template), "--excel-output", str(quote_output)],
    )
    assert quote_result.exit_code == 0

    result = runner.invoke(
        app,
        [
            "quote-report",
            str(quote_output),
            "--quantity-json",
            str(input_json),
            "--markdown-output",
            str(report_output),
            "--json-output",
            str(json_output),
            "--checklist-output",
            str(checklist_output),
        ],
    )

    assert result.exit_code == 0
    assert report_output.exists()
    assert json_output.exists()
    assert checklist_output.exists()
    report_text = report_output.read_text(encoding="utf-8")
    report_data = json.loads(json_output.read_text(encoding="utf-8"))
    checklist_workbook = load_workbook(checklist_output)
    assert "# 报价复核报告" in report_text
    assert "按模板生成" in report_text
    assert "actions" in report_data
    assert "rows" in report_data
    assert checklist_workbook.sheetnames == ["复核清单"]
    assert "Wrote" in result.output


def test_cli_quote_report_fail_on_high_blocks_high_priority_actions(tmp_path: Path):
    runner = CliRunner()
    quote_output = tmp_path / "quote.xlsx"
    report_output = tmp_path / "quote-review.md"
    _write_quote_review_workbook(
        quote_output,
        [["窗高缺失墙面项目", "m2", 10, "自动算量", "墙面净面积", "自动生成-默认推断", "窗高缺失 1 个"]],
    )

    result = runner.invoke(
        app,
        [
            "quote-report",
            str(quote_output),
            "--markdown-output",
            str(report_output),
            "--fail-on",
            "high",
        ],
    )

    assert result.exit_code == 1
    assert report_output.exists()
    assert "Quote review gate failed" in result.output
    assert "补窗高" in result.output


def test_cli_quote_report_fail_on_high_allows_medium_priority_actions(tmp_path: Path):
    runner = CliRunner()
    quote_output = tmp_path / "quote.xlsx"
    report_output = tmp_path / "quote-review.md"
    _write_quote_review_workbook(
        quote_output,
        [["包上/下水管道(单管)", "M", 3, "自动汇总", "厨房/卫生间层高合计*1.5默认长度", "自动生成-默认推断", "未识别QUOTE_PIPE_WRAP包管标识"]],
    )

    high_result = runner.invoke(
        app,
        [
            "quote-report",
            str(quote_output),
            "--markdown-output",
            str(report_output),
            "--fail-on",
            "high",
        ],
    )
    medium_result = runner.invoke(
        app,
        [
            "quote-report",
            str(quote_output),
            "--markdown-output",
            str(report_output),
            "--fail-on",
            "medium",
        ],
    )

    assert high_result.exit_code == 0
    assert medium_result.exit_code == 1
    assert "补管道/包管标识" in medium_result.output


def test_cli_init_rules_writes_default_rules_json(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "rules.json"

    result = runner.invoke(app, ["init-rules", "--output", str(output)])

    assert result.exit_code == 0
    assert output.exists()
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["wet_room_heights"]["wall_tile_height"] == 2.5
    assert data["building_area_percent_count_items"]
    assert "floor_area_percent_count_items" not in data
    assert "垃圾清运费" in data["building_area_aggregate_items"]
    rules = load_quote_rules(output)
    assert rules.wall_tile_height == 2.5
    assert "Wrote" in result.output


def test_cli_init_rules_creates_nested_output_directory(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "nested" / "rules" / "rules.json"

    result = runner.invoke(app, ["init-rules", "--output", str(output)])

    assert result.exit_code == 0
    assert output.exists()
    assert output.parent.is_dir()


def test_cli_init_rules_refuses_to_overwrite_existing_file(tmp_path: Path):
    runner = CliRunner()
    output = tmp_path / "rules.json"
    output.write_text("keep", encoding="utf-8")

    result = runner.invoke(app, ["init-rules", "--output", str(output)])

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "already exists" in error_text
    assert "Traceback" not in error_text
    assert output.read_text(encoding="utf-8") == "keep"


def test_cli_quote_uses_external_rules_file(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    rules = tmp_path / "rules.json"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload(include_kitchen=True)), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True, include_kitchen=True)
    _write_quote_rules(rules, kitchen_height=0.5, bathroom_height=1.2, tile_height=2.0)

    result = runner.invoke(
        app,
        ["quote", str(input_json), "--template", str(template), "--rules", str(rules), "--excel-output", str(output)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output, data_only=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    kitchen_waterproof = _row_containing_after(rows, "厨房工程", "墙地面防漏处理")
    kitchen_wall_tile = _row_containing_after(rows, "厨房工程", "墙面贴瓷砖(600X1200)")
    assert kitchen_waterproof[3] == 11.0
    assert kitchen_wall_tile[3] == 20.0
    assert sheet["Q7"].value == "规则来源"
    assert sheet["R7"].value == str(rules)


def test_cli_quote_reports_invalid_external_rules_without_traceback(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    rules = tmp_path / "rules.json"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)
    rules.write_text("{invalid}", encoding="utf-8")

    result = runner.invoke(
        app,
        ["quote", str(input_json), "--template", str(template), "--rules", str(rules), "--excel-output", str(output)],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to generate quote Excel" in error_text
    assert "Invalid quote rules" in error_text
    assert "Traceback" not in error_text


def test_cli_quote_reports_missing_external_rules_without_traceback(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)
    missing_rules = tmp_path / "missing-rules.json"

    result = runner.invoke(
        app,
        [
            "quote",
            str(input_json),
            "--template",
            str(template),
            "--rules",
            str(missing_rules),
            "--excel-output",
            str(output),
        ],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to generate quote Excel" in error_text
    assert "missing-rules.json" in error_text
    assert "Traceback" not in error_text


def test_cli_quote_reports_missing_template_without_traceback(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    output = tmp_path / "quote.xlsx"

    result = runner.invoke(
        app,
        ["quote", str(input_json), "--template", str(tmp_path / "missing.xlsx"), "--excel-output", str(output)],
    )

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to generate quote Excel" in error_text
    assert "Traceback" not in error_text
    assert not output.exists()


def test_cli_quote_reports_missing_fitout_sheet_without_traceback(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "result.json"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps(_quantity_result_payload()), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=False)

    result = runner.invoke(app, ["quote", str(input_json), "--template", str(template), "--excel-output", str(output)])

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Failed to generate quote Excel" in error_text
    assert "整装" in error_text
    assert "Traceback" not in error_text


def test_cli_quote_reports_invalid_quantity_json_without_traceback(tmp_path: Path):
    runner = CliRunner()
    input_json = tmp_path / "bad.json"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "quote.xlsx"
    input_json.write_text(json.dumps({"project_name": "Not A Result"}), encoding="utf-8")
    _write_quote_cli_template(template, include_fitout=True)

    result = runner.invoke(app, ["quote", str(input_json), "--template", str(template), "--excel-output", str(output)])

    assert result.exit_code == 1
    error_text = (result.stdout or "") + (result.stderr or "")
    assert "Invalid quantity result JSON" in error_text
    assert "Traceback" not in error_text


def _quantity_result_payload(*, include_kitchen: bool = False):
    rows = [
        {
            "room_id": "living",
            "floor": None,
            "room_name": "客厅",
            "space_type": "normal",
            "height": 2.8,
            "height_mode": "project_default",
            "floor_area": 20.0,
            "floor_perimeter": 0,
            "wall_measure_perimeter": 0,
            "open_boundary_length": 0,
            "gross_wall_area": 50.0,
            "window_count": 0,
            "window_area": 0,
            "door_opening_count": 0,
            "door_opening_area": 0,
            "net_wall_area": 50.0,
            "is_outdoor": False,
            "include_in_floor_quantity": True,
            "include_in_wall_paint_quantity": True,
            "status": "confirmed",
            "exception_notes": [],
        }
    ]
    if include_kitchen:
        rows.append(
            {
                "room_id": "kitchen",
                "floor": None,
                "room_name": "厨房",
                "space_type": "normal",
                "height": 2.8,
                "height_mode": "project_default",
                "floor_area": 6.0,
                "floor_perimeter": 0,
                "wall_measure_perimeter": 10.0,
                "open_boundary_length": 0,
                "gross_wall_area": 18.0,
                "window_count": 1,
                "window_area": 1.0,
                "door_opening_count": 0,
                "door_opening_area": 0,
                "net_wall_area": 18.0,
                "is_outdoor": False,
                "include_in_floor_quantity": True,
                "include_in_wall_paint_quantity": True,
                "status": "confirmed",
                "exception_notes": [],
            }
        )
    return {
        "project_name": "Quote CLI",
        "rows": rows,
        "exterior_rows": [],
        "exceptions": [],
    }


def _write_quote_cli_template(path: Path, *, include_fitout: bool, include_kitchen: bool = False) -> None:
    workbook = load_workbook(path) if path.exists() else None
    if workbook is None:
        from openpyxl import Workbook

        workbook = Workbook()
    half = workbook.active
    half.title = "半包"
    half.append(["工程(预) 算表"])
    if include_fitout:
        sheet = workbook.create_sheet("整装")
        sheet.append(["工程(预) 算表"])
        sheet.append(["名称：Demo"])
        sheet.append(["编号", "项目名称", "单位", "数量", "材料费(元)", None, "人工费\n(元)", "总价(元)", "材料及工艺说明"])
        sheet.append([None, None, None, None, "主材\n单价", "辅材\n单价"])
        sheet.append(["一", "客厅工程"])
        sheet.append([1, "顶面批嵌", "M2", 1, 0, 15, 10, None, "说明"])
        sheet.append([2, "墙面乳胶漆", "M2", 1, 10, 0, 10, None, "说明"])
        sheet.append([3, "地面砖铺贴(750X1500)", "M2", 1, 0, 36, 60, None, "说明"])
        if include_kitchen:
            sheet.append(["二", "厨房工程"])
            sheet.append([1, "地面找平", "M2", 1, 0, 26, 30, None, "说明"])
            sheet.append([2, "墙地面防漏处理", "M2", 1, 28, 10.5, 13, None, "说明"])
            sheet.append([3, "墙面贴瓷砖(600X1200)", "M2", 1, 0, 40, 60, None, "说明"])
            sheet.append([4, "地面砖铺贴(750X1500)", "M2", 1, 0, 36, 60, None, "说明"])
    workbook.save(path)


def _write_quote_review_workbook(path: Path, rows: list[list]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品房整装报价"
    headers = [
        "编号",
        "项目名称",
        "单位",
        "数量",
        "主材单价",
        "辅材单价",
        "人工单价",
        "合价",
        "工艺说明",
        "数量来源",
        "来源空间",
        "空间ID",
        "计量口径",
        "复核状态",
        "复核备注",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column).value = header
    for index, row in enumerate(rows, start=4):
        item_name, unit, quantity, source, basis, status, note = row
        values = [index - 3, item_name, unit, quantity, None, None, None, None, None, source, "全屋", None, basis, status, note]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=index, column=column).value = value
    workbook.save(path)


def _write_unit_price_cli_table(path: Path, rows: list[tuple[str, str, float, float, float]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "全局单价表"
    sheet.append(["项目名称", "单位", "主材单价", "辅材单价", "人工单价", "备注"])
    for row in rows:
        sheet.append([*row, None])
    workbook.save(path)


def _write_quote_rules(path: Path, *, kitchen_height: float, bathroom_height: float, tile_height: float) -> None:
    path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": kitchen_height,
                    "bathroom_waterproof_wall_height": bathroom_height,
                    "wall_tile_height": tile_height,
                },
                "floor_area_aggregate_items": ["垃圾清运费"],
                "tile_area_aggregate_items": ["美缝"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _row_containing_after(rows, section_name: str, item_name: str):
    section_seen = False
    for row in rows:
        if any(section_name in cell for cell in row if isinstance(cell, str)):
            section_seen = True
            continue
        if section_seen and any(item_name in cell for cell in row if isinstance(cell, str)):
            return row
    return None
