from pathlib import Path
import os
import subprocess
import sys

import ezdxf
from openpyxl import Workbook, load_workbook

from cad_budget.cad_adapter_models import CadImportOptions, CadUnit
from cad_budget.dxf_adapter import import_dxf
from cad_budget.models import ConstructionKind
from cad_budget.quantity import calculate_quantities
from cad_budget.quote_excel import export_residential_quote


def test_marker_rich_dxf_turns_quote_defaults_into_automatic_aggregates(tmp_path: Path):
    dxf_path = tmp_path / "marker_rich_plan.dxf"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _build_marker_rich_dxf(dxf_path)
    _create_marker_quote_template(template_path)

    import_result = import_dxf(CadImportOptions(source_path=dxf_path, confirmed_unit=CadUnit.MILLIMETER))
    assert not import_result.has_blockers
    assert import_result.project is not None
    quantity = calculate_quantities(import_result.project)

    assert quantity.building_area == 36.0
    assert len(quantity.exterior_rows) == 1
    assert len(quantity.construction_details) == 7
    wall_tile_details = [detail for detail in quantity.construction_details if detail.kind is ConstructionKind.WALL_TILE]
    assert len(wall_tile_details) == 1
    assert wall_tile_details[0].room_name == "阳台"
    assert wall_tile_details[0].area == 2.4

    export_residential_quote(quantity, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    expected_quantities = {
        "外墙批嵌": 70.0,
        "外墙批嵌以及修补": 6.0,
        "拆改及拆墙": 7.8,
        "砌120厚砖墙": 5.6,
        "砌240厚砖墙": 4.5,
        "打混凝土过梁孔": 4,
        "厨房、卫生间排污管包隔音棉": 2.4,
        "包上/下水管道(单管)": 2.1,
        "全屋定制": 5.2,
        "地柜": 3.0,
        "吊柜": 2.0,
        "地面瓷砖": 17,
        "墙面瓷砖": 55,
        "美缝": 55.4,
        "阳台推拉门": 4.2,
        "阳台推拉门双包套": 6.2,
    }
    for item_name, expected_quantity in expected_quantities.items():
        row = _row_containing(rows, item_name)
        assert row[3] == expected_quantity
        assert row[9] == "自动汇总"
    assert not _has_item_row(rows, "橱柜")

    lintel = _row_containing(rows, "砖墙门窗洞过梁")
    assert lintel[3] == 0
    assert lintel[9] == "自动汇总"
    assert lintel[12] == "砖墙门窗洞过梁标识数量汇总"
    assert lintel[13] == "自动生成-默认推断"
    assert "QUOTE_LINTEL" in lintel[14]
    assert _row_containing(rows, "美缝")[12] == "地砖面积+2.5m以下墙面贴砖面积+QUOTE_WALL_TILE显式墙砖面积"
    assert _summary_value(rows, "自动汇总") == 17
    assert _summary_value(rows, "模板默认") == 0


def test_marker_rich_quote_sample_script_writes_reusable_outputs(tmp_path: Path):
    output_dir = tmp_path / "marker-sample"
    env = os.environ.copy()
    env["PYTHONPATH"] = str(Path(__file__).resolve().parents[1] / "src")

    completed = subprocess.run(
        [
            sys.executable,
            "scripts/generate_marker_rich_quote_sample.py",
            "--output-dir",
            str(output_dir),
        ],
        cwd=Path(__file__).resolve().parents[1],
        env=env,
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    expected_files = {
        "marker-rich-plan.dxf",
        "marker-rich-template.xlsx",
        "project.json",
        "result.json",
        "quote.xlsx",
        "README.md",
    }
    assert {path.name for path in output_dir.iterdir()} == expected_files

    rows = list(load_workbook(output_dir / "quote.xlsx", data_only=False).active.iter_rows(values_only=True))
    assert _row_containing(rows, "外墙批嵌")[3] == 70.0
    assert _row_containing(rows, "阳台推拉门双包套")[3] == 6.2
    assert _row_containing(rows, "地柜")[3] == 3.0
    assert _row_containing(rows, "吊柜")[3] == 2.0
    assert not _has_item_row(rows, "橱柜")
    assert _row_containing(rows, "墙面瓷砖")[3] == 55
    assert _row_containing(rows, "美缝")[3] == 55.4
    assert _summary_value(rows, "自动汇总") == 17
    assert _summary_value(rows, "模板默认") == 0
    readme = (output_dir / "README.md").read_text(encoding="utf-8")
    assert "QUOTE_EXT_WALL" in readme
    assert "自动汇总: 17" in readme


def _build_marker_rich_dxf(path: Path) -> None:
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4
    doc.appids.new("CAD_BUDGET")
    doc.blocks.new("pipe_block")
    modelspace = doc.modelspace()
    for layer in [
        "QUOTE_ROOM",
        "QUOTE_TEXT",
        "QUOTE_DOOR",
        "QUOTE_EXT_WALL",
        "QUOTE_EXT_OPENING",
        "QUOTE_EXT_REPAIR",
        "QUOTE_DEMO_WALL",
        "QUOTE_NEW_WALL",
        "QUOTE_PIPE_INSULATION",
        "QUOTE_PIPE_WRAP",
        "QUOTE_CUSTOM",
        "QUOTE_CABINET",
        "QUOTE_BASE_CABINET",
        "QUOTE_WALL_CABINET",
        "QUOTE_WALL_TILE",
    ]:
        doc.layers.add(layer)

    modelspace.add_lwpolyline(
        [(0, 0), (4000, 0), (4000, 3000), (0, 3000), (0, 0)],
        dxfattribs={"layer": "QUOTE_ROOM", "closed": True},
    )
    modelspace.add_text("厨房", dxfattribs={"layer": "QUOTE_TEXT", "height": 250}).set_placement((1500, 1200))
    modelspace.add_lwpolyline(
        [(5000, 0), (8000, 0), (8000, 2000), (5000, 2000), (5000, 0)],
        dxfattribs={"layer": "QUOTE_ROOM", "closed": True},
    )
    modelspace.add_text("阳台", dxfattribs={"layer": "QUOTE_TEXT", "height": 250}).set_placement((6000, 1000))

    modelspace.add_lwpolyline(
        [(-500, -500), (8500, -500), (8500, 3500), (-500, 3500), (-500, -500)],
        dxfattribs={"layer": "QUOTE_EXT_WALL", "closed": True},
    )
    modelspace.add_lwpolyline([(1000, -500), (2000, -500)], dxfattribs={"layer": "QUOTE_EXT_OPENING"})
    modelspace.add_lwpolyline(
        [(9000, 0), (11000, 0), (11000, 3000), (9000, 3000), (9000, 0)],
        dxfattribs={"layer": "QUOTE_EXT_REPAIR", "closed": True},
    )

    demo = modelspace.add_lwpolyline([(0, 500), (3000, 500)], dxfattribs={"layer": "QUOTE_DEMO_WALL"})
    demo.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2600")])
    new_120 = modelspace.add_lwpolyline([(0, 1000), (2000, 1000)], dxfattribs={"layer": "QUOTE_NEW_WALL"})
    new_120.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2800"), (1000, "THICKNESS=120")])
    new_240 = modelspace.add_lwpolyline([(0, 1500), (1500, 1500)], dxfattribs={"layer": "QUOTE_NEW_WALL"})
    new_240.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=3000"), (1000, "THICKNESS=240")])

    insulation = modelspace.add_point((3500, 1000), dxfattribs={"layer": "QUOTE_PIPE_INSULATION"})
    insulation.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2400")])
    pipe_wrap = modelspace.add_blockref("pipe_block", (3500, 1500), dxfattribs={"layer": "QUOTE_PIPE_WRAP"})
    pipe_wrap.add_attrib("HEIGHT", "2.1")

    custom = modelspace.add_line((500, 2200), (2500, 2200), dxfattribs={"layer": "QUOTE_CUSTOM"})
    custom.set_xdata("CAD_BUDGET", [(1000, "TYPE=衣柜")])
    modelspace.add_line((500, 2600), (3500, 2600), dxfattribs={"layer": "QUOTE_BASE_CABINET"})
    modelspace.add_line((500, 2600), (2500, 2600), dxfattribs={"layer": "QUOTE_WALL_CABINET"})
    wall_tile = modelspace.add_line((5200, 1800), (7200, 1800), dxfattribs={"layer": "QUOTE_WALL_TILE"})
    wall_tile.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=1200")])

    modelspace.add_lwpolyline([(5000, 500), (7000, 500)], dxfattribs={"layer": "QUOTE_DOOR"})

    doc.saveas(path)


def _create_marker_quote_template(path: Path) -> None:
    workbook = Workbook()
    half = workbook.active
    half.title = "半包"
    _write_quote_header(half)
    half.append(["一", "半包工程"])
    half.append([1, "半包项目", "M2", 1, 2, 3, 4, None, "不应读取"])

    fitout = workbook.create_sheet("整装")
    _write_quote_header(fitout)
    fitout.append(["一", "标识自动化工程"])
    items = [
        ("外墙批嵌", "M2", 77, "外墙"),
        ("外墙批嵌以及修补", "M2", 88, "外墙修补"),
        ("拆改及拆墙", "M2", 93, "拆墙"),
        ("砌120厚砖墙", "M2", 44.8, "120墙"),
        ("砌240厚砖墙", "M2", 64.56, "240墙"),
        ("砖墙门窗洞过梁", "支", 15, "人工填写"),
        ("打混凝土过梁孔", "个", 108, "过梁孔"),
        ("厨房、卫生间排污管包隔音棉", "M", 50, "隔音棉"),
        ("包上/下水管道(单管)", "M", 36, "包管"),
        ("全屋定制", "M2", 66, "全屋定制"),
        ("地柜", "M", 18, "地柜"),
        ("吊柜", "M", 16, "吊柜"),
        ("橱柜", "M", 22, "橱柜"),
        ("地面瓷砖", "片", 99, "地砖(750X1500)"),
        ("墙面瓷砖", "片", 99, "墙砖(600x1200)"),
        ("美缝", "M2", 99, "美缝"),
        ("阳台推拉门", "M2", 66, "阳台推拉门"),
        ("阳台推拉门双包套", "M", 55, "阳台推拉门双包套"),
    ]
    for index, (name, unit, quantity, note) in enumerate(items, start=1):
        fitout.append([index, name, unit, quantity, 0, 0, 10, None, note])
    fitout.append([None, "小 计", None, None, None, None, None, "=SUM(H6:H18)"])
    fitout.append(["A", "直接费合计", None, None, None, None, None, "=H19"])
    workbook.save(path)


def _write_quote_header(sheet) -> None:
    sheet.append(["工程(预) 算表"])
    sheet.append(["名称：Marker Demo"])
    sheet.append([
        "编号",
        "项目名称",
        "单位",
        "数量",
        "材料费(元)",
        None,
        "人工费\n(元)",
        "总价(元)",
        "材  料  及  工  艺  说  明",
    ])
    sheet.append([None, None, None, None, "主材\n单价", "辅材\n单价"])


def _row_containing(rows, item_name: str):
    for row in rows:
        if row[1] == item_name:
            return row
    raise AssertionError(f"Missing quote row for {item_name}")


def _has_item_row(rows, item_name: str) -> bool:
    return any(row[1] == item_name for row in rows)


def _summary_value(rows, label: str):
    for row in rows:
        if row[16] == label:
            return row[17]
    raise AssertionError(f"Missing summary row for {label}")
