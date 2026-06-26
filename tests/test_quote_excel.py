import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cad_budget.models import (
    ConstructionKind,
    ConstructionQuantityDetail,
    DataStatus,
    DoorQuantityDetail,
    ExteriorQuantityRow,
    FixtureKind,
    FixturePricingMode,
    FixtureQuantityDetail,
    HeightMode,
    QuantityRow,
    QuantityResult,
    SpaceType,
    WindowQuantityDetail,
)
from cad_budget import quote_excel
from cad_budget.quote_excel import export_residential_quote, load_default_quote_rules, load_quote_rules, parse_quote_template


def test_parse_quote_template_reads_fitout_sheet_and_ignores_half_package(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    _create_quote_template(template_path)

    template = parse_quote_template(template_path)

    assert template.sheet_name == "\u6574\u88c5"
    assert [section.name for section in template.sections] == [
        "\u5ba2\u5385\u5de5\u7a0b",
        "\u53a8\u623f\u5de5\u7a0b",
        "\u5176\u4ed6\u5de5\u7a0b",
    ]
    assert template.sections[0].items[0].name == "\u9876\u9762\u6279\u5d4c"
    assert template.sections[0].items[0].labor_price == 10
    assert all(item.name != "\u534a\u5305\u9879\u76ee" for section in template.sections for item in section.items)


def test_load_default_quote_rules_reads_packaged_rule_file():
    rules = load_default_quote_rules()

    assert rules.kitchen_waterproof_wall_height == 0.3
    assert rules.bathroom_waterproof_wall_height == 1.8
    assert rules.wall_tile_height == 2.5
    assert "\u5783\u573e\u6e05\u8fd0\u8d39" in rules.floor_area_aggregate_items
    assert "\u7f8e\u7f1d" in rules.tile_area_aggregate_items
    assert "\u6d74\u5ba4\u67dc" in rules.bathroom_count_aggregate_items
    assert "\u536b\u751f\u95f4\u95e8" in rules.bathroom_count_aggregate_items
    assert "\u7a97\u53f0\u77f3" in rules.window_count_aggregate_items
    assert "\u94dd\u5408\u91d1\u5c01\u95e8\u7a97" in rules.window_area_aggregate_items
    assert rules.fixed_quantity_aggregate_items["\u5168\u5c4b\u4fdd\u6d01"] == 1
    assert rules.tile_piece_loss_rate == 0.05
    assert rules.wide_door_width_threshold == 1.4
    assert rules.default_door_height == 2.1
    assert "\u7a97\u5e18" in rules.curtain_wall_length_items
    assert "\u5730\u9762\u74f7\u7816" in rules.floor_tile_piece_items
    assert "\u5899\u9762\u74f7\u7816" in rules.wall_tile_piece_items
    assert "\u74f7\u7816\u52a0\u5de5\u8d39" in rules.tile_processing_area_items
    assert "\u5ba4\u5185\u95e8" in rules.interior_door_count_items
    assert "\u53a8\u623f\u63a8\u62c9\u95e8" in rules.sliding_door_area_items
    assert "\u53a8\u623f\u63a8\u62c9\u95e8\u53cc\u5305\u5957" in rules.sliding_door_trim_length_items
    assert "\u5916\u5899\u6279\u5d4c" in rules.exterior_net_area_aggregate_items
    assert "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865" not in rules.exterior_net_area_aggregate_items
    assert "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865" in rules.exterior_repair_area_items
    assert rules.sliding_door_room_keywords_by_item["\u9633\u53f0\u63a8\u62c9\u95e8"] == {"\u9633\u53f0", "\u9732\u53f0"}
    assert rules.sliding_door_trim_room_keywords_by_item["\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957"] == {
        "\u9633\u53f0",
        "\u9732\u53f0",
    }
    assert "\u6dcb\u6d74\u9694\u65ad" in rules.bathroom_count_aggregate_items
    assert "\u5168\u5c4b\u5b9a\u5236" in rules.custom_projected_area_items
    assert "\u6a71\u67dc" in rules.cabinet_length_items
    assert "\u5730\u67dc" in rules.base_cabinet_length_items
    assert "\u540a\u67dc" in rules.wall_cabinet_length_items
    assert rules.default_custom_height == 2.6
    assert rules.low_custom_height_threshold == 1.0
    assert "\u62c6\u6539\u53ca\u62c6\u5899" in rules.demo_wall_area_items
    assert rules.new_wall_area_items_by_thickness["\u780c120\u539a\u7816\u5899"] == 0.12
    assert rules.new_wall_area_items_by_thickness["\u780c240\u539a\u7816\u5899"] == 0.24
    assert "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881" in rules.lintel_count_items
    assert "\u80cc\u666f\u5899" in rules.background_wall_area_items
    assert "\u5165\u6237\u95e8" in rules.entry_door_count_items
    assert "\u73bb\u7483\u6dcb\u6d74\u623f" in rules.shower_glass_count_items
    assert "\u8e72\u5751" in rules.squat_toilet_count_items
    assert "\u53a8\u623f\u3001\u536b\u751f\u95f4\u6392\u6c61\u7ba1\u5305\u9694\u97f3\u68c9" in rules.pipe_insulation_length_items
    assert "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053(\u5355\u7ba1)" in rules.pipe_wrap_length_items
    assert "\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54" in rules.building_area_percent_count_items
    assert rules.building_area_percent_count_items["\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54"] == 0.1


def test_export_residential_quote_uses_loaded_quote_rules(tmp_path: Path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path)
    custom_rules = load_default_quote_rules()
    custom_rules.kitchen_waterproof_wall_height = 0.5
    custom_rules.bathroom_waterproof_wall_height = 1.2
    custom_rules.wall_tile_height = 2.0
    monkeypatch.setattr(quote_excel, "load_quote_rules", lambda rules_path=None: custom_rules)
    result = QuantityResult(
        project_name="Custom Rules Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                window_area=1.0,
            ),
            _quantity_row(
                "bath",
                "\u4e3b\u536b",
                floor_area=3.0,
                net_wall_area=15.0,
                wall_measure_perimeter=8.0,
                window_area=0.5,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    kitchen_waterproof = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406")
    bath_waterproof = _row_containing_after(rows, "\u4e3b\u536b\u5de5\u7a0b", "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406")
    kitchen_wall_tile = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)")
    assert kitchen_waterproof[3] == 11.0
    assert bath_waterproof[3] == 12.6
    assert kitchen_wall_tile[3] == 19.0


def test_load_quote_rules_reads_external_rule_file(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    _write_custom_quote_rules(rules_path, kitchen_height=0.5, bathroom_height=1.2, tile_height=2.0)

    rules = load_quote_rules(rules_path)

    assert rules.kitchen_waterproof_wall_height == 0.5
    assert rules.bathroom_waterproof_wall_height == 1.2
    assert rules.wall_tile_height == 2.0
    assert rules.custom_projected_area_items == set()
    assert rules.cabinet_length_items == set()
    assert rules.base_cabinet_length_items == set()
    assert rules.wall_cabinet_length_items == set()
    assert rules.default_custom_height == 2.6
    assert rules.low_custom_height_threshold == 1.0
    assert rules.source_label == str(rules_path)


def test_load_quote_rules_accepts_legacy_floor_area_percent_count_items(tmp_path: Path):
    rules_path = tmp_path / "legacy-rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "floor_area_percent_count_items": {"\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54": 0.1},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    rules = load_quote_rules(rules_path)

    assert rules.building_area_percent_count_items == {"\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54": 0.1}


def test_load_quote_rules_reports_invalid_external_rule_file(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
            }
        ),
        encoding="utf-8",
    )

    try:
        load_quote_rules(rules_path)
    except ValueError as exc:
        assert "Invalid quote rules" in str(exc)
        assert "bathroom_waterproof_wall_height" in str(exc)
    else:
        raise AssertionError("Expected invalid quote rules to raise ValueError")


def test_load_quote_rules_reports_non_numeric_height(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": "bad",
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
            }
        ),
        encoding="utf-8",
    )

    try:
        load_quote_rules(rules_path)
    except ValueError as exc:
        assert "kitchen_waterproof_wall_height must be a number" in str(exc)
    else:
        raise AssertionError("Expected invalid quote rules to raise ValueError")


def test_load_quote_rules_reports_non_numeric_fixed_quantity(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "fixed_quantity_aggregate_items": {"\u5168\u5c4b\u4fdd\u6d01": "bad"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    try:
        load_quote_rules(rules_path)
    except ValueError as exc:
        assert "fixed_quantity_aggregate_items.\u5168\u5c4b\u4fdd\u6d01 must be a number" in str(exc)
    else:
        raise AssertionError("Expected invalid quote rules to raise ValueError")


def test_load_quote_rules_reports_non_numeric_advanced_number(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "tile_piece_loss_rate": "bad",
                "wide_door_width_threshold": "wide",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    try:
        load_quote_rules(rules_path)
    except ValueError as exc:
        assert "tile_piece_loss_rate must be a number" in str(exc)
    else:
        raise AssertionError("Expected invalid quote rules to raise ValueError")


def test_load_quote_rules_reports_non_numeric_door_threshold(tmp_path: Path):
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "wide_door_width_threshold": "wide",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    try:
        load_quote_rules(rules_path)
    except ValueError as exc:
        assert "wide_door_width_threshold must be a number" in str(exc)
    else:
        raise AssertionError("Expected invalid quote rules to raise ValueError")


def test_export_residential_quote_accepts_external_rules_and_records_source(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    rules_path = tmp_path / "rules.json"
    _create_quote_template(template_path)
    _write_custom_quote_rules(rules_path, kitchen_height=0.5, bathroom_height=1.2, tile_height=2.0)
    result = QuantityResult(
        project_name="External Rules Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                window_area=1.0,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path, rules_path=rules_path)

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    kitchen_waterproof = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406")
    kitchen_wall_tile = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)")
    assert kitchen_waterproof[3] == 11.0
    assert kitchen_wall_tile[3] == 19.0
    assert sheet["Q7"].value == "\u89c4\u5219\u6765\u6e90"
    assert sheet["R7"].value == str(rules_path)


def test_export_residential_quote_generates_actual_room_sections_and_preserves_manual_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path)
    result = QuantityResult(
        project_name="Quote Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0, wall_measure_perimeter=10.0, window_area=1.0),
            _quantity_row("bath", "\u4e3b\u536b", floor_area=3.0, net_wall_area=15.0, wall_measure_perimeter=8.0, window_area=0.5),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert workbook.sheetnames == ["\u5546\u54c1\u623f\u6574\u88c5\u62a5\u4ef7"]
    assert [sheet.cell(row=3, column=column).value for column in range(10, 16)] == [
        "\u6570\u91cf\u6765\u6e90",
        "\u6765\u6e90\u7a7a\u95f4",
        "\u7a7a\u95f4ID",
        "\u8ba1\u91cf\u53e3\u5f84",
        "\u590d\u6838\u72b6\u6001",
        "\u590d\u6838\u5907\u6ce8",
    ]
    assert _row_containing(rows, "\u5ba2\u5385\u5de5\u7a0b") is not None
    assert _row_containing(rows, "\u53a8\u623f\u5de5\u7a0b") is not None
    assert _row_containing(rows, "\u4e3b\u536b\u5de5\u7a0b") is not None

    living_wall_paint = _row_containing(rows, "\u5899\u9762\u4e73\u80f6\u6f06")
    assert living_wall_paint[3] == 50.0
    assert living_wall_paint[4:7] == (10, 0, 10)
    assert living_wall_paint[9:15] == (
        "\u81ea\u52a8\u7b97\u91cf",
        "\u5ba2\u5385",
        "living",
        "\u5899\u9762\u51c0\u9762\u79ef",
        "\u81ea\u52a8\u751f\u6210",
        None,
    )

    kitchen_floor_tile = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5730\u9762\u7816\u94fa\u8d34(750X1500)")
    assert kitchen_floor_tile[3] == 6.0

    bath_waterproof = _row_containing_after(rows, "\u4e3b\u536b\u5de5\u7a0b", "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406")
    assert bath_waterproof[3] == 17.4
    assert bath_waterproof[12] == "\u5730\u9762\u9762\u79ef+1.8m\u4ee5\u4e0b\u5899\u9762\u9762\u79ef"

    kitchen_waterproof = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406")
    assert kitchen_waterproof[3] == 9.0
    assert kitchen_waterproof[12] == "\u5730\u9762\u9762\u79ef+0.3m\u4ee5\u4e0b\u5899\u9762\u9762\u79ef"

    kitchen_wall_tile = _row_containing_after(rows, "\u53a8\u623f\u5de5\u7a0b", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)")
    bath_wall_tile = _row_containing_after(rows, "\u4e3b\u536b\u5de5\u7a0b", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)")
    assert kitchen_wall_tile[3] == 24.0
    assert bath_wall_tile[3] == 19.5
    assert kitchen_wall_tile[12] == "2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef"

    manual_item = _row_containing(rows, "\u7a97\u5e18")
    assert manual_item[3] == 1
    assert manual_item[7].startswith("=D")
    assert manual_item[9:15] == (
        "\u6a21\u677f\u9ed8\u8ba4",
        None,
        None,
        "\u6a21\u677f\u9ed8\u8ba4\u6570\u91cf",
        "\u6309\u6a21\u677f\u751f\u6210",
        None,
    )

    assert _row_containing(rows, "\u76f4\u63a5\u8d39\u5408\u8ba1") is not None
    assert _row_containing(rows, "\u5de5\u7a0b\u7ba1\u7406\u8d39") is not None
    assert _row_containing(rows, "\u5de5\u7a0b\u603b\u9020\u4ef7") is not None
    assert sheet.auto_filter.ref == f"A3:O{sheet.max_row}"
    assert sheet["Q1"].value == "\u62a5\u4ef7\u81ea\u52a8\u5316\u7edf\u8ba1"
    assert [sheet.cell(row=2, column=column).value for column in range(17, 20)] == [
        "\u6570\u91cf\u6765\u6e90",
        "\u884c\u6570",
        "\u5360\u6bd4",
    ]
    assert [sheet.cell(row=row, column=17).value for row in range(3, 6)] == [
        "\u81ea\u52a8\u7b97\u91cf",
        "\u81ea\u52a8\u6c47\u603b",
        "\u6a21\u677f\u9ed8\u8ba4",
    ]
    assert [sheet.cell(row=row, column=18).value for row in range(3, 6)] == [14, 1, 1]
    assert sheet["S3"].value == "=R3/SUM(R3:R5)"
    assert sheet["S4"].value == "=R4/SUM(R3:R5)"
    assert sheet["S5"].value == "=R5/SUM(R3:R5)"


def test_export_residential_quote_auto_fills_custom_and_cabinet_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_custom_cabinet_items=True)
    result = QuantityResult(
        project_name="Custom Cabinet Quote Demo",
        rows=[
            _quantity_row(
                "bed",
                "\u4e3b\u5367",
                floor_area=12.0,
                net_wall_area=30.0,
                custom_details=[
                    FixtureQuantityDetail(
                        id="wardrobe",
                        room_id="bed",
                        room_name="\u4e3b\u5367",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        height=None,
                        effective_height=2.6,
                        height_defaulted=True,
                        projected_area=5.2,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                        fixture_type="\u8863\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="low",
                        room_id="bed",
                        room_name="\u4e3b\u5367",
                        kind=FixtureKind.CUSTOM,
                        length=1.0,
                        height=0.8,
                        effective_height=0.8,
                        height_defaulted=False,
                        projected_area=0.0,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u77ee\u67dc",
                    ),
                ],
            ),
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                cabinet_details=[
                    FixtureQuantityDetail(
                        id="base",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=3.0,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u5730\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="wall",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=3.0,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u540a\u67dc",
                    ),
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    custom = _item_row_named(rows, "\u5168\u5c4b\u5b9a\u5236")
    base = _item_row_named(rows, "\u5730\u67dc")
    wall = _item_row_named(rows, "\u540a\u67dc")
    assert custom[3] == 5.2
    assert custom[12] == "\u5168\u5c4b\u5b9a\u5236\u6295\u5f71\u9762\u79ef\u6c47\u603b"
    assert custom[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert "\u9ed8\u8ba42.6m" in custom[14]
    assert "\u9ad8\u5ea6\u5c0f\u4e8e1m" in custom[14]
    assert _item_row_named(rows, "\u6a71\u67dc") is None
    assert base[3] == 3.0
    assert base[12] == "\u5730\u67dc\u957f\u5ea6\u6c47\u603b"
    assert wall[3] == 3.0
    assert wall[12] == "\u540a\u67dc\u957f\u5ea6\u6c47\u603b"


def test_export_residential_quote_auto_fills_split_base_and_wall_cabinet_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_split_cabinet_items=True)
    result = QuantityResult(
        project_name="Split Cabinet Quote Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                cabinet_details=[
                    FixtureQuantityDetail(
                        id="base",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=3.2,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u5730\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="wall",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=2.1,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u540a\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="unknown",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=1.0,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type=None,
                    ),
                ],
            )
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    base = _item_row_named(rows, "\u5730\u67dc")
    wall = _item_row_named(rows, "\u540a\u67dc")
    assert base[3] == 3.2
    assert base[12] == "\u5730\u67dc\u957f\u5ea6\u6c47\u603b"
    assert base[13] == "\u81ea\u52a8\u751f\u6210"
    assert base[14] is None
    assert wall[3] == 2.1
    assert wall[12] == "\u540a\u67dc\u957f\u5ea6\u6c47\u603b"
    assert wall[13] == "\u81ea\u52a8\u751f\u6210"
    assert _item_row_named(rows, "\u6a71\u67dc") is None


def test_export_residential_quote_splits_generic_cabinet_item_by_base_and_wall_lengths(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_custom_cabinet_items=True)
    result = QuantityResult(
        project_name="Generic Cabinet Split Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                cabinet_details=[
                    FixtureQuantityDetail(
                        id="base",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=3.2,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u5730\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="wall",
                        room_id="kitchen",
                        room_name="\u53a8\u623f",
                        kind=FixtureKind.CABINET,
                        length=2.1,
                        pricing_mode=FixturePricingMode.LENGTH,
                        fixture_type="\u540a\u67dc",
                    ),
                ],
            )
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    assert _item_row_named(rows, "\u6a71\u67dc") is None
    base = _item_row_named(rows, "\u5730\u67dc")
    wall = _item_row_named(rows, "\u540a\u67dc")
    assert base[3] == 3.2
    assert base[12] == "\u5730\u67dc\u957f\u5ea6\u6c47\u603b"
    assert wall[3] == 2.1
    assert wall[12] == "\u540a\u67dc\u957f\u5ea6\u6c47\u603b"


def test_export_residential_quote_excludes_low_height_projected_custom_details(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_custom_cabinet_items=True)
    result = QuantityResult(
        project_name="Persisted Low Custom Detail Demo",
        rows=[
            _quantity_row(
                "bed",
                "\u4e3b\u5367",
                floor_area=12.0,
                net_wall_area=30.0,
                custom_details=[
                    FixtureQuantityDetail(
                        id="low-persisted",
                        room_id="bed",
                        room_name="\u4e3b\u5367",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        height=0.8,
                        effective_height=0.8,
                        height_defaulted=False,
                        projected_area=1.6,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                        fixture_type="\u77ee\u67dc",
                    ),
                    FixtureQuantityDetail(
                        id="wardrobe",
                        room_id="bed",
                        room_name="\u4e3b\u5367",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        height=2.4,
                        effective_height=2.4,
                        height_defaulted=False,
                        projected_area=4.8,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                        fixture_type="\u8863\u67dc",
                    ),
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    custom = _item_row_named(rows, "\u5168\u5c4b\u5b9a\u5236")
    assert custom[3] == 4.8
    assert custom[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert "\u9ad8\u5ea6\u5c0f\u4e8e1m" in custom[14]


def test_export_residential_quote_keeps_custom_template_default_when_only_low_height_projected_details(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_custom_cabinet_items=True)
    result = QuantityResult(
        project_name="Only Low Custom Detail Demo",
        rows=[
            _quantity_row(
                "bed",
                "\u4e3b\u5367",
                floor_area=12.0,
                net_wall_area=30.0,
                custom_details=[
                    FixtureQuantityDetail(
                        id="low-persisted",
                        room_id="bed",
                        room_name="\u4e3b\u5367",
                        kind=FixtureKind.CUSTOM,
                        length=2.0,
                        height=0.8,
                        effective_height=0.8,
                        height_defaulted=False,
                        projected_area=1.6,
                        pricing_mode=FixturePricingMode.PROJECTED_AREA,
                        fixture_type="\u77ee\u67dc",
                    ),
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    custom = _item_row_named(rows, "\u5168\u5c4b\u5b9a\u5236")
    assert custom[3] == 99
    assert custom[12] == "\u6a21\u677f\u9ed8\u8ba4\u6570\u91cf"
    assert custom[13] == "\u6309\u6a21\u677f\u751f\u6210"


def test_export_residential_quote_keeps_custom_and_cabinet_template_defaults_without_details(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_custom_cabinet_items=True)
    result = QuantityResult(
        project_name="No Fixture Details Demo",
        rows=[_quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0)],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    custom = _item_row_named(rows, "\u5168\u5c4b\u5b9a\u5236")
    cabinet = _item_row_named(rows, "\u6a71\u67dc")
    assert custom[3] == 99
    assert custom[9:15] == (
        "\u6a21\u677f\u9ed8\u8ba4",
        None,
        None,
        "\u6a21\u677f\u9ed8\u8ba4\u6570\u91cf",
        "\u6309\u6a21\u677f\u751f\u6210",
        None,
    )
    assert cabinet[3] == 99
    assert cabinet[9:15] == custom[9:15]


def test_export_residential_quote_auto_fills_whole_house_area_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_area_summary_items=True)
    result = QuantityResult(
        project_name="Area Summary Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0, wall_measure_perimeter=10.0, window_area=1.0),
            _quantity_row("bath", "\u4e3b\u536b", floor_area=3.0, net_wall_area=15.0, wall_measure_perimeter=8.0, window_area=0.5),
            _quantity_row("shaft", "\u7535\u68af\u4e95", floor_area=4.0, net_wall_area=12.0, status=DataStatus.EXCLUDED),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    garbage = _row_containing(rows, "\u5783\u573e\u6e05\u8fd0\u8d39")
    wiring = _row_containing(rows, "\u5f3a\u7535\u5e03\u7ebf")
    maintenance = _row_containing(rows, "\u5730\u9762\u7816\u73b0\u573a\u7ef4\u62a4\u8d39")
    assert garbage[3] == 29.0
    assert wiring[3] == 29.0
    assert maintenance[3] == 29.0
    assert garbage[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u5ba4\u5185\u5730\u9762\u9762\u79ef\u6c47\u603b",
        "\u81ea\u52a8\u751f\u6210",
        None,
    )


def test_export_residential_quote_auto_fills_tile_area_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_area_summary_items=True)
    result = QuantityResult(
        project_name="Tile Area Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0, wall_measure_perimeter=10.0, window_area=1.0),
            _quantity_row("bath", "\u4e3b\u536b", floor_area=3.0, net_wall_area=15.0, wall_measure_perimeter=8.0, window_area=0.5),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    tile_grout = _row_containing(rows, "\u7f8e\u7f1d")
    assert tile_grout[3] == 72.5
    assert tile_grout[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u5730\u7816\u9762\u79ef+2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef",
        "\u81ea\u52a8\u751f\u6210",
        None,
    )


def test_export_residential_quote_keeps_aggregate_review_notes_item_specific(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    rules_path = tmp_path / "rules.json"
    _create_quote_template(template_path, include_area_summary_items=True, include_count_summary_items=True)
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": ["\u5783\u573e\u6e05\u8fd0\u8d39"],
                "tile_area_aggregate_items": ["\u7f8e\u7f1d"],
                "window_count_aggregate_items": ["\u7a97\u53f0\u77f3"],
                "window_area_aggregate_items": ["\u94dd\u5408\u91d1\u5c01\u95e8\u7a97"],
                "door_count_aggregate_items": ["\u5ba4\u5185\u95e8"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    result = QuantityResult(
        project_name="Aggregate Review Demo",
        rows=[
            _quantity_row(
                "living",
                "\u5ba2\u5385",
                floor_area=20.0,
                net_wall_area=50.0,
                window_count=1,
                window_area=2.0,
                window_details=[
                    WindowQuantityDetail(
                        id="w1",
                        width=1.0,
                        height=2.0,
                        area=2.0,
                        height_defaulted=True,
                        wall_segment_key="living:0",
                        wall_segment_length=4.0,
                    )
                ],
                door_opening_count=1,
                door_details=[
                    DoorQuantityDetail(
                        id="d1",
                        room_id="living",
                        width=0.9,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=1.89,
                    )
                ],
            ),
            _quantity_row(
                "bath",
                "\u4e3b\u536b",
                floor_area=3.0,
                net_wall_area=15.0,
                wall_measure_perimeter=8.0,
                window_count=1,
                window_area=0.5,
                window_details=[
                    WindowQuantityDetail(
                        id="w2",
                        width=0.5,
                        height=1.0,
                        area=0.5,
                        height_defaulted=True,
                        wall_segment_key="bath:0",
                        wall_segment_length=2.0,
                    )
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path, rules_path=rules_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    garbage = _item_row_named(rows, "\u5783\u573e\u6e05\u8fd0\u8d39")
    tile_grout = _item_row_named(rows, "\u7f8e\u7f1d")
    windowsill = _item_row_named(rows, "\u7a97\u53f0\u77f3")
    sealed_windows = _item_row_named(rows, "\u94dd\u5408\u91d1\u5c01\u95e8\u7a97")
    door_count = _item_row_named(rows, "\u5ba4\u5185\u95e8")
    assert garbage[13:15] == ("\u81ea\u52a8\u751f\u6210", None)
    assert tile_grout[13:15] == ("\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad", "\u5899\u7816\u9762\u79ef\u6263\u7a97\u4f7f\u7528\u9ed8\u8ba4\u7a97\u9ad8\uff0c\u9700\u590d\u6838\u7a97\u9ad8")
    assert windowsill[13:15] == ("\u81ea\u52a8\u751f\u6210", None)
    assert sealed_windows[13:15] == ("\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad", "\u7a97\u9762\u79ef\u4f7f\u7528\u9ed8\u8ba4\u7a97\u9ad8\uff0c\u9700\u590d\u6838\u7a97\u9ad8")
    assert door_count[13:15] == ("\u81ea\u52a8\u751f\u6210", None)


def test_export_residential_quote_auto_fills_count_and_opening_aggregate_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    rules_path = tmp_path / "rules.json"
    _create_quote_template(template_path, include_count_summary_items=True)
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "room_count_aggregate_items": ["\u623f\u95f4\u6210\u54c1\u4fdd\u62a4"],
                "bathroom_count_aggregate_items": ["\u6d74\u5ba4\u67dc"],
                "window_count_aggregate_items": ["\u7a97\u53f0\u77f3"],
                "window_area_aggregate_items": ["\u94dd\u5408\u91d1\u5c01\u95e8\u7a97"],
                "door_count_aggregate_items": ["\u5ba4\u5185\u95e8"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    result = QuantityResult(
        project_name="Count Summary Demo",
        rows=[
            _quantity_row(
                "living",
                "\u5ba2\u5385",
                floor_area=20.0,
                net_wall_area=50.0,
                window_count=2,
                window_area=5.5,
                door_opening_count=1,
            ),
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                window_count=1,
                window_area=1.0,
                door_opening_count=1,
            ),
            _quantity_row(
                "bath",
                "\u4e3b\u536b",
                floor_area=3.0,
                net_wall_area=15.0,
                wall_measure_perimeter=8.0,
                window_count=1,
                window_area=0.5,
                door_opening_count=1,
            ),
            _quantity_row(
                "shaft",
                "\u7535\u68af\u4e95",
                floor_area=4.0,
                net_wall_area=12.0,
                window_count=3,
                window_area=9.0,
                door_opening_count=2,
                status=DataStatus.EXCLUDED,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path, rules_path=rules_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    room_protection = _row_containing(rows, "\u623f\u95f4\u6210\u54c1\u4fdd\u62a4")
    vanity = _row_containing(rows, "\u6d74\u5ba4\u67dc")
    windowsill = _item_row_named(rows, "\u7a97\u53f0\u77f3")
    sealed_windows = _row_containing(rows, "\u94dd\u5408\u91d1\u5c01\u95e8\u7a97")
    interior_door = _item_row_named(rows, "\u5ba4\u5185\u95e8")
    assert room_protection[3] == 3
    assert vanity[3] == 1
    assert windowsill[3] == 4
    assert sealed_windows[3] == 7.0
    assert interior_door[3] == 3
    assert room_protection[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u6709\u6548\u7a7a\u95f4\u6570\u91cf\u6c47\u603b",
        "\u81ea\u52a8\u751f\u6210",
        None,
    )
    assert vanity[12] == "\u536b\u751f\u95f4\u6570\u91cf\u6c47\u603b"
    assert windowsill[12] == "\u7a97\u6570\u91cf\u6c47\u603b"
    assert sealed_windows[12] == "\u7a97\u9762\u79ef\u6c47\u603b"
    assert interior_door[12] == "\u95e8\u6d1e\u6570\u91cf\u6c47\u603b"


def test_export_residential_quote_auto_fills_fixed_quantity_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    rules_path = tmp_path / "rules.json"
    _create_quote_template(template_path, include_fixed_summary_items=True)
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "fixed_quantity_aggregate_items": {
                    "\u5168\u5c4b\u4fdd\u6d01": 1,
                    "\u5168\u5c4b\u63d2\u5ea7\u5f00\u5173": 2,
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    result = QuantityResult(
        project_name="Fixed Summary Demo",
        rows=[_quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0)],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path, rules_path=rules_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    cleaning = _item_row_named(rows, "\u5168\u5c4b\u4fdd\u6d01")
    switches = _item_row_named(rows, "\u5168\u5c4b\u63d2\u5ea7\u5f00\u5173")
    assert cleaning[3] == 1
    assert switches[3] == 2
    assert cleaning[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u56fa\u5b9a\u6570\u91cf\u6c47\u603b",
        "\u81ea\u52a8\u751f\u6210",
        None,
    )


def test_export_residential_quote_auto_fills_curtains_by_unique_window_wall_length(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Curtain Demo",
        rows=[
            _quantity_row(
                "living",
                "\u5ba2\u5385",
                floor_area=20.0,
                net_wall_area=50.0,
                window_details=[
                    WindowQuantityDetail(
                        id="w1",
                        width=1.2,
                        height=1.5,
                        area=1.8,
                        height_defaulted=False,
                        wall_segment_key="living:0",
                        wall_segment_length=4.0,
                    ),
                    WindowQuantityDetail(
                        id="w2",
                        width=0.8,
                        height=1.5,
                        area=1.2,
                        height_defaulted=True,
                        wall_segment_key="living:0",
                        wall_segment_length=4.0,
                    ),
                    WindowQuantityDetail(
                        id="w3",
                        width=1.0,
                        height=1.5,
                        area=1.5,
                        height_defaulted=False,
                        wall_segment_key="living:1",
                        wall_segment_length=3.0,
                    ),
                ],
            )
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    curtains = _item_row_named(rows, "\u7a97\u5e18")
    assert curtains[3] == 7.0
    assert curtains[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u7a97\u6240\u5728\u5899\u9762\u957f\u5ea6\u6c47\u603b",
        "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad",
        "\u7a97\u5e18\u6309\u540c\u623f\u95f4\u540c\u5899\u6bb5\u53bb\u91cd\uff0cL\u5f62\u7a97\u9700\u4eba\u5de5\u786e\u8ba4",
    )


def test_export_residential_quote_auto_fills_tile_piece_counts_from_area_and_spec(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Tile Piece Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0, wall_measure_perimeter=10.0, window_area=1.0),
            _quantity_row("bath", "\u4e3b\u536b", floor_area=3.0, net_wall_area=15.0, wall_measure_perimeter=8.0, window_area=0.5),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    floor_tiles = _item_row_named(rows, "\u5730\u9762\u74f7\u7816")
    wall_tiles = _item_row_named(rows, "\u5899\u9762\u74f7\u7816")
    assert floor_tiles[3] == 28
    assert floor_tiles[12] == "\u5730\u7816\u9762\u79ef\u6309750X1500\u89c4\u683c+5%\u635f\u8017\u6298\u7b97\u7247\u6570"
    assert wall_tiles[3] == 64
    assert wall_tiles[12] == "\u5899\u7816\u9762\u79ef\u6309600x1200\u89c4\u683c+5%\u635f\u8017\u6298\u7b97\u7247\u6570"


def test_export_residential_quote_adds_explicit_non_wet_wall_tile_markers(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_area_summary_items=True, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Explicit Wall Tile Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                window_area=1.0,
            ),
            _quantity_row("balcony", "\u9633\u53f0", floor_area=4.0, net_wall_area=8.0),
        ],
        construction_details=[
            ConstructionQuantityDetail(
                id="balcony-wall-tile",
                kind=ConstructionKind.WALL_TILE,
                room_id="balcony",
                room_name="\u9633\u53f0",
                length=3.0,
                height=1.2,
                effective_height=1.2,
                area=3.6,
            ),
            ConstructionQuantityDetail(
                id="kitchen-wall-tile",
                kind=ConstructionKind.WALL_TILE,
                room_id="kitchen",
                room_name="\u53a8\u623f",
                length=2.0,
                height=1.2,
                effective_height=1.2,
                area=2.4,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    wall_tiles = _item_row_named(rows, "\u5899\u9762\u74f7\u7816")
    tile_grout = _item_row_named(rows, "\u7f8e\u7f1d")
    assert wall_tiles[3] == 41
    assert wall_tiles[13] == "\u81ea\u52a8\u751f\u6210"
    assert tile_grout[3] == 57.6
    assert tile_grout[12] == "\u5730\u7816\u9762\u79ef+2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef+QUOTE_WALL_TILE\u663e\u5f0f\u5899\u7816\u9762\u79ef"


def test_export_residential_quote_auto_fills_tile_processing_by_house_area(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Tile Processing Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0),
            _quantity_row("shaft", "\u7535\u68af\u4e95", floor_area=4.0, net_wall_area=12.0, status=DataStatus.EXCLUDED),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    processing = _item_row_named(rows, "\u74f7\u7816\u52a0\u5de5\u8d39")
    assert processing[3] == 26.0
    assert processing[9:15] == (
        "\u81ea\u52a8\u6c47\u603b",
        "\u5168\u5c4b",
        None,
        "\u623f\u5b50\u9762\u79ef\u6c47\u603b",
        "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad",
        "\u74f7\u7816\u52a0\u5de5\u8d39\u6309\u623f\u5b50\u9762\u79ef\u8ba1\u7b97\uff0c\u9700\u8bbe\u8ba1\u5e08\u4fee\u6539\u786e\u8ba4",
    )


def test_export_residential_quote_parses_tile_spec_with_star_separator(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    workbook = load_workbook(template_path)
    sheet = workbook["\u6574\u88c5"]
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "\u5730\u9762\u74f7\u7816":
            sheet.cell(row=row, column=9).value = "750*1500\u74f7\u7816"
        if sheet.cell(row=row, column=2).value == "\u5899\u9762\u74f7\u7816":
            sheet.cell(row=row, column=9).value = "600*1200\u74f7\u7816"
    workbook.save(template_path)
    result = QuantityResult(
        project_name="Star Tile Spec Demo",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0),
            _quantity_row("kitchen", "\u53a8\u623f", floor_area=6.0, net_wall_area=18.0, wall_measure_perimeter=10.0, window_area=1.0),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    floor_tiles = _item_row_named(rows, "\u5730\u9762\u74f7\u7816")
    wall_tiles = _item_row_named(rows, "\u5899\u9762\u74f7\u7816")
    assert floor_tiles[3] == 25
    assert floor_tiles[12] == "\u5730\u7816\u9762\u79ef\u6309750*1500\u89c4\u683c+5%\u635f\u8017\u6298\u7b97\u7247\u6570"
    assert wall_tiles[3] == 35
    assert wall_tiles[12] == "\u5899\u7816\u9762\u79ef\u6309600*1200\u89c4\u683c+5%\u635f\u8017\u6298\u7b97\u7247\u6570"


def test_export_residential_quote_reports_unparseable_tile_spec(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    workbook = load_workbook(template_path)
    sheet = workbook["\u6574\u88c5"]
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "\u5730\u9762\u74f7\u7816":
            sheet.cell(row=row, column=9).value = "\u5730\u7816"
            break
    workbook.save(template_path)
    result = QuantityResult(
        project_name="Bad Tile Spec Demo",
        rows=[_quantity_row("living", "\u5ba2\u5385", floor_area=20.0, net_wall_area=50.0)],
        exceptions=[],
    )

    try:
        export_residential_quote(result, template_path, output_path)
    except ValueError as exc:
        assert "Cannot parse tile specification" in str(exc)
        assert "\u5730\u9762\u74f7\u7816 \u5730\u7816" in str(exc)
    else:
        raise AssertionError("Expected invalid tile specification to raise ValueError")


def test_export_residential_quote_auto_fills_doors_and_shower_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Door Demo",
        rows=[
            _quantity_row(
                "living",
                "\u5ba2\u5385",
                floor_area=20.0,
                net_wall_area=50.0,
                door_details=[
                    DoorQuantityDetail(id="d1", room_id="living", width=0.9, height=2.1, effective_height=2.1, area=1.89),
                ],
            ),
            _quantity_row(
                "bedroom",
                "\u6b21\u5367",
                floor_area=10.0,
                net_wall_area=30.0,
                door_details=[
                    DoorQuantityDetail(id="d1", room_id="bedroom", width=0.9, height=2.1, effective_height=2.1, area=1.89),
                ],
            ),
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                door_details=[
                    DoorQuantityDetail(
                        id="slide",
                        room_id="kitchen",
                        width=1.6,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=3.36,
                    )
                ],
            ),
            _quantity_row(
                "study",
                "\u4e66\u623f",
                floor_area=8.0,
                net_wall_area=24.0,
                door_details=[
                    DoorQuantityDetail(
                        id="slide",
                        room_id="study",
                        width=1.6,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=3.36,
                    )
                ],
            ),
            _quantity_row(
                "bath",
                "\u4e3b\u536b",
                floor_area=3.0,
                net_wall_area=15.0,
                wall_measure_perimeter=8.0,
                door_details=[
                    DoorQuantityDetail(id="bath-door", room_id="bath", width=0.8, height=2.1, effective_height=2.1, area=1.68),
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    interior_door = _item_row_named(rows, "\u5ba4\u5185\u95e8")
    sliding_door = _item_row_named(rows, "\u53a8\u623f\u63a8\u62c9\u95e8")
    shower_partition = _item_row_named(rows, "\u6dcb\u6d74\u9694\u65ad")
    glass_shower = _item_row_named(rows, "\u73bb\u7483\u6dcb\u6d74\u623f")
    assert interior_door[3] == 1
    assert interior_door[12] == "\u666e\u901a\u5ba4\u5185\u95e8\u6d1e\u6570\u91cf\u6c47\u603b"
    assert sliding_door[3] == 3.36
    assert sliding_door[12] == "\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u9762\u79ef\u6c47\u603b"
    assert "\u9ed8\u8ba4\u95e8\u9ad82.1m" in sliding_door[14]
    assert shower_partition[3] == 1
    assert shower_partition[12] == "\u536b\u751f\u95f4\u6570\u91cf\u6c47\u603b"
    assert glass_shower[3] == 0
    assert glass_shower[9] == "\u81ea\u52a8\u6c47\u603b"
    assert glass_shower[12] == "\u6dcb\u6d74\u623f\u6807\u8bc6\u6570\u91cf\u6c47\u603b"


def test_export_residential_quote_auto_fills_sliding_door_trim_length(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    result = QuantityResult(
        project_name="Sliding Trim Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                door_details=[
                    DoorQuantityDetail(
                        id="slide",
                        room_id="kitchen",
                        width=1.6,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=3.36,
                    )
                ],
            ),
            _quantity_row(
                "study",
                "\u4e66\u623f",
                floor_area=8.0,
                net_wall_area=24.0,
                door_details=[
                    DoorQuantityDetail(
                        id="slide",
                        room_id="study",
                        width=1.6,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=3.36,
                    )
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    trim = _item_row_named(rows, "\u53a8\u623f\u63a8\u62c9\u95e8\u53cc\u5305\u5957")
    assert trim[3] == 5.8
    assert trim[12] == "\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u5957\u7ebf\u957f\u5ea6\u6c47\u603b"
    assert trim[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert "\u9ed8\u8ba4\u95e8\u9ad82.1m" in trim[14]


def test_export_residential_quote_uses_custom_default_door_height_for_sliding_doors(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    rules_path = tmp_path / "rules.json"
    _create_quote_template(template_path, include_advanced_summary_items=True)
    rules_path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": 0.3,
                    "bathroom_waterproof_wall_height": 1.8,
                    "wall_tile_height": 2.5,
                },
                "floor_area_aggregate_items": [],
                "tile_area_aggregate_items": [],
                "sliding_door_area_items": ["\u53a8\u623f\u63a8\u62c9\u95e8"],
                "wide_door_width_threshold": 1.4,
                "default_door_height": 2.4,
                "sliding_door_room_keywords": ["\u53a8\u623f"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    result = QuantityResult(
        project_name="Custom Door Height Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                door_details=[
                    DoorQuantityDetail(
                        id="slide",
                        room_id="kitchen",
                        width=1.6,
                        height=None,
                        effective_height=2.1,
                        height_defaulted=True,
                        area=3.36,
                    )
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path, rules_path=rules_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    sliding_door = _item_row_named(rows, "\u53a8\u623f\u63a8\u62c9\u95e8")
    assert sliding_door[3] == 3.84
    assert "\u9ed8\u8ba4\u95e8\u9ad82.4m" in sliding_door[14]


def test_export_residential_quote_auto_fills_exterior_plaster_from_included_net_area(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_exterior_summary_items=True)
    result = QuantityResult(
        project_name="Exterior Quote Demo",
        rows=[],
        exterior_rows=[
            ExteriorQuantityRow(
                exterior_wall_id="ext-included",
                floor=None,
                height=3.0,
                measure_length=4.0,
                opening_length=1.0,
                gross_area=12.0,
                net_area=9.0,
                include_in_quote=True,
            ),
            ExteriorQuantityRow(
                exterior_wall_id="ext-excluded",
                floor=None,
                height=3.0,
                measure_length=5.0,
                opening_length=2.0,
                gross_area=15.0,
                net_area=9.0,
                include_in_quote=False,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    exterior_plaster = _item_row_named(rows, "\u5916\u5899\u6279\u5d4c")
    exterior_repair = _item_row_named(rows, "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865")
    assert exterior_plaster[3] == 9.0
    assert exterior_plaster[9] == "\u81ea\u52a8\u6c47\u603b"
    assert exterior_plaster[12] == "\u9009\u5b9a\u5916\u5899\u51c0\u9762\u79ef\u6c47\u603b"
    assert exterior_plaster[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert "\u6263\u9664\u5916\u5899\u6d1e\u53e3" in exterior_plaster[14]
    assert "\u786e\u8ba4\u5916\u5899\u6279\u5d4c\u65bd\u5de5\u9762" in exterior_plaster[14]
    assert exterior_repair[3] == 0
    assert exterior_repair[9] == "\u81ea\u52a8\u6c47\u603b"
    assert "\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165" in exterior_repair[14]


def test_export_residential_quote_auto_fills_exterior_repair_from_marked_area(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_exterior_summary_items=True)
    result = QuantityResult(
        project_name="Exterior Repair Quote Demo",
        rows=[],
        exceptions=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="repair-outline",
                kind=ConstructionKind.EXTERIOR_REPAIR,
                length=10.0,
                effective_height=None,
                height_defaulted=False,
                area=6.0,
            ),
            ConstructionQuantityDetail(
                id="repair-line",
                kind=ConstructionKind.EXTERIOR_REPAIR,
                length=2.0,
                effective_height=1.5,
                height_defaulted=False,
                area=3.0,
            ),
        ],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    exterior_repair = _item_row_named(rows, "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865")
    assert exterior_repair[3] == 9.0
    assert exterior_repair[9] == "\u81ea\u52a8\u6c47\u603b"
    assert exterior_repair[12] == "\u5916\u5899\u4fee\u8865\u8303\u56f4\u9762\u79ef\u6c47\u603b"
    assert exterior_repair[13] == "\u81ea\u52a8\u751f\u6210"


def test_export_residential_quote_uses_zero_for_optional_missing_exterior_repair_and_balcony_doors(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(
        template_path,
        include_advanced_summary_items=True,
        include_balcony_sliding_items=True,
        include_exterior_summary_items=True,
    )
    result = QuantityResult(
        project_name="Defaults Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                door_details=[
                    DoorQuantityDetail(
                        id="kitchen-slide",
                        room_id="kitchen",
                        width=1.6,
                        height=2.1,
                        effective_height=2.1,
                        area=3.36,
                    )
                ],
            )
        ],
        exterior_rows=[],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert _item_row_named(rows, "\u5916\u5899\u6279\u5d4c")[3] == 77
    exterior_repair = _item_row_named(rows, "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865")
    balcony_door = _item_row_named(rows, "\u9633\u53f0\u63a8\u62c9\u95e8")
    balcony_trim = _item_row_named(rows, "\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957")
    assert exterior_repair[3] == 0
    assert exterior_repair[9] == "\u81ea\u52a8\u6c47\u603b"
    assert "\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165" in exterior_repair[14]
    assert balcony_door[3] == 0
    assert balcony_door[9] == "\u81ea\u52a8\u6c47\u603b"
    assert balcony_trim[3] == 0
    assert balcony_trim[9] == "\u81ea\u52a8\u6c47\u603b"


def test_export_residential_quote_auto_fills_construction_marker_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_construction_items=True)
    result = QuantityResult(
        project_name="Construction Quote",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=10.0, net_wall_area=20.0),
            _quantity_row("bedroom", "\u5367\u5ba4", floor_area=5.0, net_wall_area=18.0),
        ],
        building_area=116.0,
        construction_details=[
            ConstructionQuantityDetail(
                id="demo-1",
                kind=ConstructionKind.DEMO_WALL,
                length=3.0,
                effective_height=2.8,
                height_defaulted=True,
                area=8.4,
            ),
            ConstructionQuantityDetail(
                id="new-120",
                kind=ConstructionKind.NEW_WALL,
                length=2.0,
                height=3.0,
                effective_height=3.0,
                thickness=0.12,
                area=6.0,
            ),
            ConstructionQuantityDetail(
                id="new-240",
                kind=ConstructionKind.NEW_WALL,
                length=1.5,
                height=3.0,
                effective_height=3.0,
                thickness=0.24,
                area=4.5,
            ),
            ConstructionQuantityDetail(id="lintel-1", kind=ConstructionKind.LINTEL, count=1),
            ConstructionQuantityDetail(id="hole-1", kind=ConstructionKind.LINTEL_HOLE, count=1),
            ConstructionQuantityDetail(id="hole-2", kind=ConstructionKind.LINTEL_HOLE, count=1),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    demo = _item_row_named(rows, "\u62c6\u6539\u53ca\u62c6\u5899")
    wall_120 = _item_row_named(rows, "\u780c120\u539a\u7816\u5899")
    wall_240 = _item_row_named(rows, "\u780c240\u539a\u7816\u5899")
    lintel = _item_row_named(rows, "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881")
    hole = _item_row_named(rows, "\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54")
    assert demo[3] == 8.4
    assert demo[9] == "\u81ea\u52a8\u6c47\u603b"
    assert demo[12] == "\u62c6\u6539\u5899\u9762\u79ef\u6c47\u603b"
    assert demo[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert wall_120[3] == 6.0
    assert wall_240[3] == 4.5
    assert lintel[3] == 1
    assert lintel[9] == "\u81ea\u52a8\u6c47\u603b"
    assert lintel[12] == "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert hole[3] == 12
    assert hole[9] == "\u81ea\u52a8\u6c47\u603b"
    assert hole[12] == "\u5efa\u7b51\u9762\u79ef\u768410%\u53d6\u6574"
    assert hole[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"


def test_export_residential_quote_auto_fills_background_wall_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_background_wall_items=True)
    result = QuantityResult(
        project_name="Background Wall Quote",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="background-1",
                kind=ConstructionKind.BACKGROUND_WALL,
                length=3.0,
                height=1.8,
                effective_height=1.8,
                area=5.4,
            ),
            ConstructionQuantityDetail(
                id="background-2",
                kind=ConstructionKind.BACKGROUND_WALL,
                length=2.0,
                effective_height=2.8,
                height_defaulted=True,
                area=5.6,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    background = _item_row_named(rows, "\u80cc\u666f\u5899")
    assert background[3] == 11.0
    assert background[9] == "\u81ea\u52a8\u6c47\u603b"
    assert background[12] == "\u80cc\u666f\u5899\u9762\u79ef\u6c47\u603b"
    assert background[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert background[14] == "\u80cc\u666f\u5899\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838"


def test_export_residential_quote_defaults_missing_lintel_markers_to_zero(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_construction_items=True)
    result = QuantityResult(project_name="No Lintels", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    lintel = _item_row_named(rows, "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881")
    assert lintel[3] == 0
    assert lintel[9] == "\u81ea\u52a8\u6c47\u603b"
    assert lintel[12] == "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert lintel[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert lintel[14] == "\u672a\u8bc6\u522bQUOTE_LINTEL\u8fc7\u6881\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u8fc7\u6881\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199"


def test_export_residential_quote_defaults_missing_background_wall_markers_to_zero(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_background_wall_items=True)
    result = QuantityResult(project_name="No Background Wall", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    background = _item_row_named(rows, "\u80cc\u666f\u5899")
    assert background[3] == 0
    assert background[9] == "\u81ea\u52a8\u6c47\u603b"
    assert background[12] == "\u80cc\u666f\u5899\u9762\u79ef\u6c47\u603b"
    assert background[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert background[14] == "\u672a\u8bc6\u522bQUOTE_BACKGROUND_WALL\u80cc\u666f\u5899\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u80cc\u666f\u5899\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199"


def test_export_residential_quote_defaults_entry_door_to_zero(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_entry_door_items=True)
    result = QuantityResult(project_name="No Entry Door", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    entry_door = _item_row_named(rows, "\u5165\u6237\u95e8")
    assert entry_door[3] == 0
    assert entry_door[9] == "\u81ea\u52a8\u6c47\u603b"
    assert entry_door[12] == "\u5165\u6237\u95e8\u9ed8\u8ba40"
    assert entry_door[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert entry_door[14] == "\u5165\u6237\u95e8\u7ecf\u5e38\u4e0d\u5728\u62a5\u4ef7\u8303\u56f4\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u66f4\u6362\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199"


def test_export_residential_quote_counts_shower_glass_markers(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_shower_glass_items=True)
    result = QuantityResult(
        project_name="Shower Glass Quote",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(id="shower-glass-1", kind=ConstructionKind.SHOWER_GLASS, count=1),
            ConstructionQuantityDetail(id="shower-glass-2", kind=ConstructionKind.SHOWER_GLASS, count=1),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    shower_glass = _item_row_named(rows, "\u73bb\u7483\u6dcb\u6d74\u623f")
    assert shower_glass[3] == 2
    assert shower_glass[9] == "\u81ea\u52a8\u6c47\u603b"
    assert shower_glass[12] == "\u6dcb\u6d74\u623f\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert shower_glass[13] == "\u81ea\u52a8\u751f\u6210"


def test_export_residential_quote_defaults_missing_shower_glass_markers_to_zero(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_shower_glass_items=True)
    result = QuantityResult(project_name="No Shower Glass", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    shower_glass = _item_row_named(rows, "\u73bb\u7483\u6dcb\u6d74\u623f")
    assert shower_glass[3] == 0
    assert shower_glass[9] == "\u81ea\u52a8\u6c47\u603b"
    assert shower_glass[12] == "\u6dcb\u6d74\u623f\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert shower_glass[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert shower_glass[14] == "\u672a\u8bc6\u522bQUOTE_SHOWER_GLASS\u6dcb\u6d74\u623f\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff0c\u907f\u514d\u4e0e\u6dcb\u6d74\u9694\u65ad\u91cd\u590d\u62a5\u4ef7"


def test_export_residential_quote_counts_squat_toilet_markers(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_squat_toilet_items=True)
    result = QuantityResult(
        project_name="Squat Toilet Quote",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(id="squat-toilet-1", kind=ConstructionKind.SQUAT_TOILET, count=1),
            ConstructionQuantityDetail(id="squat-toilet-2", kind=ConstructionKind.SQUAT_TOILET, count=1),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    squat_toilet = _item_row_named(rows, "\u8e72\u5751")
    assert squat_toilet[3] == 2
    assert squat_toilet[9] == "\u81ea\u52a8\u6c47\u603b"
    assert squat_toilet[12] == "\u8e72\u5751\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert squat_toilet[13] == "\u81ea\u52a8\u751f\u6210"


def test_export_residential_quote_defaults_missing_squat_toilet_markers_to_zero(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_squat_toilet_items=True)
    result = QuantityResult(project_name="No Squat Toilet", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    squat_toilet = _item_row_named(rows, "\u8e72\u5751")
    assert squat_toilet[3] == 0
    assert squat_toilet[9] == "\u81ea\u52a8\u6c47\u603b"
    assert squat_toilet[12] == "\u8e72\u5751\u6807\u8bc6\u6570\u91cf\u6c47\u603b"
    assert squat_toilet[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert squat_toilet[14] == "\u672a\u8bc6\u522bQUOTE_SQUAT_TOILET\u8e72\u5751\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff0c\u907f\u514d\u4e0e\u9a6c\u6876\u91cd\u590d\u62a5\u4ef7"


def test_export_residential_quote_auto_fills_pipe_marker_items(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_pipe_items=True)
    result = QuantityResult(
        project_name="Pipe Quote",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="pipe-insulation-1",
                kind=ConstructionKind.PIPE_INSULATION,
                length=2.4,
                height=2.4,
                effective_height=2.4,
                count=1,
            ),
            ConstructionQuantityDetail(
                id="pipe-wrap-1",
                kind=ConstructionKind.PIPE_WRAP,
                length=2.8,
                effective_height=2.8,
                height_defaulted=True,
                count=1,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    insulation = _item_row_named(rows, "\u53a8\u623f\u3001\u536b\u751f\u95f4\u6392\u6c61\u7ba1\u5305\u9694\u97f3\u68c9")
    pipe_wrap = _item_row_named(rows, "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053(\u5355\u7ba1)")
    assert insulation[3] == 2.4
    assert insulation[9] == "\u81ea\u52a8\u6c47\u603b"
    assert insulation[12] == "\u6392\u6c61\u7ba1\u9694\u97f3\u68c9\u7acb\u7ba1\u957f\u5ea6\u6c47\u603b"
    assert insulation[13] == "\u81ea\u52a8\u751f\u6210"
    assert pipe_wrap[3] == 2.8
    assert pipe_wrap[9] == "\u81ea\u52a8\u6c47\u603b"
    assert pipe_wrap[12] == "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053\u7acb\u7ba1\u957f\u5ea6\u6c47\u603b"
    assert pipe_wrap[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert pipe_wrap[14] == "\u7ba1\u9053\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838"


def test_export_residential_quote_keeps_percent_count_default_without_building_area(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_construction_items=True)
    result = QuantityResult(
        project_name="Missing Building Area",
        rows=[
            _quantity_row("living", "\u5ba2\u5385", floor_area=80.0, net_wall_area=20.0),
            _quantity_row("bedroom", "\u5367\u5ba4", floor_area=36.0, net_wall_area=18.0),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    hole = _item_row_named(rows, "\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54")
    assert hole[3] == 108
    assert hole[9] == "\u6a21\u677f\u9ed8\u8ba4"
    assert hole[13] == "\u6309\u6a21\u677f\u751f\u6210"


def test_export_residential_quote_uses_zero_for_missing_demo_wall_markers(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_construction_items=True)
    result = QuantityResult(project_name="No Construction", rows=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    demo = _item_row_named(rows, "\u62c6\u6539\u53ca\u62c6\u5899")
    hole = _item_row_named(rows, "\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54")
    assert demo[3] == 0
    assert demo[9] == "\u81ea\u52a8\u6c47\u603b"
    assert demo[12] == "\u62c6\u6539\u5899\u9762\u79ef\u6c47\u603b"
    assert "\u6ca1\u6709\u8981\u62c6\u7684\u5899" in demo[14]
    assert hole[3] == 108
    assert hole[9] == "\u6a21\u677f\u9ed8\u8ba4"


def test_export_residential_quote_defaults_missing_new_wall_thickness_to_240(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_construction_items=True)
    result = QuantityResult(
        project_name="Construction Thickness",
        rows=[],
        construction_details=[
            ConstructionQuantityDetail(
                id="new-missing-thickness",
                kind=ConstructionKind.NEW_WALL,
                length=2.0,
                effective_height=3.0,
                area=6.0,
            ),
            ConstructionQuantityDetail(
                id="new-125",
                kind=ConstructionKind.NEW_WALL,
                length=2.0,
                effective_height=3.0,
                thickness=0.125,
                area=6.0,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    wall_120 = _item_row_named(rows, "\u780c120\u539a\u7816\u5899")
    wall_240 = _item_row_named(rows, "\u780c240\u539a\u7816\u5899")
    assert wall_120[3] == 0
    assert wall_120[9] == "\u81ea\u52a8\u6c47\u603b"
    assert wall_240[3] == 6.0
    assert wall_240[9] == "\u81ea\u52a8\u6c47\u603b"
    assert "\u672a\u586b\u5199\u539a\u5ea6\u7684\u65b0\u780c\u5899\u6309240mm\u8ba1\u7b97" in wall_240[14]


def test_export_residential_quote_uses_zero_for_missing_pipe_markers(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_pipe_items=True)
    result = QuantityResult(project_name="No Pipes", rows=[], construction_details=[], exceptions=[])

    export_residential_quote(result, template_path, output_path)

    rows = list(load_workbook(output_path, data_only=False).active.iter_rows(values_only=True))
    insulation = _item_row_named(rows, "\u53a8\u623f\u3001\u536b\u751f\u95f4\u6392\u6c61\u7ba1\u5305\u9694\u97f3\u68c9")
    pipe_wrap = _item_row_named(rows, "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053(\u5355\u7ba1)")
    assert insulation[3] == 0
    assert insulation[9] == "\u81ea\u52a8\u6c47\u603b"
    assert "\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165" in insulation[14]
    assert pipe_wrap[3] == 0
    assert pipe_wrap[9] == "\u81ea\u52a8\u6c47\u603b"
    assert "\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165" in pipe_wrap[14]


def test_export_residential_quote_keeps_kitchen_and_balcony_sliding_doors_separate(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_advanced_summary_items=True, include_balcony_sliding_items=True)
    result = QuantityResult(
        project_name="Balcony Door Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                door_details=[
                    DoorQuantityDetail(
                        id="kitchen-slide",
                        room_id="kitchen",
                        width=1.6,
                        height=2.1,
                        effective_height=2.1,
                        area=3.36,
                    )
                ],
            ),
            _quantity_row(
                "balcony",
                "\u9633\u53f0",
                floor_area=4.0,
                net_wall_area=10.0,
                door_details=[
                    DoorQuantityDetail(
                        id="balcony-slide",
                        room_id="balcony",
                        width=2.0,
                        height=2.2,
                        effective_height=2.2,
                        area=4.4,
                    )
                ],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert _item_row_named(rows, "\u53a8\u623f\u63a8\u62c9\u95e8")[3] == 3.36
    assert _item_row_named(rows, "\u53a8\u623f\u63a8\u62c9\u95e8\u53cc\u5305\u5957")[3] == 5.8
    assert _item_row_named(rows, "\u9633\u53f0\u63a8\u62c9\u95e8")[3] == 4.4
    assert _item_row_named(rows, "\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957")[3] == 6.4


def test_export_residential_quote_marks_default_inferred_rows_as_auto_generated(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path)
    result = QuantityResult(
        project_name="Default Inferred Demo",
        rows=[
            _quantity_row(
                "living",
                "\u5ba2\u5385",
                floor_area=20.0,
                net_wall_area=50.0,
                status=DataStatus.DEFAULT_INFERRED,
                exception_notes=["window_height_defaulted: Window w1 used default height 1.5"],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    living_wall_paint = _row_containing(rows, "\u5899\u9762\u4e73\u80f6\u6f06")
    assert living_wall_paint[13] == "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    assert "window_height_defaulted" in living_wall_paint[14]
    assert "default height 1.5" in living_wall_paint[14]


def test_export_residential_quote_marks_needs_review_rows_as_auto_generated_with_issue_note(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path)
    result = QuantityResult(
        project_name="Needs Review Demo",
        rows=[
            _quantity_row(
                "stair",
                "\u697c\u68af",
                floor_area=8.0,
                net_wall_area=22.0,
                status=DataStatus.NEEDS_REVIEW,
                exception_notes=["stair_special_quantity_manual: Room stair requires stair-specific quantities"],
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    stair_wall_paint = _row_containing(rows, "\u5899\u9762\u4e73\u80f6\u6f06")
    assert stair_wall_paint[13] == "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    assert "stair_special_quantity_manual" in stair_wall_paint[14]


def test_export_residential_quote_skips_excluded_rooms(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path)
    result = QuantityResult(
        project_name="Excluded Demo",
        rows=[
            _quantity_row(
                "shaft",
                "\u7535\u68af\u4e95",
                floor_area=4.0,
                net_wall_area=12.0,
                status=DataStatus.EXCLUDED,
            ),
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert _row_containing(rows, "\u7535\u68af\u4e95\u5de5\u7a0b") is None


def test_export_residential_quote_uses_one_wall_tile_variant_for_wet_rooms(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "quote.xlsx"
    _create_quote_template(template_path, include_both_wall_tile_variants=True)
    result = QuantityResult(
        project_name="Wet Room Demo",
        rows=[
            _quantity_row(
                "kitchen",
                "\u53a8\u623f",
                floor_area=6.0,
                net_wall_area=18.0,
                wall_measure_perimeter=10.0,
                window_area=1.0,
            )
        ],
        exceptions=[],
    )

    export_residential_quote(result, template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    kitchen_rows = _rows_between_section_and_subtotal(rows, "\u53a8\u623f\u5de5\u7a0b")
    tile_rows = [row for row in kitchen_rows if row[1] in {"\u5899\u9762\u8d34\u74f7\u7816(600x1200)", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)"}]
    assert len(tile_rows) == 1
    assert tile_rows[0][3] == 24.0


def _create_quote_template(
    path: Path,
    *,
    include_both_wall_tile_variants: bool = False,
    include_area_summary_items: bool = False,
    include_count_summary_items: bool = False,
    include_fixed_summary_items: bool = False,
    include_advanced_summary_items: bool = False,
    include_custom_cabinet_items: bool = False,
    include_balcony_sliding_items: bool = False,
    include_exterior_summary_items: bool = False,
    include_construction_items: bool = False,
    include_pipe_items: bool = False,
    include_split_cabinet_items: bool = False,
    include_background_wall_items: bool = False,
    include_shower_glass_items: bool = False,
    include_squat_toilet_items: bool = False,
    include_entry_door_items: bool = False,
) -> None:
    workbook = Workbook()
    half = workbook.active
    half.title = "\u534a\u5305"
    _write_quote_header(half)
    half.append(["\u4e00", "\u534a\u5305\u5de5\u7a0b"])
    half.append([1, "\u534a\u5305\u9879\u76ee", "M2", 1, 2, 3, 4, "=D6*(E6+F6+G6)", "\u4e0d\u5e94\u8bfb\u53d6"])

    fitout = workbook.create_sheet("\u6574\u88c5")
    _write_quote_header(fitout)
    fitout.append(["\u4e00", "\u5ba2\u5385\u5de5\u7a0b"])
    fitout.append([1, "\u9876\u9762\u6279\u5d4c", "M2", 26.8, 0, 15, 10, None, "\u9876\u9762\u8bf4\u660e"])
    fitout.append([2, "\u9876\u9762\u4e73\u80f6\u6f06", "M2", 26.8, 10, 0, 10, None, "\u9876\u6f06\u8bf4\u660e"])
    fitout.append([3, "\u5899\u9762\u754c\u9762\u5242\u5904\u7406", "M2", 52.15, 0, 4, 3, None, "\u754c\u9762\u5242\u8bf4\u660e"])
    fitout.append([4, "\u5899\u9762\u6279\u5d4c", "M2", 52.15, 0, 15, 10, None, "\u5899\u6279\u8bf4\u660e"])
    fitout.append([5, "\u5899\u9762\u4e73\u80f6\u6f06", "M2", 52.15, 10, 0, 10, None, "\u5899\u6f06\u8bf4\u660e"])
    fitout.append([6, "\u5730\u9762\u7816\u94fa\u8d34(750X1500)", "M2", 26.8, 0, 36, 60, None, "\u5730\u7816\u8bf4\u660e"])
    fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H6:H11)"])
    fitout.append(["\u4e8c", "\u53a8\u623f\u5de5\u7a0b"])
    fitout.append([1, "\u5730\u9762\u627e\u5e73", "M2", 2.6, 0, 26, 30, None, "\u627e\u5e73\u8bf4\u660e"])
    fitout.append([2, "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406", "M2", 6.6, 28, 10.5, 13, None, "\u9632\u6c34\u8bf4\u660e"])
    fitout.append([3, "\u5899\u9762\u8d34\u74f7\u7816(600X1200)", "M2", 17.7, 0, 40, 60, None, "\u5899\u7816\u8bf4\u660e"])
    fitout.append([4, "\u5730\u9762\u7816\u94fa\u8d34(750X1500)", "M2", 6.6, 0, 36, 60, None, "\u5730\u7816\u8bf4\u660e"])
    if include_both_wall_tile_variants:
        fitout.append([5, "\u5899\u9762\u8d34\u74f7\u7816(600x1200)", "M2", 17.7, 0, 40, 60, None, "\u5899\u7816\u8bf4\u660e"])
    fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H14:H17)"])
    fitout.append(["\u4e09", "\u5176\u4ed6\u5de5\u7a0b"])
    fitout.append([1, "\u5783\u573e\u6e05\u8fd0\u8d39", "M2", 99, 0, 0, 10, None, "\u4eba\u5de5\u6e05\u8fd0"])
    fitout.append([2, "\u7a97\u5e18", "\u5957", 1, 0, 0, 10, None, "\u7a97\u5e18"])
    if include_area_summary_items:
        fitout.append([3, "\u5730\u9762\u7816\u73b0\u573a\u7ef4\u62a4\u8d39", "M2", 99, 0, 0, 10, None, "\u5730\u7816\u7ef4\u62a4"])
        fitout.append([4, "\u5f3a\u7535\u5e03\u7ebf", "M2", 99, 0, 0, 10, None, "\u5f3a\u7535"])
        fitout.append([5, "\u7f8e\u7f1d", "M2", 99, 0, 0, 10, None, "\u7f8e\u7f1d"])
    if include_custom_cabinet_items:
        fitout.append([6, "\u5168\u5c4b\u5b9a\u5236", "M2", 99, 0, 0, 10, None, "\u5168\u5c4b\u5b9a\u5236"])
        fitout.append([7, "\u6a71\u67dc", "M", 99, 0, 0, 10, None, "\u6a71\u67dc"])
    if include_background_wall_items:
        fitout.append([8, "\u80cc\u666f\u5899", "M2", 60, 0, 0, 10, None, "\u80cc\u666f\u5899"])
    if include_shower_glass_items:
        fitout.append([9, "\u73bb\u7483\u6dcb\u6d74\u623f", "\u4e2a", 6, 0, 0, 10, None, "\u73bb\u7483\u6dcb\u6d74\u623f"])
    if include_squat_toilet_items:
        fitout.append([10, "\u8e72\u5751", "\u4e2a", 1, 0, 0, 10, None, "\u8e72\u5751"])
    if include_entry_door_items:
        fitout.append([11, "\u5165\u6237\u95e8", "\u6a18", 1, 0, 0, 10, None, "\u5165\u6237\u95e8"])
    if include_split_cabinet_items:
        fitout.append([6, "\u5730\u67dc", "M", 99, 0, 0, 10, None, "\u5730\u67dc"])
        fitout.append([7, "\u540a\u67dc", "M", 99, 0, 0, 10, None, "\u540a\u67dc"])
        fitout.append([8, "\u6a71\u67dc", "M", 99, 0, 0, 10, None, "\u6a71\u67dc"])
    fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H20:H20)"])
    if include_count_summary_items:
        fitout.append(["\u56db", "\u5ba4\u5185\u95e8"])
        fitout.append([1, "\u5ba4\u5185\u95e8", "\u6a18", 99, 0, 0, 10, None, "\u95e8"])
        fitout.append([2, "\u94dd\u5408\u91d1\u5c01\u95e8\u7a97", "M2", 99, 0, 0, 10, None, "\u5c01\u7a97"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H22:H23)"])
        fitout.append(["\u4e94", "\u536b\u6d74"])
        fitout.append([1, "\u6d74\u5ba4\u67dc", "\u5957", 99, 0, 0, 10, None, "\u6d74\u5ba4\u67dc"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H25:H25)"])
        fitout.append(["\u516d", "\u5176\u4ed6\uff08\u7a97\u5e18\u3001\u7f8e\u7f1d\u3001\u7a97\u53f0\u77f3\u7b49\uff09"])
        fitout.append([1, "\u7a97\u53f0\u77f3", "\u5957", 99, 0, 0, 10, None, "\u7a97\u53f0\u77f3"])
        fitout.append([2, "\u623f\u95f4\u6210\u54c1\u4fdd\u62a4", "\u95f4", 99, 0, 0, 10, None, "\u6210\u54c1\u4fdd\u62a4"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H27:H28)"])
    if include_fixed_summary_items:
        fitout.append(["\u56db", "\u96c6\u6210\u540a\u9876\u3001\u536b\u6d74\u3001\u5168\u5c4b\u5f00\u5173\u706f\u9970"])
        fitout.append([1, "\u5168\u5c4b\u63d2\u5ea7\u5f00\u5173", "\u5957", 99, 0, 0, 10, None, "\u5f00\u5173"])
        fitout.append([2, "\u5168\u5c4b\u706f\u9970", "\u5957", 99, 0, 0, 10, None, "\u706f\u9970"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H22:H23)"])
        fitout.append(["\u4e94", "\u5176\u4ed6\uff08\u7a97\u5e18\u3001\u7f8e\u7f1d\u3001\u7a97\u53f0\u77f3\u7b49\uff09"])
        fitout.append([1, "\u5168\u5c4b\u4fdd\u6d01", "\u5957", 99, 0, 0, 10, None, "\u4fdd\u6d01"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H25:H25)"])
    if include_advanced_summary_items:
        fitout.append(["\u56db", "\u4e3b\u6750\u9879\u76ee"])
        fitout.append([1, "\u5730\u9762\u74f7\u7816", "\u7247", 99, 0, 0, 10, None, "\u5730\u7816(750X1500)"])
        fitout.append([2, "\u5899\u9762\u74f7\u7816", "\u7247", 99, 0, 0, 10, None, "\u5899\u7816(600x1200)"])
        fitout.append([3, "\u74f7\u7816\u52a0\u5de5\u8d39", "M", 99, 0, 0, 10, None, "\u74f7\u7816\u52a0\u5de5"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H22:H24)"])
        fitout.append(["\u4e94", "\u5ba4\u5185\u95e8"])
        fitout.append([1, "\u5ba4\u5185\u95e8", "\u6a18", 99, 0, 0, 10, None, "\u95e8"])
        fitout.append([2, "\u53a8\u623f\u63a8\u62c9\u95e8", "M2", 99, 0, 0, 10, None, "\u63a8\u62c9\u95e8"])
        fitout.append([3, "\u53a8\u623f\u63a8\u62c9\u95e8\u53cc\u5305\u5957", "M", 99, 0, 0, 10, None, "\u63a8\u62c9\u95e8\u53cc\u5305\u5957"])
        if include_balcony_sliding_items:
            fitout.append([4, "\u9633\u53f0\u63a8\u62c9\u95e8", "M2", 66, 0, 0, 10, None, "\u9633\u53f0\u63a8\u62c9\u95e8"])
            fitout.append([5, "\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957", "M", 55, 0, 0, 10, None, "\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H26:H28)"])
        fitout.append(["\u516d", "\u96c6\u6210\u540a\u9876\u3001\u536b\u6d74\u3001\u5168\u5c4b\u5f00\u5173\u706f\u9970"])
        fitout.append([1, "\u6dcb\u6d74\u9694\u65ad", "\u5957", 99, 0, 0, 10, None, "\u6dcb\u6d74\u9694\u65ad"])
        fitout.append([2, "\u73bb\u7483\u6dcb\u6d74\u623f", "\u5957", 99, 0, 0, 10, None, "\u73bb\u7483\u6dcb\u6d74\u623f"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H28:H29)"])
        fitout.append(["\u4e03", "\u5176\u4ed6\uff08\u7a97\u5e18\u3001\u7f8e\u7f1d\u3001\u7a97\u53f0\u77f3\u7b49\uff09"])
        fitout.append([1, "\u7a97\u5e18", "\u7c73", 99, 0, 0, 10, None, "\u7a97\u5e18"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H31:H31)"])
    if include_exterior_summary_items:
        fitout.append(["\u516b", "\u5916\u5899\u5de5\u7a0b"])
        fitout.append([1, "\u5916\u5899\u6279\u5d4c", "M2", 77, 0, 0, 10, None, "\u5916\u5899\u6279\u5d4c"])
        fitout.append([2, "\u5916\u5899\u6279\u5d4c\u4ee5\u53ca\u4fee\u8865", "M2", 88, 0, 0, 10, None, "\u5916\u5899\u4fee\u8865"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H34:H35)"])
    if include_construction_items:
        fitout.append(["\u4e5d", "\u62c6\u6539\u53ca\u65b0\u5efa\u5de5\u7a0b"])
        fitout.append([1, "\u62c6\u6539\u53ca\u62c6\u5899", "M2", 93, 0, 0, 10, None, "\u62c6\u5899"])
        fitout.append([2, "\u780c120\u539a\u7816\u5899", "M2", 44.8, 0, 0, 10, None, "\u780c120\u5899"])
        fitout.append([3, "\u780c240\u539a\u7816\u5899", "M2", 64.56, 0, 0, 10, None, "\u780c240\u5899"])
        fitout.append([4, "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881", "\u652f", 15, 0, 0, 10, None, "\u8fc7\u6881"])
        fitout.append([5, "\u6253\u6df7\u51dd\u571f\u8fc7\u6881\u5b54", "\u4e2a", 108, 0, 0, 10, None, "\u8fc7\u6881\u5b54"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H37:H41)"])
    if include_pipe_items:
        fitout.append(["\u5341", "\u7ba1\u9053\u53ca\u5305\u7ba1\u5de5\u7a0b"])
        fitout.append([1, "\u53a8\u623f\u3001\u536b\u751f\u95f4\u6392\u6c61\u7ba1\u5305\u9694\u97f3\u68c9", "M", 50, 0, 0, 10, None, "\u9694\u97f3\u68c9"])
        fitout.append([2, "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053(\u5355\u7ba1)", "M", 36, 0, 0, 10, None, "\u5305\u7ba1"])
        fitout.append([None, "\u5c0f \u8ba1", None, None, None, None, None, "=SUM(H43:H44)"])
    fitout.append(["A", "\u76f4\u63a5\u8d39\u5408\u8ba1(\u4e00+\u2026...\u5341)", None, None, None, None, None, "=H12+H18+H21"])
    fitout.append(["B", "\u5de5\u7a0b\u7ba1\u7406\u8d39(D=A* 5%)", None, None, None, None, None, "=H22*0.05"])
    fitout.append(["C", "\u7a0e\u91d1E=(A+B)* 3%", None, None, None, None, None, 0])
    fitout.append(["D", "\u5de5\u7a0b\u603b\u9020\u4ef7F=(A+B+C)", None, None, None, None, None, "=SUM(H22:H24)"])
    workbook.save(path)


def _write_quote_header(sheet) -> None:
    sheet.append(["\u5de5\u7a0b(\u9884) \u7b97\u8868"])
    sheet.append(["\u540d\u79f0\uff1aDemo"])
    sheet.append([
        "\u7f16\u53f7",
        "\u9879\u76ee\u540d\u79f0",
        "\u5355\u4f4d",
        "\u6570\u91cf",
        "\u6750\u6599\u8d39(\u5143)",
        None,
        "\u4eba\u5de5\u8d39\n(\u5143)",
        "\u603b\u4ef7(\u5143)",
        "\u6750  \u6599  \u53ca  \u5de5  \u827a  \u8bf4  \u660e",
    ])
    sheet.append([None, None, None, None, "\u4e3b\u6750\n\u5355\u4ef7", "\u8f85\u6750\n\u5355\u4ef7"])


def _write_custom_quote_rules(path: Path, *, kitchen_height: float, bathroom_height: float, tile_height: float) -> None:
    path.write_text(
        json.dumps(
            {
                "wet_room_heights": {
                    "kitchen_waterproof_wall_height": kitchen_height,
                    "bathroom_waterproof_wall_height": bathroom_height,
                    "wall_tile_height": tile_height,
                },
                "floor_area_aggregate_items": ["\u5783\u573e\u6e05\u8fd0\u8d39"],
                "tile_area_aggregate_items": ["\u7f8e\u7f1d"],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _quantity_row(
    room_id: str,
    name: str,
    *,
    floor_area: float,
    net_wall_area: float,
    wall_measure_perimeter: float = 0,
    window_count: int = 0,
    window_area: float = 0,
    window_details: list[WindowQuantityDetail] | None = None,
    door_opening_count: int = 0,
    door_opening_area: float = 0,
    door_details: list[DoorQuantityDetail] | None = None,
    custom_details: list[FixtureQuantityDetail] | None = None,
    cabinet_details: list[FixtureQuantityDetail] | None = None,
    status: DataStatus = DataStatus.CONFIRMED,
    exception_notes: list[str] | None = None,
) -> QuantityRow:
    return QuantityRow(
        room_id=room_id,
        floor=None,
        room_name=name,
        space_type=SpaceType.NORMAL,
        height=2.8,
        height_mode=HeightMode.PROJECT_DEFAULT,
        floor_area=floor_area,
        floor_perimeter=0,
        wall_measure_perimeter=wall_measure_perimeter,
        open_boundary_length=0,
        gross_wall_area=net_wall_area,
        window_count=window_count,
        window_area=window_area,
        window_details=window_details or [],
        door_opening_count=door_opening_count,
        door_opening_area=door_opening_area,
        door_details=door_details or [],
        custom_details=custom_details or [],
        cabinet_details=cabinet_details or [],
        net_wall_area=net_wall_area,
        is_outdoor=False,
        include_in_floor_quantity=True,
        include_in_wall_paint_quantity=True,
        status=status,
        exception_notes=exception_notes or [],
    )


def _row_containing(rows, text: str):
    for row in rows:
        if any(text in cell for cell in row if isinstance(cell, str)):
            return row
    return None


def _item_row_named(rows, item_name: str):
    for row in rows:
        if isinstance(row[0], int) and row[1] == item_name:
            return row
    return None


def _row_containing_after(rows, section_name: str, item_name: str):
    section_seen = False
    for row in rows:
        if any(section_name in cell for cell in row if isinstance(cell, str)):
            section_seen = True
            continue
        if section_seen and any(item_name in cell for cell in row if isinstance(cell, str)):
            return row
    return None


def _rows_between_section_and_subtotal(rows, section_name: str):
    section_seen = False
    result = []
    for row in rows:
        if any(section_name in cell for cell in row if isinstance(cell, str)):
            section_seen = True
            continue
        if section_seen and row[1] == "\u5c0f \u8ba1":
            return result
        if section_seen:
            result.append(row)
    return result
