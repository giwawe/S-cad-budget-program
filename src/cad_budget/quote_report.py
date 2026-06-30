from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from cad_budget.models import ConstructionKind, QuantityResult


_HEADER_ROW = 3
_FIRST_DATA_ROW = 4
_REVIEW_STATUSES = [
    "自动生成-默认推断",
    "自动生成-异常提示",
    "按模板生成",
]
_ACTION_RULES = [
    ("补窗高", ["窗高", "默认窗高"]),
    ("补新砌墙高度/厚度", ["新砌", "墙体标识", "THICKNESS", "厚度"]),
    ("补管道/包管标识", ["QUOTE_PIPE_INSULATION", "QUOTE_PIPE_WRAP", "立管", "管道", "包管"]),
    ("补门洞/推拉门高度", ["推拉门门高", "门洞缺少高度", "默认门洞高度", "门高缺少"]),
    ("补全屋定制高度/类型", ["全屋定制", "定制项", "缺少类型"]),
    ("复核外墙修补范围", ["外墙修补", "修补范围"]),
]
_ACTION_METADATA = {
    "补窗高": {
        "priority": "high",
        "suggested_action": "在 QUOTE_WINDOW 窗块属性或窗洞轮廓 XDATA 中补充 HEIGHT；也可由预算员在报价 Excel 中复核默认窗高。",
    },
    "补新砌墙高度/厚度": {
        "priority": "high",
        "suggested_action": "在 QUOTE_NEW_WALL 标识中补充 HEIGHT 和 THICKNESS，确认新砌墙高度及 120/240mm 厚度归类。",
    },
    "补管道/包管标识": {
        "priority": "medium",
        "suggested_action": "按实际方案补充 QUOTE_PIPE_INSULATION 或 QUOTE_PIPE_WRAP 立管标识；不补标时由预算员复核默认长度。",
    },
    "补门洞/推拉门高度": {
        "priority": "medium",
        "suggested_action": "在 QUOTE_DOOR 门块属性或门洞轮廓 XDATA 中补充 HEIGHT，重点确认推拉门 2.4m 和门洞 2.2m 默认高度。",
    },
    "补全屋定制高度/类型": {
        "priority": "medium",
        "suggested_action": "在 QUOTE_CUSTOM 标识中补充 HEIGHT 和 TYPE，或由预算员复核默认 2.6m 投影面积及定制类型。",
    },
    "复核外墙修补范围": {
        "priority": "high",
        "suggested_action": "用 QUOTE_EXT_REPAIR 标识明确外墙修补范围；若不在报价范围内，由预算员保持 0 或手工填写。",
    },
}


@dataclass(frozen=True)
class QuoteReviewRow:
    excel_row: int
    number: int
    item_name: str
    quantity: Any
    quantity_source: str | None
    source_room: str | None
    basis: str | None
    review_status: str
    review_note: str | None


def generate_quote_review_report(
    input_excel: Path,
    markdown_output: Path,
    *,
    quantity_result: QuantityResult | None = None,
    json_output: Path | None = None,
) -> str:
    report_text = build_quote_review_report(input_excel, quantity_result=quantity_result)
    markdown_output.parent.mkdir(parents=True, exist_ok=True)
    markdown_output.write_text(report_text, encoding="utf-8")
    if json_output is not None:
        report_data = build_quote_review_data(input_excel, quantity_result=quantity_result)
        json_output.parent.mkdir(parents=True, exist_ok=True)
        json_output.write_text(
            json.dumps(report_data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
    return report_text


def build_quote_review_report(input_excel: Path, *, quantity_result: QuantityResult | None = None) -> str:
    rows = _load_review_rows(input_excel)
    action_contexts = _quantity_action_contexts(quantity_result)
    lines = ["# 报价复核报告", ""]
    lines.extend(_action_summary(rows, action_contexts))
    lines.append("")
    lines.extend(_status_summary(rows))
    lines.append("")
    lines.extend(_source_summary(rows))
    for status in _REVIEW_STATUSES:
        status_rows = [row for row in rows if row.review_status == status]
        if not status_rows:
            continue
        lines.append("")
        lines.append(f"## {status}")
        lines.append("")
        lines.append("| Excel行 | 编号 | 项目名称 | 数量 | 计量口径 | 复核状态 | 复核备注 |")
        lines.append("| ---: | --- | --- | ---: | --- | --- | --- |")
        for row in status_rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        str(row.excel_row),
                        str(row.number),
                        _markdown_cell(row.item_name),
                        _markdown_cell(_format_quantity(row.quantity)),
                        _markdown_cell(row.basis),
                        _markdown_cell(row.review_status),
                        _markdown_cell(row.review_note),
                    ]
                )
                + " |"
            )
    lines.append("")
    return "\n".join(lines)


def build_quote_review_data(input_excel: Path, *, quantity_result: QuantityResult | None = None) -> dict[str, Any]:
    rows = _load_review_rows(input_excel)
    action_contexts = _quantity_action_contexts(quantity_result)
    return {
        "actions": _action_items(rows, action_contexts),
        "status_counts": _status_counts(rows),
        "source_counts": _source_counts(rows),
        "rows": [_row_to_dict(row) for row in rows],
    }


def _load_review_rows(input_excel: Path) -> list[QuoteReviewRow]:
    workbook = load_workbook(input_excel, data_only=False)
    sheet = workbook.active
    headers = _header_columns(sheet)
    return _review_rows(sheet, headers)


def _header_columns(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row=_HEADER_ROW, column=column).value
        if isinstance(value, str) and value:
            headers[value] = column
    required = ["项目名称", "数量", "数量来源", "计量口径", "复核状态", "复核备注"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"Quote workbook is missing review column(s): {', '.join(missing)}")
    return headers


def _review_rows(sheet, headers: dict[str, int]) -> list[QuoteReviewRow]:
    rows: list[QuoteReviewRow] = []
    for row_index in range(_FIRST_DATA_ROW, sheet.max_row + 1):
        number = sheet.cell(row=row_index, column=1).value
        item_name = _text(sheet.cell(row=row_index, column=headers["项目名称"]).value)
        review_status = _text(sheet.cell(row=row_index, column=headers["复核状态"]).value)
        if not isinstance(number, int) or item_name is None or review_status in {None, "自动生成"}:
            continue
        rows.append(
            QuoteReviewRow(
                excel_row=row_index,
                number=number,
                item_name=item_name,
                quantity=sheet.cell(row=row_index, column=headers["数量"]).value,
                quantity_source=_text(sheet.cell(row=row_index, column=headers["数量来源"]).value),
                source_room=_text(sheet.cell(row=row_index, column=headers.get("来源空间", 0)).value)
                if "来源空间" in headers
                else None,
                basis=_text(sheet.cell(row=row_index, column=headers["计量口径"]).value),
                review_status=review_status,
                review_note=_text(sheet.cell(row=row_index, column=headers["复核备注"]).value),
            )
        )
    return rows


def _status_summary(rows: list[QuoteReviewRow]) -> list[str]:
    lines = ["## 复核状态统计", ""]
    counts = _status_counts(rows)
    for status, count in counts.items():
        lines.append(f"- {status}：{count} 行")
    return lines


def _status_counts(rows: list[QuoteReviewRow]) -> dict[str, int]:
    return {status: sum(1 for row in rows if row.review_status == status) for status in _REVIEW_STATUSES}


def _action_summary(rows: list[QuoteReviewRow], action_contexts: dict[str, list[str]]) -> list[str]:
    action_items = _action_items(rows, action_contexts)
    lines = ["## 复核行动建议", ""]
    if not action_items:
        lines.append("- 暂无明确补图建议")
        return lines
    for action in action_items:
        context = "、".join(action["objects"])
        context_text = f"；涉及对象：{context}" if context else ""
        lines.append(
            f"- {action['label']}：影响 {action['quote_row_count']} 个报价行，"
            f"涉及项目：{'、'.join(action['item_names'])}；"
            f"Excel 行 {_format_excel_rows(action['excel_rows'])}{context_text}"
        )
    return lines


def _action_items(rows: list[QuoteReviewRow], action_contexts: dict[str, list[str]]) -> list[dict[str, Any]]:
    actions: dict[str, list[QuoteReviewRow]] = {label: [] for label, _ in _ACTION_RULES}
    for row in rows:
        searchable_text = " ".join(
            part or "" for part in [row.item_name, row.basis, row.review_status, row.review_note]
        )
        for label, keywords in _ACTION_RULES:
            if any(keyword in searchable_text for keyword in keywords):
                actions[label].append(row)
                break

    actionable = [(label, action_rows) for label, action_rows in actions.items() if action_rows]
    items: list[dict[str, Any]] = []
    for label, action_rows in actionable:
        metadata = _ACTION_METADATA[label]
        items.append(
            {
                "label": label,
                "priority": metadata["priority"],
                "suggested_action": metadata["suggested_action"],
                "quote_row_count": len(action_rows),
                "item_names": _item_names(action_rows),
                "excel_rows": [row.excel_row for row in action_rows],
                "objects": action_contexts.get(label, []),
            }
        )
    return items


def _quantity_action_contexts(quantity_result: QuantityResult | None) -> dict[str, list[str]]:
    if quantity_result is None:
        return {}

    contexts: dict[str, list[str]] = {}
    window_context = _summarize_room_object_counts(
        (
            row.room_name,
            "窗高",
            sum(1 for window in row.window_details if window.height_defaulted),
            "个",
        )
        for row in quantity_result.rows
    )
    if window_context:
        contexts["补窗高"] = window_context

    door_context = _summarize_room_object_counts(
        (
            row.room_name,
            "门洞/推拉门高度",
            sum(1 for door in row.door_details if door.height_defaulted),
            "个",
        )
        for row in quantity_result.rows
    )
    if door_context:
        contexts["补门洞/推拉门高度"] = door_context

    custom_context = _summarize_room_object_counts(
        (
            row.room_name,
            "全屋定制高度/类型",
            sum(
                1
                for custom in row.custom_details
                if custom.height_defaulted or custom.fixture_type is None
            ),
            "处",
        )
        for row in quantity_result.rows
    )
    if custom_context:
        contexts["补全屋定制高度/类型"] = custom_context

    new_wall_count = sum(
        1
        for detail in quantity_result.construction_details
        if detail.kind is ConstructionKind.NEW_WALL and (detail.height_defaulted or detail.thickness is None)
    )
    if new_wall_count:
        contexts["补新砌墙高度/厚度"] = [f"新砌墙标识 {new_wall_count} 处"]

    wet_room_names = _unique_ordered(
        row.room_name for row in quantity_result.rows if "厨房" in row.room_name or "卫" in row.room_name
    )
    if wet_room_names:
        contexts["补管道/包管标识"] = [f"{'、'.join(wet_room_names)}湿区空间"]

    return contexts


def _unique_ordered(values) -> list[str]:
    result: list[str] = []
    for value in values:
        if value not in result:
            result.append(value)
    return result


def _summarize_room_object_counts(items) -> list[str] | None:
    counts: dict[tuple[str, str, str], int] = {}
    ordered_keys: list[tuple[str, str, str]] = []
    for room_name, label, count, unit in items:
        if count <= 0:
            continue
        key = (room_name, label, unit)
        if key not in counts:
            counts[key] = 0
            ordered_keys.append(key)
        counts[key] += count
    if not ordered_keys:
        return None
    return [f"{room_name}{label} {counts[key]} {unit}" for key in ordered_keys for room_name, label, unit in [key]]


def _source_summary(rows: list[QuoteReviewRow]) -> list[str]:
    counts = _source_counts(rows)
    lines = ["## 数量来源统计", ""]
    if not counts:
        lines.append("- 无需要复核的报价行")
        return lines
    for source in sorted(counts):
        lines.append(f"- {source}：{counts[source]} 行")
    return lines


def _source_counts(rows: list[QuoteReviewRow]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        source = row.quantity_source or "未标注"
        counts[source] = counts.get(source, 0) + 1
    return counts


def _row_to_dict(row: QuoteReviewRow) -> dict[str, Any]:
    return {
        "excel_row": row.excel_row,
        "number": row.number,
        "item_name": row.item_name,
        "quantity": row.quantity,
        "quantity_source": row.quantity_source,
        "source_room": row.source_room,
        "basis": row.basis,
        "review_status": row.review_status,
        "review_note": row.review_note,
    }


def _format_excel_rows(row_numbers: list[int]) -> str:
    return "、".join(str(row_number) for row_number in row_numbers)


def _format_item_names(rows: list[QuoteReviewRow]) -> str:
    return "、".join(_item_names(rows))


def _item_names(rows: list[QuoteReviewRow]) -> list[str]:
    names: list[str] = []
    for row in rows:
        if row.item_name not in names:
            names.append(row.item_name)
    return names


def _format_quantity(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _markdown_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", "<br>")
