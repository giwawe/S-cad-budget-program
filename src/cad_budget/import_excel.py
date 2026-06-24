from pathlib import Path
import json
from typing import Any

from openpyxl import load_workbook

from cad_budget.export_excel import EXTERIOR_HEADERS, HEADERS
from cad_budget.models import (
    DataStatus,
    DoorQuantityDetail,
    ExteriorQuantityRow,
    FixtureQuantityDetail,
    HeightMode,
    QuantityRow,
    QuantityResult,
    SpaceType,
    WindowQuantityDetail,
)


_FLOOR = HEADERS[0]
_ROOM_NAME = HEADERS[1]
_SPACE_TYPE = HEADERS[2]
_HEIGHT = HEADERS[3]
_FLOOR_AREA = HEADERS[4]
_FLOOR_PERIMETER = HEADERS[5]
_WALL_MEASURE_PERIMETER = HEADERS[6]
_OPEN_BOUNDARY_LENGTH = HEADERS[7]
_WINDOW_COUNT = HEADERS[9]
_WINDOW_AREA = HEADERS[10]
_DOOR_OPENING_COUNT = HEADERS[11]
_DOOR_OPENING_AREA = HEADERS[12]
_IS_OUTDOOR = HEADERS[14]
_INCLUDE_FLOOR = HEADERS[15]
_INCLUDE_WALL_PAINT = HEADERS[16]
_STATUS = HEADERS[17]
_EXCEPTION_NOTES = HEADERS[18]
_ROOM_ID = "\u7a7a\u95f4ID"
_DETAILS_JSON = "\u62a5\u4ef7\u660e\u7ec6JSON"

_EXTERIOR_FLOOR = EXTERIOR_HEADERS[0]
_EXTERIOR_WALL_ID = EXTERIOR_HEADERS[1]
_EXTERIOR_HEIGHT = EXTERIOR_HEADERS[2]
_EXTERIOR_MEASURE_LENGTH = EXTERIOR_HEADERS[3]
_EXTERIOR_OPENING_LENGTH = EXTERIOR_HEADERS[4]


def import_quantity_result(workbook_path: Path) -> QuantityResult:
    workbook = load_workbook(workbook_path, data_only=False)
    quantity_sheet = _find_sheet_with_headers(workbook, HEADERS)
    project_name = _as_text(quantity_sheet["B1"].value) or "Untitled"

    rows = [_read_quantity_row(quantity_sheet, row_index) for row_index in range(4, quantity_sheet.max_row + 1)]
    rows = [row for row in rows if row is not None]

    exterior_rows: list[ExteriorQuantityRow] = []
    exterior_sheet = _find_sheet_with_headers(workbook, EXTERIOR_HEADERS, required=False)
    if exterior_sheet is not None:
        exterior_rows = [
            row
            for row in (
                _read_exterior_row(exterior_sheet, row_index) for row_index in range(4, exterior_sheet.max_row + 1)
            )
            if row is not None
        ]

    return QuantityResult(project_name=project_name, rows=rows, exterior_rows=exterior_rows, exceptions=[])


def _find_sheet_with_headers(workbook, headers: list[str], *, required: bool = True):
    for sheet in workbook.worksheets:
        values = [sheet.cell(row=3, column=index).value for index in range(1, len(headers) + 1)]
        if values == headers:
            return sheet
    if required:
        raise ValueError("Workbook does not contain the expected quantity sheet headers.")
    return None


def _read_quantity_row(sheet, row_index: int) -> QuantityRow | None:
    values = _row_values_by_header(sheet, row_index)
    room_id = _as_text(values.get(_ROOM_ID))
    room_name = _as_text(values.get(_ROOM_NAME))
    if not room_id and not room_name:
        return None

    height = _as_float(values.get(_HEIGHT))
    wall_measure_perimeter = _as_float(values.get(_WALL_MEASURE_PERIMETER))
    window_area = _as_float(values.get(_WINDOW_AREA))
    gross_wall_area = _round_quantity(wall_measure_perimeter * height)
    net_wall_area = _round_quantity(gross_wall_area - window_area)

    exception_text = _as_text(values.get(_EXCEPTION_NOTES))
    exception_notes = _split_exception_notes(exception_text)
    details = _quote_details_from_json(values.get(_DETAILS_JSON))

    return QuantityRow(
        room_id=room_id or f"excel-row-{row_index}",
        floor=_as_text(values.get(_FLOOR)),
        room_name=room_name or "",
        space_type=SpaceType(_as_text(values.get(_SPACE_TYPE)) or SpaceType.NORMAL.value),
        height=height,
        height_mode=HeightMode.MANUAL,
        floor_area=_as_float(values.get(_FLOOR_AREA)),
        floor_perimeter=_as_float(values.get(_FLOOR_PERIMETER)),
        wall_measure_perimeter=wall_measure_perimeter,
        open_boundary_length=_as_float(values.get(_OPEN_BOUNDARY_LENGTH)),
        gross_wall_area=gross_wall_area,
        window_count=_as_int(values.get(_WINDOW_COUNT)),
        window_area=window_area,
        door_opening_count=_as_int(values.get(_DOOR_OPENING_COUNT)),
        door_opening_area=_as_float(values.get(_DOOR_OPENING_AREA)),
        net_wall_area=net_wall_area,
        is_outdoor=_as_bool(values.get(_IS_OUTDOOR)),
        include_in_floor_quantity=_as_bool(values.get(_INCLUDE_FLOOR)),
        include_in_wall_paint_quantity=_as_bool(values.get(_INCLUDE_WALL_PAINT)),
        status=DataStatus(_as_text(values.get(_STATUS)) or DataStatus.MANUALLY_EDITED.value),
        exception_notes=exception_notes,
        window_details=details["window_details"],
        door_details=details["door_details"],
        custom_details=details["custom_details"],
        cabinet_details=details["cabinet_details"],
    )


def _read_exterior_row(sheet, row_index: int) -> ExteriorQuantityRow | None:
    values = _row_values_by_header(sheet, row_index)
    exterior_wall_id = _as_text(values.get(_EXTERIOR_WALL_ID))
    if not exterior_wall_id:
        return None

    height = _as_float(values.get(_EXTERIOR_HEIGHT))
    measure_length = _as_float(values.get(_EXTERIOR_MEASURE_LENGTH))
    opening_length = _as_float(values.get(_EXTERIOR_OPENING_LENGTH))
    gross_area = _round_quantity(measure_length * height)
    net_area = _round_quantity(gross_area - opening_length * height)
    return ExteriorQuantityRow(
        exterior_wall_id=exterior_wall_id,
        floor=_as_text(values.get(_EXTERIOR_FLOOR)),
        height=height,
        measure_length=measure_length,
        opening_length=opening_length,
        gross_area=gross_area,
        net_area=net_area,
    )


def _row_values_by_header(sheet, row_index: int) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for column_index in range(1, sheet.max_column + 1):
        header = sheet.cell(row=3, column=column_index).value
        if header:
            values[str(header)] = sheet.cell(row=row_index, column=column_index).value
    return values


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _as_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(value)


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "\u662f"}


def _round_quantity(value: float) -> float:
    return round(value, 6)


def _split_exception_notes(text: str | None) -> list[str]:
    if not text:
        return []
    normalized = text.replace(";", "\uff1b")
    return [part.strip() for part in normalized.split("\uff1b") if part.strip()]


def _quote_details_from_json(value: Any) -> dict[str, list]:
    text = _as_text(value)
    if not text:
        return {
            "window_details": [],
            "door_details": [],
            "custom_details": [],
            "cabinet_details": [],
        }
    data = json.loads(text)
    if not isinstance(data, dict):
        raise ValueError("Quote detail metadata must be a JSON object.")
    return {
        "window_details": [
            WindowQuantityDetail.model_validate(item) for item in data.get("window_details", [])
        ],
        "door_details": [
            DoorQuantityDetail.model_validate(item) for item in data.get("door_details", [])
        ],
        "custom_details": [
            FixtureQuantityDetail.model_validate(item) for item in data.get("custom_details", [])
        ],
        "cabinet_details": [
            FixtureQuantityDetail.model_validate(item) for item in data.get("cabinet_details", [])
        ],
    }
