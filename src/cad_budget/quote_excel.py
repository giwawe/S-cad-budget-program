from dataclasses import dataclass, field, replace
import json
import math
from importlib import resources
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from cad_budget.models import (
    ConstructionKind,
    ConstructionQuantityDetail,
    DataStatus,
    ExteriorQuantityRow,
    FixturePricingMode,
    QuantityResult,
    QuantityRow,
)


FITOUT_SHEET_NAME = "\u6574\u88c5"
QUOTE_SHEET_NAME = "\u5546\u54c1\u623f\u6574\u88c5\u62a5\u4ef7"
QUOTE_HEADERS = [
    "\u7f16\u53f7",
    "\u9879\u76ee\u540d\u79f0",
    "\u5355\u4f4d",
    "\u6570\u91cf",
    "\u6750\u6599\u8d39(\u5143)",
    None,
    "\u4eba\u5de5\u8d39\n(\u5143)",
    "\u603b\u4ef7(\u5143)",
    "\u6750  \u6599  \u53ca  \u5de5  \u827a  \u8bf4  \u660e",
    "\u6570\u91cf\u6765\u6e90",
    "\u6765\u6e90\u7a7a\u95f4",
    "\u7a7a\u95f4ID",
    "\u8ba1\u91cf\u53e3\u5f84",
    "\u590d\u6838\u72b6\u6001",
    "\u590d\u6838\u5907\u6ce8",
]
QUOTE_SUBHEADERS = [None, None, None, None, "\u4e3b\u6750\n\u5355\u4ef7", "\u8f85\u6750\n\u5355\u4ef7"]
UNIT_PRICE_SHEET_NAME = "\u5168\u5c40\u5355\u4ef7\u8868"
UNIT_PRICE_HEADERS = ["\u9879\u76ee\u540d\u79f0", "\u5355\u4f4d", "\u4e3b\u6750\u5355\u4ef7", "\u8f85\u6750\u5355\u4ef7", "\u4eba\u5de5\u5355\u4ef7", "\u5907\u6ce8"]

_SPACE_TEMPLATE_KEYWORDS = {
    "\u5ba2\u5385",
    "\u9910\u5385",
    "\u4e3b\u5367",
    "\u6b21\u5367",
    "\u5ba2\u5367",
    "\u5367\u5ba4",
    "\u7384\u5173",
    "\u8fc7\u9053",
    "\u53a8\u623f",
    "\u536b\u751f\u95f4",
    "\u4e3b\u536b",
    "\u516c\u536b",
    "\u9732\u53f0",
    "\u9633\u53f0",
}
_NON_SPACE_SECTION_KEYWORDS = {
    "\u5176\u4ed6",
    "\u6c34\u7535",
    "\u4e3b\u6750",
    "\u5168\u5c4b\u5b9a\u5236",
    "\u5ba4\u5185\u95e8",
    "\u536b\u6d74",
    "\u5f00\u5173",
    "\u706f\u9970",
    "\u7a97\u5e18",
    "\u7f8e\u7f1d",
}
_GENERAL_ROOM_ITEMS = [
    "\u8f7b\u94a2\u9f99\u9aa8\u5e73\u9876",
    "\u9876\u9762\u6279\u5d4c",
    "\u9876\u9762\u4e73\u80f6\u6f06",
    "\u5899\u9762\u754c\u9762\u5242\u5904\u7406",
    "\u5899\u9762\u6279\u5d4c",
    "\u5899\u9762\u4e73\u80f6\u6f06",
    "\u5730\u9762\u7816\u94fa\u8d34(750X1500)",
]
_WET_ROOM_BASE_ITEMS = [
    "\u5730\u9762\u627e\u5e73",
    "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406",
    "\u5730\u9762\u7816\u94fa\u8d34(750X1500)",
]
_WALL_TILE_VARIANTS = ["\u5899\u9762\u8d34\u74f7\u7816(600x1200)", "\u5899\u9762\u8d34\u74f7\u7816(600X1200)"]
_CURTAIN_BOX_ITEMS = {"\u6697\u7a97\u5e18\u7bb1"}
_WALL_TILE_WINDOW_DEDUCTION_THRESHOLD = 3.0
_INTERIOR_DOOR_SOURCE_ROOM_KEYWORDS = {"\u5367\u5ba4", "\u4e3b\u5367", "\u6b21\u5367", "\u5ba2\u5367", "\u513f\u7ae5\u623f", "\u8001\u4eba\u623f", "\u4e66\u623f"}
_TERRACE_ITEMS = [
    "\u5730\u9762\u627e\u5e73",
    "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406",
    "\u9876\u9762\u6279\u5d4c",
    "\u9876\u9762\u4e73\u80f6\u6f06",
    "\u5899\u9762\u6279\u5d4c",
    "\u5899\u9762\u754c\u9762\u5242\u5904\u7406",
    "\u5899\u9762\u4e73\u80f6\u6f06",
    "\u5899\u9762\u8d34\u74f7\u7816(600x1200)",
    "\u5730\u9762\u7816\u94fa\u8d34(750X1500)",
]
_SECTION_NUMERALS = [
    "\u4e00",
    "\u4e8c",
    "\u4e09",
    "\u56db",
    "\u4e94",
    "\u516d",
    "\u4e03",
    "\u516b",
    "\u4e5d",
    "\u5341",
    "\u5341\u4e00",
    "\u5341\u4e8c",
    "\u5341\u4e09",
    "\u5341\u56db",
    "\u5341\u4e94",
    "\u5341\u516d",
    "\u5341\u4e03",
    "\u5341\u516b",
    "\u5341\u4e5d",
    "\u4e8c\u5341",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAD3")
_REVIEW_FILL = PatternFill("solid", fgColor="EADCF8")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)
_QUOTE_BORDER_SIDE = Side(style="thin", color="000000")
_QUOTE_BORDER = Border(
    left=_QUOTE_BORDER_SIDE,
    right=_QUOTE_BORDER_SIDE,
    top=_QUOTE_BORDER_SIDE,
    bottom=_QUOTE_BORDER_SIDE,
)
_QUOTE_FOOTER_NOTES = [
    "1、本报价不含其他管理处所增加任何费用，如果管理处所增此费用业主承担；",
    "2、施工中如有项目需增补或有设计变更，工作量及价格由甲乙双方协商认定，签字为准并进入工程总造价。",
    "3、甲方变更认可后即须交齐增补款项，不缴款者乙方有权停工，由此引起的工期延误不属乙方责任。",
    "4、所有款项甲方应缴给乙方财务（开据公司盖章票据），如因款项不是支付乙方财务产生的损失不属乙方责任。",
    "5、甲方自购材料运费自理。",
    "6、物业押金及物业管理费由甲方承担。",
    "7、本报价单不含装修期间所产生的水电费。",
    "8、施工以合同与预算为准，口头承诺无效。",
    "9、上述材料若因市场缺货,按同等档次产品品牌，若选择超过材料名单价值的材料应按市场价格补差额。",
    "10、工程竣工验收合格后，乙方向甲方办理移交及发放保修卡。不按约缴款者视甲方自动放弃保修权益，乙方并将依法追缴。",
    "11、如未办理竣工验收即入住（或更换装修钥匙）视为验收合格。",
    "12、本预算未经签字确认前，请勿带走，  否则，必须交付本报价总额3％的预算费，合同签订后计入工程款；",
    "13、本报介未含：  室内所有广告画面及字体、活动家具、音响及功放、空调、电器、室内所有监控线路铺设及设备安装；",
    "14、施工监理员按本预算进行施工，预算没有的项目不在施工范围内；",
    "15、本报价一式两份，双方各执一份，   同具法律效力。",
]


@dataclass
class QuoteTemplateItem:
    number: int
    name: str
    unit: str | None
    template_quantity: float | int | None
    main_material_price: float | int
    auxiliary_material_price: float | int
    labor_price: float | int
    description: str | None


@dataclass
class QuoteTemplateSection:
    name: str
    items: list[QuoteTemplateItem] = field(default_factory=list)


@dataclass
class QuoteTemplate:
    sheet_name: str
    sections: list[QuoteTemplateSection]


@dataclass(frozen=True)
class QuoteUnitPrice:
    item_name: str
    unit: str | None
    main_material_price: float | int
    auxiliary_material_price: float | int
    labor_price: float | int


@dataclass(frozen=True)
class QuoteUnitPriceIssue:
    row: int
    code: str
    message: str
    item_name: str | None = None
    unit: str | None = None
    field: str | None = None


@dataclass
class QuoteAggregateQuantity:
    quantity: float
    basis: str
    rooms: list[QuantityRow]
    review_status: str | None = None
    review_note: str | None = None


@dataclass
class ResidentialQuoteRules:
    kitchen_waterproof_wall_height: float
    bathroom_waterproof_wall_height: float
    wall_tile_height: float
    floor_area_aggregate_items: set[str]
    building_area_aggregate_items: set[str]
    tile_area_aggregate_items: set[str]
    room_count_aggregate_items: set[str]
    wet_room_count_aggregate_items: set[str]
    kitchen_count_aggregate_items: set[str]
    bathroom_count_aggregate_items: set[str]
    window_count_aggregate_items: set[str]
    window_area_aggregate_items: set[str]
    door_count_aggregate_items: set[str]
    door_area_aggregate_items: set[str]
    fixed_quantity_aggregate_items: dict[str, float]
    curtain_wall_length_items: set[str]
    floor_tile_piece_items: set[str]
    wall_tile_piece_items: set[str]
    tile_processing_area_items: set[str]
    interior_door_count_items: set[str]
    sliding_door_area_items: set[str]
    sliding_door_trim_length_items: set[str]
    exterior_net_area_aggregate_items: set[str]
    exterior_repair_area_items: set[str]
    demo_wall_area_items: set[str]
    new_wall_area_items_by_thickness: dict[str, float]
    lintel_count_items: set[str]
    lintel_hole_count_items: set[str]
    background_wall_area_items: set[str]
    entry_door_count_items: set[str]
    shower_glass_count_items: set[str]
    squat_toilet_count_items: set[str]
    pipe_insulation_length_items: set[str]
    pipe_wrap_length_items: set[str]
    pipe_default_length_factor: float
    building_area_percent_count_items: dict[str, float]
    custom_projected_area_items: set[str]
    cabinet_length_items: set[str]
    base_cabinet_length_items: set[str]
    wall_cabinet_length_items: set[str]
    tile_piece_loss_rate: float
    wide_door_width_threshold: float
    default_door_height: float
    default_sliding_door_height: float
    default_custom_height: float
    low_custom_height_threshold: float
    sliding_door_room_keywords: set[str]
    sliding_door_room_keywords_by_item: dict[str, set[str]]
    sliding_door_trim_room_keywords_by_item: dict[str, set[str]]
    source_label: str


def default_quote_rules_text() -> str:
    rules_path = resources.files("cad_budget").joinpath("config/residential_quote_rules.json")
    return rules_path.read_text(encoding="utf-8")


def load_default_quote_rules() -> ResidentialQuoteRules:
    return load_quote_rules(None)


def load_quote_rules(rules_path: Path | None = None) -> ResidentialQuoteRules:
    source_label = "\u5185\u7f6e\u9ed8\u8ba4\u89c4\u5219"
    if rules_path is None:
        resolved_rules_path = resources.files("cad_budget").joinpath("config/residential_quote_rules.json")
    else:
        resolved_rules_path = rules_path
        source_label = str(rules_path)
    try:
        data = json.loads(resolved_rules_path.read_text(encoding="utf-8"))
        return _quote_rules_from_dict(data, source_label)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid quote rules '{source_label}': invalid JSON: {exc}") from exc
    except KeyError as exc:
        raise ValueError(f"Invalid quote rules '{source_label}': missing field {exc.args[0]!r}") from exc
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid quote rules '{source_label}': {exc}") from exc


def _quote_rules_from_dict(data: dict[str, Any], source_label: str) -> ResidentialQuoteRules:
    wet_room_heights = data["wet_room_heights"]
    floor_area_items = data["floor_area_aggregate_items"]
    tile_area_items = data["tile_area_aggregate_items"]
    return ResidentialQuoteRules(
        kitchen_waterproof_wall_height=_required_float(wet_room_heights, "kitchen_waterproof_wall_height"),
        bathroom_waterproof_wall_height=_required_float(wet_room_heights, "bathroom_waterproof_wall_height"),
        wall_tile_height=_required_float(wet_room_heights, "wall_tile_height"),
        floor_area_aggregate_items=_required_item_set(floor_area_items, "floor_area_aggregate_items"),
        building_area_aggregate_items=_optional_item_set(data, "building_area_aggregate_items"),
        tile_area_aggregate_items=_required_item_set(tile_area_items, "tile_area_aggregate_items"),
        room_count_aggregate_items=_optional_item_set(data, "room_count_aggregate_items"),
        wet_room_count_aggregate_items=_optional_item_set(data, "wet_room_count_aggregate_items"),
        kitchen_count_aggregate_items=_optional_item_set(data, "kitchen_count_aggregate_items"),
        bathroom_count_aggregate_items=_optional_item_set(data, "bathroom_count_aggregate_items"),
        window_count_aggregate_items=_optional_item_set(data, "window_count_aggregate_items"),
        window_area_aggregate_items=_optional_item_set(data, "window_area_aggregate_items"),
        door_count_aggregate_items=_optional_item_set(data, "door_count_aggregate_items"),
        door_area_aggregate_items=_optional_item_set(data, "door_area_aggregate_items"),
        fixed_quantity_aggregate_items=_optional_quantity_map(data, "fixed_quantity_aggregate_items"),
        curtain_wall_length_items=_optional_item_set(data, "curtain_wall_length_items"),
        floor_tile_piece_items=_optional_item_set(data, "floor_tile_piece_items"),
        wall_tile_piece_items=_optional_item_set(data, "wall_tile_piece_items"),
        tile_processing_area_items=_optional_item_set(data, "tile_processing_area_items"),
        interior_door_count_items=_optional_item_set(data, "interior_door_count_items"),
        sliding_door_area_items=_optional_item_set(data, "sliding_door_area_items"),
        sliding_door_trim_length_items=_optional_item_set(data, "sliding_door_trim_length_items"),
        exterior_net_area_aggregate_items=_optional_item_set(data, "exterior_net_area_aggregate_items"),
        exterior_repair_area_items=_optional_item_set(data, "exterior_repair_area_items"),
        demo_wall_area_items=_optional_item_set(data, "demo_wall_area_items"),
        new_wall_area_items_by_thickness=_optional_quantity_map(data, "new_wall_area_items_by_thickness"),
        lintel_count_items=_optional_item_set(data, "lintel_count_items"),
        lintel_hole_count_items=_optional_item_set(data, "lintel_hole_count_items"),
        background_wall_area_items=_optional_item_set(data, "background_wall_area_items"),
        entry_door_count_items=_optional_item_set(data, "entry_door_count_items"),
        shower_glass_count_items=_optional_item_set(data, "shower_glass_count_items"),
        squat_toilet_count_items=_optional_item_set(data, "squat_toilet_count_items"),
        pipe_insulation_length_items=_optional_item_set(data, "pipe_insulation_length_items"),
        pipe_wrap_length_items=_optional_item_set(data, "pipe_wrap_length_items"),
        pipe_default_length_factor=_optional_float(data, "pipe_default_length_factor", 1.5),
        building_area_percent_count_items=_building_area_percent_count_items(data),
        custom_projected_area_items=_optional_item_set(data, "custom_projected_area_items"),
        cabinet_length_items=_optional_item_set(data, "cabinet_length_items"),
        base_cabinet_length_items=_optional_item_set(data, "base_cabinet_length_items"),
        wall_cabinet_length_items=_optional_item_set(data, "wall_cabinet_length_items"),
        tile_piece_loss_rate=_optional_float(data, "tile_piece_loss_rate", 0.05),
        wide_door_width_threshold=_optional_float(data, "wide_door_width_threshold", 1.4),
        default_door_height=_optional_float(data, "default_door_height", 2.2),
        default_sliding_door_height=_optional_float(
            data,
            "default_sliding_door_height",
            _optional_float(data, "default_door_height", 2.4) if "default_door_height" in data else 2.4,
        ),
        default_custom_height=_optional_float(data, "default_custom_height", 2.6),
        low_custom_height_threshold=_optional_float(data, "low_custom_height_threshold", 1.0),
        sliding_door_room_keywords=_optional_item_set(data, "sliding_door_room_keywords"),
        sliding_door_room_keywords_by_item=_optional_item_set_map(data, "sliding_door_room_keywords_by_item"),
        sliding_door_trim_room_keywords_by_item=_optional_item_set_map(data, "sliding_door_trim_room_keywords_by_item"),
        source_label=source_label,
    )


def _required_item_set(value: Any, key: str) -> set[str]:
    if not isinstance(value, list):
        raise ValueError(f"{key} must be a list")
    return {str(item) for item in value}


def _optional_item_set(data: dict[str, Any], key: str) -> set[str]:
    value = data.get(key, [])
    return _required_item_set(value, key)


def _optional_item_set_map(data: dict[str, Any], key: str) -> dict[str, set[str]]:
    value = data.get(key, {})
    if not isinstance(value, dict):
        raise ValueError(f"{key} must be an object")
    return {str(item_name): _required_item_set(keywords, f"{key}.{item_name}") for item_name, keywords in value.items()}


def _optional_quantity_map(data: dict[str, Any], key: str) -> dict[str, float]:
    value = data.get(key, {})
    if not isinstance(value, dict):
        raise ValueError(f"{key} must be an object")
    result: dict[str, float] = {}
    for item_name, quantity in value.items():
        try:
            result[str(item_name)] = float(quantity)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{key}.{item_name} must be a number") from exc
    return result


def _building_area_percent_count_items(data: dict[str, Any]) -> dict[str, float]:
    legacy_items = _optional_quantity_map(data, "floor_area_percent_count_items")
    current_items = _optional_quantity_map(data, "building_area_percent_count_items")
    return {**legacy_items, **current_items}


def _required_float(data: dict[str, Any], key: str) -> float:
    try:
        return float(data[key])
    except KeyError:
        raise
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{key} must be a number") from exc


def _optional_float(data: dict[str, Any], key: str, default: float) -> float:
    if key not in data:
        return default
    try:
        return float(data[key])
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{key} must be a number") from exc


def parse_quote_template(template_path: Path) -> QuoteTemplate:
    workbook = load_workbook(template_path, data_only=False)
    if FITOUT_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Quote template must contain an '\u6574\u88c5' worksheet.")
    sheet = workbook[FITOUT_SHEET_NAME]

    sections: list[QuoteTemplateSection] = []
    current_section: QuoteTemplateSection | None = None
    for row_index in range(5, sheet.max_row + 1):
        number = sheet.cell(row=row_index, column=1).value
        name = _as_text(sheet.cell(row=row_index, column=2).value)
        if not name:
            continue
        if _is_section_row(number, sheet, row_index):
            current_section = QuoteTemplateSection(name=name)
            sections.append(current_section)
            continue
        if current_section is None or not isinstance(number, int) or _is_skip_row(name):
            continue
        current_section.items.append(
            QuoteTemplateItem(
                number=number,
                name=name,
                unit=_as_text(sheet.cell(row=row_index, column=3).value),
                template_quantity=sheet.cell(row=row_index, column=4).value,
                main_material_price=_number_or_zero(sheet.cell(row=row_index, column=5).value),
                auxiliary_material_price=_number_or_zero(sheet.cell(row=row_index, column=6).value),
                labor_price=_number_or_zero(sheet.cell(row=row_index, column=7).value),
                description=_as_text(sheet.cell(row=row_index, column=9).value),
            )
        )

    return QuoteTemplate(sheet_name=FITOUT_SHEET_NAME, sections=[section for section in sections if section.items])


def export_quote_unit_price_table(input_excel: Path, output_path: Path) -> None:
    workbook = load_workbook(input_excel, data_only=True)
    sheet = workbook[FITOUT_SHEET_NAME] if FITOUT_SHEET_NAME in workbook.sheetnames else workbook.active

    prices: dict[tuple[str, str], QuoteUnitPrice] = {}
    for row_index in range(1, sheet.max_row + 1):
        number = sheet.cell(row=row_index, column=1).value
        name = _as_text(sheet.cell(row=row_index, column=2).value)
        if not isinstance(number, int) or not name or _is_skip_row(name):
            continue
        unit = _as_text(sheet.cell(row=row_index, column=3).value)
        key = _unit_price_key(name, unit)
        if key in prices:
            continue
        prices[key] = QuoteUnitPrice(
            item_name=name,
            unit=unit,
            main_material_price=_number_or_zero(sheet.cell(row=row_index, column=5).value),
            auxiliary_material_price=_number_or_zero(sheet.cell(row=row_index, column=6).value),
            labor_price=_number_or_zero(sheet.cell(row=row_index, column=7).value),
        )

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = UNIT_PRICE_SHEET_NAME
    output_sheet.append(UNIT_PRICE_HEADERS)
    for price in prices.values():
        output_sheet.append(
            [
                price.item_name,
                price.unit,
                price.main_material_price,
                price.auxiliary_material_price,
                price.labor_price,
                None,
            ]
        )
    for cell in output_sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [34, 10, 14, 14, 14, 32]
    for column_index, width in enumerate(widths, start=1):
        output_sheet.column_dimensions[output_sheet.cell(row=1, column=column_index).column_letter].width = width
    output_sheet.freeze_panes = "A2"
    output_sheet.auto_filter.ref = f"A1:F{max(output_sheet.max_row, 1)}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(output_path)


def load_quote_unit_prices(unit_prices_path: Path | None = None) -> dict[tuple[str, str], QuoteUnitPrice]:
    if unit_prices_path is None:
        return {}
    workbook = load_workbook(unit_prices_path, data_only=True)
    sheet = workbook[UNIT_PRICE_SHEET_NAME] if UNIT_PRICE_SHEET_NAME in workbook.sheetnames else workbook.active
    header_row = _find_unit_price_header_row(sheet)
    columns = _unit_price_columns(sheet, header_row)
    issues = _quote_unit_price_issues(sheet, header_row, columns)
    if issues:
        issue_summary = "; ".join(f"row {issue.row} {issue.code}: {issue.message}" for issue in issues[:10])
        raise ValueError(f"Invalid unit price workbook '{unit_prices_path}': {issue_summary}")
    prices: dict[tuple[str, str], QuoteUnitPrice] = {}
    for row_index in range(header_row + 1, sheet.max_row + 1):
        name = _as_text(sheet.cell(row=row_index, column=columns["name"]).value)
        if not name:
            continue
        unit = _as_text(sheet.cell(row=row_index, column=columns["unit"]).value)
        price = QuoteUnitPrice(
            item_name=name,
            unit=unit,
            main_material_price=_strict_price_value(sheet.cell(row=row_index, column=columns["main"]).value),
            auxiliary_material_price=_strict_price_value(sheet.cell(row=row_index, column=columns["auxiliary"]).value),
            labor_price=_strict_price_value(sheet.cell(row=row_index, column=columns["labor"]).value),
        )
        prices[_unit_price_key(name, unit)] = price
    return prices


def validate_quote_unit_price_table(unit_prices_path: Path) -> list[QuoteUnitPriceIssue]:
    workbook = load_workbook(unit_prices_path, data_only=True)
    sheet = workbook[UNIT_PRICE_SHEET_NAME] if UNIT_PRICE_SHEET_NAME in workbook.sheetnames else workbook.active
    header_row = _find_unit_price_header_row(sheet)
    columns = _unit_price_columns(sheet, header_row)
    return _quote_unit_price_issues(sheet, header_row, columns)


def export_residential_quote(
    result: QuantityResult,
    template_path: Path,
    output_path: Path,
    rules_path: Path | None = None,
    unit_prices_path: Path | None = None,
) -> None:
    rules = load_quote_rules(rules_path)
    unit_prices = load_quote_unit_prices(unit_prices_path)
    template = parse_quote_template(template_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = QUOTE_SHEET_NAME
    _write_quote_title(sheet, result.building_area)

    subtotal_rows: list[int] = []
    section_index = 0
    item_index = _build_item_index(template)
    included_rooms = [room for room in result.rows if _should_generate_room_section(room)]

    for room in included_rooms:
        section_index += 1
        item_names = _item_names_for_room(room, item_index, result.construction_details)
        items = [item_index[name] for name in item_names if name in item_index]
        if not items:
            continue
        subtotal_rows.append(
            _append_section(
                sheet,
                _section_label(section_index),
                f"{room.room_name}\u5de5\u7a0b",
                items,
                room,
                rules=rules,
                unit_prices=unit_prices,
                construction_details=result.construction_details,
            )
        )

    for section in template.sections:
        if _is_space_template_section(section.name):
            continue
        section_index += 1
        subtotal_rows.append(
            _append_section(
                sheet,
                _section_label(section_index),
                section.name,
                section.items,
                None,
                included_rooms,
                rules,
                result.exterior_rows,
                result.construction_details,
                result.building_area,
                unit_prices,
            )
        )

    _append_summary_rows(sheet, subtotal_rows)
    _append_quote_footer(sheet)
    _write_automation_summary(sheet, rules.source_label)
    _configure_sheet(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_quote_title(sheet, building_area: float | None = None) -> None:
    sheet.append(["\u5de5\u7a0b(\u9884) \u7b97\u8868"])
    area_label = "" if building_area is None else str(building_area)
    sheet.append(["\u5730\u5740\u540d\u79f0\uff1a", None, "\u5ba2\u6237\uff1a", None, None, None, f"\u88c5\u4fee\u9762\u79ef\uff1a{area_label}", None, "\u65e5\u671f\uff1a"])
    sheet.append(QUOTE_HEADERS)
    sheet.append(QUOTE_SUBHEADERS)
    for row in sheet.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[4]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=6)
    sheet.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    sheet.row_dimensions[1].height = 23
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 28.75
    sheet.row_dimensions[4].height = 14.75


def _find_unit_price_header_row(sheet) -> int:
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        values = [_as_text(sheet.cell(row=row_index, column=column).value) for column in range(1, sheet.max_column + 1)]
        if "\u9879\u76ee\u540d\u79f0" in values and "\u4e3b\u6750\u5355\u4ef7" in values:
            return row_index
    raise ValueError("Unit price workbook must contain headers: 项目名称, 单位, 主材单价, 辅材单价, 人工单价")


def _unit_price_columns(sheet, header_row: int) -> dict[str, int]:
    headers = {_as_text(sheet.cell(row=header_row, column=column).value): column for column in range(1, sheet.max_column + 1)}
    required = {
        "name": "\u9879\u76ee\u540d\u79f0",
        "unit": "\u5355\u4f4d",
        "main": "\u4e3b\u6750\u5355\u4ef7",
        "auxiliary": "\u8f85\u6750\u5355\u4ef7",
        "labor": "\u4eba\u5de5\u5355\u4ef7",
    }
    missing = [header for header in required.values() if header not in headers]
    if missing:
        raise ValueError(f"Unit price workbook missing headers: {', '.join(missing)}")
    return {key: headers[header] for key, header in required.items()}


def _quote_unit_price_issues(sheet, header_row: int, columns: dict[str, int]) -> list[QuoteUnitPriceIssue]:
    issues: list[QuoteUnitPriceIssue] = []
    seen: dict[tuple[str, str], QuoteUnitPrice] = {}
    field_labels = {
        "main": "\u4e3b\u6750\u5355\u4ef7",
        "auxiliary": "\u8f85\u6750\u5355\u4ef7",
        "labor": "\u4eba\u5de5\u5355\u4ef7",
    }
    for row_index in range(header_row + 1, sheet.max_row + 1):
        name = _as_text(sheet.cell(row=row_index, column=columns["name"]).value)
        if not name:
            continue
        unit = _as_text(sheet.cell(row=row_index, column=columns["unit"]).value)
        parsed_prices: dict[str, float | int] = {}
        for field, label in field_labels.items():
            value = sheet.cell(row=row_index, column=columns[field]).value
            parsed = _parse_unit_price_number(value)
            if parsed is None:
                code = "MISSING_PRICE" if value is None or (isinstance(value, str) and not value.strip()) else "INVALID_PRICE"
                issues.append(
                    QuoteUnitPriceIssue(
                        row=row_index,
                        code=code,
                        item_name=name,
                        unit=unit,
                        field=label,
                        message=f"{name}({unit or ''}) {label} is empty" if code == "MISSING_PRICE" else f"{name}({unit or ''}) {label} is not numeric: {value}",
                    )
                )
            else:
                parsed_prices[field] = parsed
        if len(parsed_prices) != 3:
            continue
        price = QuoteUnitPrice(
            item_name=name,
            unit=unit,
            main_material_price=parsed_prices["main"],
            auxiliary_material_price=parsed_prices["auxiliary"],
            labor_price=parsed_prices["labor"],
        )
        key = _unit_price_key(name, unit)
        existing = seen.get(key)
        if existing is not None and existing != price:
            issues.append(
                QuoteUnitPriceIssue(
                    row=row_index,
                    code="DUPLICATE_CONFLICT",
                    item_name=name,
                    unit=unit,
                    field=None,
                    message=f"{name}({unit or ''}) has conflicting duplicate unit prices",
                )
            )
        else:
            seen[key] = price
    return issues


def _parse_unit_price_number(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value if math.isfinite(value) else None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            parsed = float(text)
        except ValueError:
            return None
        if not math.isfinite(parsed):
            return None
        return int(parsed) if parsed.is_integer() else parsed
    return None


def _strict_price_value(value: Any) -> float | int:
    parsed = _parse_unit_price_number(value)
    if parsed is None:
        raise ValueError(f"Invalid unit price value: {value}")
    return parsed


def _unit_price_key(item_name: str, unit: str | None) -> tuple[str, str]:
    return (item_name.strip(), (unit or "").strip())


def _apply_unit_price_override(
    item: QuoteTemplateItem,
    unit_prices: dict[tuple[str, str], QuoteUnitPrice],
) -> QuoteTemplateItem:
    price = unit_prices.get(_unit_price_key(item.name, item.unit))
    if price is None:
        return item
    return replace(
        item,
        main_material_price=price.main_material_price,
        auxiliary_material_price=price.auxiliary_material_price,
        labor_price=price.labor_price,
    )


def _append_section(
    sheet,
    section_number: str,
    section_name: str,
    items: list[QuoteTemplateItem],
    room: QuantityRow | None,
    included_rooms: list[QuantityRow] | None = None,
    rules: ResidentialQuoteRules | None = None,
    exterior_rows: list[ExteriorQuantityRow] | None = None,
    construction_details: list[ConstructionQuantityDetail] | None = None,
    building_area: float | None = None,
    unit_prices: dict[tuple[str, str], QuoteUnitPrice] | None = None,
) -> int:
    if rules is None:
        rules = load_default_quote_rules()
    sheet.append([section_number, section_name])
    section_row = sheet.max_row
    for cell in sheet[section_row]:
        cell.fill = _SECTION_FILL
        cell.font = _BOLD_FONT

    if room is None:
        items = _expand_generic_cabinet_items(items, included_rooms or [], rules)

    first_item_row = sheet.max_row + 1
    for item_number, item in enumerate(items, start=1):
        item = _apply_unit_price_override(item, unit_prices or {})
        row_index = sheet.max_row + 1
        aggregate = (
            _aggregate_quantity_for_item(
                item,
                included_rooms or [],
                rules,
                exterior_rows or [],
                construction_details or [],
                building_area,
            )
            if room is None
            else None
        )
        if room is not None:
            quantity = _quantity_for_item(item.name, room, rules, construction_details or [])
        elif aggregate is not None:
            quantity = aggregate.quantity
        else:
            quantity = item.template_quantity
        review_values = _review_values_for_item(item.name, room, aggregate, rules)
        sheet.append(
            [
                item_number,
                item.name,
                item.unit,
                quantity,
                item.main_material_price,
                item.auxiliary_material_price,
                item.labor_price,
                f"=D{row_index}*(E{row_index}+F{row_index}+G{row_index})",
                item.description,
                *review_values,
            ]
        )
        _style_item_row(sheet, row_index)

    subtotal_row = sheet.max_row + 1
    sheet.append([None, "\u5c0f \u8ba1", None, None, None, None, None, f"=SUM(H{first_item_row}:H{subtotal_row - 1})"])
    sheet.cell(row=subtotal_row, column=2).font = _BOLD_FONT
    sheet.cell(row=subtotal_row, column=8).font = _BOLD_FONT
    return subtotal_row


def _expand_generic_cabinet_items(
    items: list[QuoteTemplateItem],
    rooms: list[QuantityRow],
    rules: ResidentialQuoteRules,
) -> list[QuoteTemplateItem]:
    if not any(room.cabinet_details for room in rooms):
        return items
    has_base_item = any(item.name in rules.base_cabinet_length_items for item in items)
    has_wall_item = any(item.name in rules.wall_cabinet_length_items for item in items)
    expanded: list[QuoteTemplateItem] = []
    for item in items:
        if item.name not in rules.cabinet_length_items:
            expanded.append(item)
            continue
        if has_base_item or has_wall_item:
            continue
        expanded.append(replace(item, name="\u5730\u67dc"))
        expanded.append(replace(item, name="\u540a\u67dc"))
    return expanded


def _append_summary_rows(sheet, subtotal_rows: list[int]) -> None:
    direct_row = sheet.max_row + 1
    direct_formula = "=" + "+".join(f"H{row}" for row in subtotal_rows) if subtotal_rows else "=0"
    sheet.append(["A", "\u76f4\u63a5\u8d39\u5408\u8ba1(\u4e00+\u2026...\u5341)", None, None, None, None, None, direct_formula])
    manage_row = sheet.max_row + 1
    sheet.append(["B", "\u5de5\u7a0b\u7ba1\u7406\u8d39(D=A* 5%)", None, None, None, None, None, f"=H{direct_row}*0.05"])
    tax_row = sheet.max_row + 1
    sheet.append(["C", "\u7a0e\u91d1E=(A+B)* 3%", None, None, None, None, None, 0])
    total_row = sheet.max_row + 1
    sheet.append(["D", "\u5de5\u7a0b\u603b\u9020\u4ef7F=(A+B+C)", None, None, None, None, None, f"=SUM(H{direct_row}:H{tax_row})"])
    for row_index in range(direct_row, total_row + 1):
        sheet.cell(row=row_index, column=1).font = _BOLD_FONT
        sheet.cell(row=row_index, column=2).font = _BOLD_FONT
        sheet.cell(row=row_index, column=8).font = _BOLD_FONT


def _append_quote_footer(sheet) -> None:
    heading_row = sheet.max_row + 1
    sheet.append(["\u7f16\u5236\u8bf4\u660e\uff1a"])
    sheet.merge_cells(start_row=heading_row, start_column=1, end_row=heading_row, end_column=9)
    sheet.cell(row=heading_row, column=1).font = _BOLD_FONT
    sheet.row_dimensions[heading_row].height = 15
    for note in _QUOTE_FOOTER_NOTES:
        row_index = sheet.max_row + 1
        sheet.append([note])
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=9)
        sheet.cell(row=row_index, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 17 if note.startswith(("9、", "10、", "11、")) else 15

    signature_row = sheet.max_row + 1
    sheet.append(["\u5ba2\u6237\u7b7e\u540d\uff1a", None, "\u8bbe\u8ba1\u5e08\uff1a", None, None, None, None, "\u62a5\u4ef7\u5458\uff1a"])
    sheet.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=2)
    sheet.merge_cells(start_row=signature_row, start_column=3, end_row=signature_row, end_column=7)
    sheet.merge_cells(start_row=signature_row, start_column=8, end_row=signature_row, end_column=9)
    sheet.row_dimensions[signature_row].height = 14
    for column_index in (1, 3, 8):
        sheet.cell(row=signature_row, column=column_index).font = _BOLD_FONT


def _style_item_row(sheet, row_index: int) -> None:
    for column_index in range(1, 16):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.alignment = Alignment(vertical="top", wrap_text=column_index in {9, 13, 15})
        if column_index in {4, 5, 6, 7, 8}:
            cell.number_format = "0.###"
        if column_index >= 10:
            cell.fill = _REVIEW_FILL


def _configure_sheet(sheet) -> None:
    widths = {
        "A": 8,
        "B": 24,
        "C": 8,
        "D": 10,
        "E": 12,
        "F": 34,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 18,
        "N": 14,
        "O": 18,
        "Q": 16,
        "R": 10,
        "S": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A3:O{sheet.max_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = f"A1:I{sheet.max_row}"
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = _QUOTE_BORDER


def _write_automation_summary(sheet, rules_source: str) -> None:
    labels = [
        "\u81ea\u52a8\u7b97\u91cf",
        "\u81ea\u52a8\u6c47\u603b",
        "\u6a21\u677f\u9ed8\u8ba4",
    ]
    counts = {label: 0 for label in labels}
    for row_index in range(5, sheet.max_row + 1):
        source = sheet.cell(row=row_index, column=10).value
        if source in counts:
            counts[source] += 1

    sheet["Q1"] = "\u62a5\u4ef7\u81ea\u52a8\u5316\u7edf\u8ba1"
    sheet["Q1"].font = _BOLD_FONT
    sheet["Q1"].fill = _SECTION_FILL
    for column_index, header in enumerate(["\u6570\u91cf\u6765\u6e90", "\u884c\u6570", "\u5360\u6bd4"], start=17):
        cell = sheet.cell(row=2, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for offset, label in enumerate(labels, start=3):
        sheet.cell(row=offset, column=17, value=label)
        sheet.cell(row=offset, column=18, value=counts[label])
        percent_cell = sheet.cell(row=offset, column=19, value=f"=R{offset}/SUM(R3:R5)")
        percent_cell.number_format = "0.0%"
        for column_index in range(17, 20):
            sheet.cell(row=offset, column=column_index).fill = _REVIEW_FILL
    sheet["Q7"] = "\u89c4\u5219\u6765\u6e90"
    sheet["R7"] = rules_source
    sheet["Q7"].font = _BOLD_FONT
    sheet["Q7"].fill = _SECTION_FILL
    sheet["R7"].fill = _REVIEW_FILL


def _build_item_index(template: QuoteTemplate) -> dict[str, QuoteTemplateItem]:
    item_index: dict[str, QuoteTemplateItem] = {}
    for section in template.sections:
        for item in section.items:
            item_index.setdefault(item.name, item)
    return item_index


def _item_names_for_room(
    room: QuantityRow,
    item_index: dict[str, QuoteTemplateItem],
    construction_details: list[ConstructionQuantityDetail] | None = None,
) -> list[str]:
    name = room.room_name
    if any(keyword in name for keyword in ["\u53a8\u623f", "\u536b", "\u6d17\u624b\u95f4", "\u536b\u751f\u95f4"]):
        names = _WET_ROOM_BASE_ITEMS.copy()
        wall_tile = next((variant for variant in _WALL_TILE_VARIANTS if variant in item_index), None)
        if wall_tile is not None:
            names.insert(2, wall_tile)
        return names
    if any(keyword in name for keyword in ["\u9732\u53f0", "\u9633\u53f0"]):
        return _TERRACE_ITEMS
    names = _GENERAL_ROOM_ITEMS.copy()
    if _explicit_wall_tile_area_for_room(room, construction_details or []) > 0:
        wall_tile = next((variant for variant in _WALL_TILE_VARIANTS if variant in item_index), None)
        if wall_tile is not None and wall_tile not in names:
            names.append(wall_tile)
    return _with_window_room_items(room, names, item_index)


def _with_window_room_items(
    room: QuantityRow,
    names: list[str],
    item_index: dict[str, QuoteTemplateItem],
) -> list[str]:
    if not room.window_details:
        return names
    result = names.copy()
    for item_name in _CURTAIN_BOX_ITEMS:
        if item_name in item_index and item_name not in result:
            result.append(item_name)
    return result


def _quantity_for_item(
    item_name: str,
    room: QuantityRow,
    rules: ResidentialQuoteRules,
    construction_details: list[ConstructionQuantityDetail] | None = None,
) -> float:
    if item_name in {
        "\u8f7b\u94a2\u9f99\u9aa8\u5e73\u9876",
        "\u9876\u9762\u6279\u5d4c",
        "\u9876\u9762\u4e73\u80f6\u6f06",
        "\u5730\u9762\u7816\u94fa\u8d34(750X1500)",
        "\u5730\u9762\u627e\u5e73",
    }:
        return _round_quantity(room.floor_area)
    if item_name == "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406":
        if "\u53a8\u623f" in room.room_name:
            return _round_quantity(room.floor_area + room.wall_measure_perimeter * rules.kitchen_waterproof_wall_height)
        return _round_quantity(room.floor_area + room.wall_measure_perimeter * rules.bathroom_waterproof_wall_height)
    if item_name in {
        "\u5899\u9762\u754c\u9762\u5242\u5904\u7406",
        "\u5899\u9762\u6279\u5d4c",
        "\u5899\u9762\u4e73\u80f6\u6f06",
    }:
        return _round_quantity(room.net_wall_area)
    if item_name in {
        "\u5899\u9762\u8d34\u74f7\u7816(600x1200)",
        "\u5899\u9762\u8d34\u74f7\u7816(600X1200)",
    }:
        if not _is_wet_room(room):
            return _explicit_wall_tile_area_for_room(room, construction_details or [])
        return _wet_wall_tile_area(room, rules)
    if item_name in _CURTAIN_BOX_ITEMS:
        return _curtain_wall_length([room])
    return 0


def _aggregate_quantity_for_item(
    item: QuoteTemplateItem,
    rooms: list[QuantityRow],
    rules: ResidentialQuoteRules,
    exterior_rows: list[ExteriorQuantityRow] | None = None,
    construction_details: list[ConstructionQuantityDetail] | None = None,
    building_area: float | None = None,
) -> QuoteAggregateQuantity | None:
    item_name = item.name
    if item_name in rules.fixed_quantity_aggregate_items:
        return QuoteAggregateQuantity(
            quantity=_round_quantity(rules.fixed_quantity_aggregate_items[item_name]),
            basis="\u56fa\u5b9a\u6570\u91cf\u6c47\u603b",
            rooms=[],
        )
    if item_name in rules.exterior_net_area_aggregate_items:
        return _exterior_net_area_aggregate(exterior_rows or [])
    if item_name in rules.exterior_repair_area_items:
        return _construction_area_aggregate(
            construction_details or [],
            ConstructionKind.EXTERIOR_REPAIR,
            "\u5916\u5899\u4fee\u8865\u8303\u56f4\u9762\u79ef\u6c47\u603b",
            zero_note="\u672a\u8bc6\u522bQUOTE_EXT_REPAIR\u5916\u5899\u4fee\u8865\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u4fee\u8865\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165",
        )
    if item_name in rules.demo_wall_area_items:
        return _construction_area_aggregate(
            construction_details or [],
            ConstructionKind.DEMO_WALL,
            "\u62c6\u6539\u5899\u9762\u79ef\u6c47\u603b",
            zero_note="\u672a\u8bc6\u522bQUOTE_DEMO_WALL\u62c6\u5899\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff0c\u8868\u793a\u6ca1\u6709\u8981\u62c6\u7684\u5899",
        )
    if item_name in rules.new_wall_area_items_by_thickness:
        return _new_wall_area_aggregate(
            construction_details or [],
            rules.new_wall_area_items_by_thickness[item_name],
        )
    if item_name in rules.lintel_count_items:
        return _construction_count_aggregate(
            construction_details or [],
            ConstructionKind.LINTEL,
            "\u7816\u5899\u95e8\u7a97\u6d1e\u8fc7\u6881\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
            zero_note="\u672a\u8bc6\u522bQUOTE_LINTEL\u8fc7\u6881\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u8fc7\u6881\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199",
            zero_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.lintel_hole_count_items:
        return _construction_count_aggregate(
            construction_details or [],
            ConstructionKind.LINTEL_HOLE,
            "\u6df7\u51dd\u571f\u8fc7\u6881\u5b54\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
        )
    if item_name in rules.background_wall_area_items:
        return _background_wall_area_aggregate(construction_details or [])
    if item_name in rules.entry_door_count_items:
        return _zero_aggregate(
            "\u5165\u6237\u95e8\u9ed8\u8ba40",
            "\u5165\u6237\u95e8\u7ecf\u5e38\u4e0d\u5728\u62a5\u4ef7\u8303\u56f4\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u66f4\u6362\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199",
            status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.shower_glass_count_items:
        return _shower_glass_count_aggregate(construction_details or [])
    if item_name in rules.squat_toilet_count_items:
        return _squat_toilet_count_aggregate(construction_details or [])
    if item_name in rules.pipe_insulation_length_items:
        return _pipe_length_aggregate(
            construction_details or [],
            ConstructionKind.PIPE_INSULATION,
            "\u6392\u6c61\u7ba1\u9694\u97f3\u68c9\u7acb\u7ba1\u957f\u5ea6\u6c47\u603b",
            "\u7ba1\u9053\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838",
            zero_note="\u672a\u8bc6\u522bQUOTE_PIPE_INSULATION\u6392\u6c61\u7ba1\u9694\u97f3\u68c9\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u8ba1\u7b97\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165",
            rooms=rooms,
            default_length_factor=rules.pipe_default_length_factor,
        )
    if item_name in rules.pipe_wrap_length_items:
        return _pipe_length_aggregate(
            construction_details or [],
            ConstructionKind.PIPE_WRAP,
            "\u5305\u4e0a/\u4e0b\u6c34\u7ba1\u9053\u7acb\u7ba1\u957f\u5ea6\u6c47\u603b",
            "\u7ba1\u9053\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838",
            zero_note="\u672a\u8bc6\u522bQUOTE_PIPE_WRAP\u5305\u7ba1\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u8ba1\u7b97\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165",
            rooms=rooms,
            default_length_factor=rules.pipe_default_length_factor,
        )
    if item_name in rules.building_area_percent_count_items:
        return _building_area_percent_count_aggregate(building_area, rules.building_area_percent_count_items[item_name])
    if not rooms:
        return None
    if item_name in rules.building_area_aggregate_items:
        return _building_area_aggregate(building_area)
    if item_name in rules.custom_projected_area_items:
        return _custom_projected_area_aggregate(rooms, rules)
    if item_name in rules.base_cabinet_length_items:
        return _cabinet_length_aggregate(rooms, fixture_type="\u5730\u67dc", basis="\u5730\u67dc\u957f\u5ea6\u6c47\u603b")
    if item_name in rules.wall_cabinet_length_items:
        return _cabinet_length_aggregate(rooms, fixture_type="\u540a\u67dc", basis="\u540a\u67dc\u957f\u5ea6\u6c47\u603b")
    if item_name in rules.cabinet_length_items:
        return _cabinet_length_aggregate(rooms)
    if item_name in rules.curtain_wall_length_items:
        curtain_quantity = _curtain_wall_length(rooms)
        curtain_rooms = [room for room in rooms if room.window_details]
        if curtain_quantity <= 0 or not curtain_rooms:
            return None
        return QuoteAggregateQuantity(
            quantity=curtain_quantity,
            basis="\u7a97\u6240\u5728\u5899\u9762\u957f\u5ea6\u6c47\u603b",
            rooms=curtain_rooms,
            review_status="\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad",
            review_note="\u7a97\u5e18\u6309\u540c\u623f\u95f4\u540c\u5899\u6bb5\u53bb\u91cd\uff0cL\u5f62\u7a97\u9700\u4eba\u5de5\u786e\u8ba4",
        )
    if item_name in rules.floor_tile_piece_items:
        tile_area_rooms = [room for room in rooms if room.include_in_floor_quantity]
        spec = _tile_spec_from_text(item.name, item.description)
        return QuoteAggregateQuantity(
            quantity=_tile_piece_count(sum(room.floor_area for room in tile_area_rooms), spec, rules.tile_piece_loss_rate),
            basis=_tile_piece_basis("\u5730\u7816\u9762\u79ef", spec, rules.tile_piece_loss_rate),
            rooms=tile_area_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.wall_tile_piece_items:
        wall_tile_rooms = [room for room in rooms if _room_has_wall_tile(room)]
        explicit_wall_tile_details = _explicit_wall_tile_details(construction_details or [], rooms)
        spec = _tile_spec_from_text(item.name, item.description)
        return QuoteAggregateQuantity(
            quantity=_tile_piece_count(
                sum(_wet_wall_tile_area(room, rules) for room in wall_tile_rooms)
                + _explicit_wall_tile_area(explicit_wall_tile_details),
                spec,
                rules.tile_piece_loss_rate,
            ),
            basis=_tile_piece_basis("\u5899\u7816\u9762\u79ef", spec, rules.tile_piece_loss_rate),
            rooms=_wall_tile_rooms_with_explicit_details(wall_tile_rooms, explicit_wall_tile_details, rooms),
            review_status=_wall_tile_review_status(wall_tile_rooms, explicit_wall_tile_details),
            review_note=_wall_tile_review_note(wall_tile_rooms, explicit_wall_tile_details),
        )
    if item_name in rules.tile_processing_area_items:
        if building_area is None or building_area <= 0:
            return None
        return QuoteAggregateQuantity(
            quantity=_round_quantity(building_area),
            basis="QUOTE_EXT_WALL\u56f4\u5408\u5efa\u7b51\u9762\u79ef",
            rooms=[],
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.interior_door_count_items:
        door_count, door_rooms = _ordinary_interior_door_count(rooms, rules)
        if door_count <= 0:
            return None
        return QuoteAggregateQuantity(
            quantity=door_count,
            basis="\u666e\u901a\u5ba4\u5185\u95e8\u6d1e\u6570\u91cf\u6c47\u603b",
            rooms=door_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.sliding_door_area_items:
        door_area, door_rooms = _sliding_door_area(rooms, rules, item_name)
        if door_area <= 0:
            if _is_optional_zero_sliding_item(item_name, rules.sliding_door_room_keywords_by_item):
                return _zero_aggregate(
                    "\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u9762\u79ef\u6c47\u603b",
                    "\u672a\u8bc6\u522b\u9633\u53f0/\u9732\u53f0\u63a8\u62c9\u95e8\u5bbd\u95e8\u6d1e\uff0c\u9ed8\u8ba40\uff1b\u5982\u6709\u9633\u53f0\u63a8\u62c9\u95e8\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165",
                )
            return None
        return QuoteAggregateQuantity(
            quantity=door_area,
            basis="\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u9762\u79ef\u6c47\u603b",
            rooms=door_rooms,
            review_status=_door_area_review_status(door_rooms),
            review_note=_door_default_note_for_rooms(door_rooms, rules, default_kind="sliding"),
        )
    if item_name in rules.sliding_door_trim_length_items:
        trim_length, door_rooms = _sliding_door_trim_length(rooms, rules, item_name)
        if trim_length <= 0:
            if _is_optional_zero_sliding_item(item_name, rules.sliding_door_trim_room_keywords_by_item):
                return _zero_aggregate(
                    "\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u5957\u7ebf\u957f\u5ea6\u6c47\u603b",
                    "\u672a\u8bc6\u522b\u9633\u53f0/\u9732\u53f0\u63a8\u62c9\u95e8\u5bbd\u95e8\u6d1e\uff0c\u9ed8\u8ba40\uff1b\u5982\u6709\u9633\u53f0\u63a8\u62c9\u95e8\u53cc\u5305\u5957\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u8f93\u5165",
                )
            return None
        return QuoteAggregateQuantity(
            quantity=trim_length,
            basis="\u5bbd\u5ea6>=1.4m\u95e8\u6d1e\u5957\u7ebf\u957f\u5ea6\u6c47\u603b",
            rooms=door_rooms,
            review_status=_door_area_review_status(door_rooms),
            review_note=_door_default_note_for_rooms(door_rooms, rules, default_kind="door_opening"),
        )
    if item_name in rules.room_count_aggregate_items:
        return QuoteAggregateQuantity(
            quantity=len(rooms),
            basis="\u6709\u6548\u7a7a\u95f4\u6570\u91cf\u6c47\u603b",
            rooms=rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.wet_room_count_aggregate_items:
        wet_rooms = [room for room in rooms if _is_wet_room(room)]
        return QuoteAggregateQuantity(
            quantity=len(wet_rooms),
            basis="\u6e7f\u533a\u7a7a\u95f4\u6570\u91cf\u6c47\u603b",
            rooms=wet_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.kitchen_count_aggregate_items:
        kitchen_rooms = [room for room in rooms if _is_kitchen(room)]
        return QuoteAggregateQuantity(
            quantity=len(kitchen_rooms),
            basis="\u53a8\u623f\u6570\u91cf\u6c47\u603b",
            rooms=kitchen_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.bathroom_count_aggregate_items:
        bathroom_rooms = [room for room in rooms if _is_bathroom(room)]
        return QuoteAggregateQuantity(
            quantity=len(bathroom_rooms),
            basis="\u536b\u751f\u95f4\u6570\u91cf\u6c47\u603b",
            rooms=bathroom_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.window_count_aggregate_items:
        window_rooms = [room for room in rooms if room.window_count > 0]
        return QuoteAggregateQuantity(
            quantity=sum(room.window_count for room in window_rooms),
            basis="\u7a97\u6570\u91cf\u6c47\u603b",
            rooms=window_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.window_area_aggregate_items:
        window_rooms = [room for room in rooms if room.window_area > 0]
        return QuoteAggregateQuantity(
            quantity=_round_quantity(sum(room.window_area for room in window_rooms)),
            basis="\u7a97\u9762\u79ef\u6c47\u603b",
            rooms=window_rooms,
            review_status=_window_area_review_status(window_rooms),
            review_note=_window_area_default_note_for_rooms(window_rooms),
        )
    if item_name in rules.door_count_aggregate_items:
        door_rooms = [room for room in rooms if room.door_opening_count > 0]
        return QuoteAggregateQuantity(
            quantity=sum(room.door_opening_count for room in door_rooms),
            basis="\u95e8\u6d1e\u6570\u91cf\u6c47\u603b",
            rooms=door_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.door_area_aggregate_items:
        door_rooms = [room for room in rooms if room.door_opening_area > 0]
        return QuoteAggregateQuantity(
            quantity=_round_quantity(sum(room.door_opening_area for room in door_rooms)),
            basis="\u95e8\u6d1e\u9762\u79ef\u6c47\u603b",
            rooms=door_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.floor_area_aggregate_items:
        floor_rooms = [room for room in rooms if room.include_in_floor_quantity]
        return QuoteAggregateQuantity(
            quantity=_round_quantity(sum(room.floor_area for room in floor_rooms)),
            basis="\u5ba4\u5185\u5730\u9762\u9762\u79ef\u6c47\u603b",
            rooms=floor_rooms,
            review_status="\u81ea\u52a8\u751f\u6210",
        )
    if item_name in rules.tile_area_aggregate_items:
        floor_rooms = [room for room in rooms if room.include_in_floor_quantity]
        wall_tile_rooms = [room for room in rooms if _room_has_wall_tile(room)]
        explicit_wall_tile_details = _explicit_wall_tile_details(construction_details or [], rooms)
        explicit_wall_tile_area = _explicit_wall_tile_area(explicit_wall_tile_details)
        return QuoteAggregateQuantity(
            quantity=_round_quantity(
                sum(room.floor_area for room in floor_rooms)
                + sum(_wet_wall_tile_area(room, rules) for room in wall_tile_rooms)
                + explicit_wall_tile_area
            ),
            basis=(
                "\u5730\u7816\u9762\u79ef+2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef+QUOTE_WALL_TILE\u663e\u5f0f\u5899\u7816\u9762\u79ef"
                if explicit_wall_tile_area > 0
                else "\u5730\u7816\u9762\u79ef+2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef"
            ),
            rooms=list({
                room.room_id: room
                for room in [*floor_rooms, *_wall_tile_rooms_with_explicit_details(wall_tile_rooms, explicit_wall_tile_details, rooms)]
            }.values()),
            review_status=_wall_tile_review_status(wall_tile_rooms, explicit_wall_tile_details),
            review_note=_wall_tile_review_note(wall_tile_rooms, explicit_wall_tile_details),
        )
    return None


def _exterior_net_area_aggregate(exterior_rows: list[ExteriorQuantityRow]) -> QuoteAggregateQuantity | None:
    included_rows = [row for row in exterior_rows if row.include_in_quote]
    quantity = _round_quantity(sum(row.gross_area for row in included_rows))
    if not included_rows or quantity <= 0:
        return None
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis="\u5916\u5899\u957f\u5ea6*\u5c42\u9ad8\u9762\u79ef\u6c47\u603b",
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _construction_area_aggregate(
    construction_details: list[ConstructionQuantityDetail],
    kind: ConstructionKind,
    basis: str,
    zero_note: str | None = None,
) -> QuoteAggregateQuantity | None:
    details = [detail for detail in construction_details if detail.kind is kind]
    quantity = _round_quantity(sum(detail.area for detail in details))
    if not details or quantity <= 0:
        if zero_note is not None:
            return _zero_aggregate(basis, zero_note)
        return None
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis=basis,
        rooms=[],
        review_status=_construction_review_status(details),
        review_note=_construction_review_note(details),
    )


def _new_wall_area_aggregate(
    construction_details: list[ConstructionQuantityDetail],
    thickness: float,
) -> QuoteAggregateQuantity | None:
    new_wall_details = [detail for detail in construction_details if detail.kind is ConstructionKind.NEW_WALL]
    if not new_wall_details:
        return None
    details = [
        detail
        for detail in new_wall_details
        if (detail.thickness is not None and abs(detail.thickness - thickness) <= 1e-6)
        or (detail.thickness is None and abs(thickness - 0.24) <= 1e-6)
    ]
    quantity = _round_quantity(sum(detail.area for detail in details))
    if not details or quantity <= 0:
        thickness_label = f"{thickness * 1000:g}mm"
        if not _has_near_miss_new_wall_thickness(new_wall_details, thickness):
            return _zero_aggregate(
                f"\u65b0\u780c{thickness_label}\u7816\u5899\u9762\u79ef\u6c47\u603b",
                status="\u81ea\u52a8\u751f\u6210",
            )
        return _zero_aggregate(
            f"\u65b0\u780c{thickness_label}\u7816\u5899\u9762\u79ef\u6c47\u603b",
            (
                f"\u6709QUOTE_NEW_WALL\u65b0\u780c\u5899\u6807\u8bc6\uff0c\u4f46\u672a\u8bc6\u522b{thickness_label}"
                "\u65b0\u780c\u5899\uff0c\u9ed8\u8ba40\uff1b\u8bf7\u786e\u8ba4THICKNESS\u662f\u5426\u586b\u5199"
                f"\u4e3a{thickness_label}/{thickness:g}m"
            ),
        )
    note = _construction_review_note(details)
    if abs(thickness - 0.24) <= 1e-6 and any(detail.thickness is None for detail in details):
        default_note = "\u672a\u586b\u5199\u539a\u5ea6\u7684\u65b0\u780c\u5899\u6309240mm\u8ba1\u7b97"
        note = f"{note}\uff1b{default_note}" if note else default_note
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis=f"\u65b0\u780c{thickness * 1000:g}mm\u7816\u5899\u9762\u79ef\u6c47\u603b",
        rooms=[],
        review_status=_construction_review_status(details),
        review_note=note,
    )


def _has_near_miss_new_wall_thickness(
    new_wall_details: list[ConstructionQuantityDetail],
    thickness: float,
) -> bool:
    return any(
        detail.thickness is not None
        and abs(detail.thickness - thickness) > 1e-6
        and abs(detail.thickness - thickness) <= 0.02
        for detail in new_wall_details
    )


def _construction_count_aggregate(
    construction_details: list[ConstructionQuantityDetail],
    kind: ConstructionKind,
    basis: str,
    zero_note: str | None = None,
    zero_status: str | None = None,
) -> QuoteAggregateQuantity | None:
    details = [detail for detail in construction_details if detail.kind is kind]
    quantity = sum(detail.count for detail in details)
    if not details or quantity <= 0:
        if zero_note is not None:
            return _zero_aggregate(basis, zero_note, status=zero_status)
        return None
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis=basis,
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _background_wall_area_aggregate(
    construction_details: list[ConstructionQuantityDetail],
) -> QuoteAggregateQuantity | None:
    details = [detail for detail in construction_details if detail.kind is ConstructionKind.BACKGROUND_WALL]
    quantity = _round_quantity(sum(detail.area for detail in details))
    if not details or quantity <= 0:
        return _zero_aggregate(
            "\u80cc\u666f\u5899\u9762\u79ef\u6c47\u603b",
            "\u672a\u8bc6\u522bQUOTE_BACKGROUND_WALL\u80cc\u666f\u5899\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff1b\u5982\u9700\u80cc\u666f\u5899\u8bf7\u8bbe\u8ba1\u5e08\u624b\u5de5\u586b\u5199",
            status="\u81ea\u52a8\u751f\u6210",
        )
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis="\u80cc\u666f\u5899\u9762\u79ef\u6c47\u603b",
        rooms=[],
        review_status=_construction_review_status(details),
        review_note=(
            "\u80cc\u666f\u5899\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838"
            if any(detail.height_defaulted for detail in details)
            else None
        ),
    )


def _shower_glass_count_aggregate(
    construction_details: list[ConstructionQuantityDetail],
) -> QuoteAggregateQuantity:
    details = [detail for detail in construction_details if detail.kind is ConstructionKind.SHOWER_GLASS]
    quantity = sum(detail.count for detail in details)
    if not details or quantity <= 0:
        return _zero_aggregate(
            "\u6dcb\u6d74\u623f\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
            "\u672a\u8bc6\u522bQUOTE_SHOWER_GLASS\u6dcb\u6d74\u623f\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff0c\u907f\u514d\u4e0e\u6dcb\u6d74\u9694\u65ad\u91cd\u590d\u62a5\u4ef7",
            status="\u81ea\u52a8\u751f\u6210",
        )
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis="\u6dcb\u6d74\u623f\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _squat_toilet_count_aggregate(
    construction_details: list[ConstructionQuantityDetail],
) -> QuoteAggregateQuantity:
    details = [detail for detail in construction_details if detail.kind is ConstructionKind.SQUAT_TOILET]
    quantity = sum(detail.count for detail in details)
    if not details or quantity <= 0:
        return _zero_aggregate(
            "\u8e72\u5751\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
            "\u672a\u8bc6\u522bQUOTE_SQUAT_TOILET\u8e72\u5751\u6807\u8bc6\uff0c\u9ed8\u8ba40\uff0c\u907f\u514d\u4e0e\u9a6c\u6876\u91cd\u590d\u62a5\u4ef7",
            status="\u81ea\u52a8\u751f\u6210",
        )
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis="\u8e72\u5751\u6807\u8bc6\u6570\u91cf\u6c47\u603b",
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _construction_length_aggregate(
    construction_details: list[ConstructionQuantityDetail],
    kind: ConstructionKind,
    basis: str,
    default_height_note: str,
    zero_note: str | None = None,
) -> QuoteAggregateQuantity | None:
    details = [detail for detail in construction_details if detail.kind is kind]
    quantity = _round_quantity(sum(detail.length for detail in details))
    if not details or quantity <= 0:
        if zero_note is not None:
            return _zero_aggregate(basis, zero_note)
        return None
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis=basis,
        rooms=[],
        review_status=_construction_review_status(details),
        review_note=default_height_note if any(detail.height_defaulted for detail in details) else None,
    )


def _pipe_length_aggregate(
    construction_details: list[ConstructionQuantityDetail],
    kind: ConstructionKind,
    basis: str,
    default_height_note: str,
    zero_note: str | None = None,
    rooms: list[QuantityRow] | None = None,
    default_length_factor: float = 1.5,
) -> QuoteAggregateQuantity | None:
    details = [detail for detail in construction_details if detail.kind is kind]
    quantity = _round_quantity(sum(detail.length for detail in details))
    if details and quantity > 0:
        return QuoteAggregateQuantity(
            quantity=quantity,
            basis=basis,
            rooms=[],
            review_status=_construction_review_status(details),
            review_note=default_height_note if any(detail.height_defaulted for detail in details) else None,
        )

    wet_rooms = [room for room in rooms or [] if _is_wet_room(room)]
    factor_label = f"{default_length_factor:g}"
    default_quantity = _round_quantity(sum(room.height for room in wet_rooms) * default_length_factor)
    if wet_rooms and default_quantity > 0:
        marker_note = zero_note or "\u672a\u8bc6\u522b\u7ba1\u9053/\u5305\u7ba1\u6807\u8bc6"
        return QuoteAggregateQuantity(
            quantity=default_quantity,
            basis=f"\u53a8\u623f/\u536b\u751f\u95f4\u5c42\u9ad8\u5408\u8ba1*{factor_label}\u9ed8\u8ba4\u957f\u5ea6",
            rooms=wet_rooms,
            review_status="\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad",
            review_note=(
                f"{marker_note}\uff1b\u672a\u6807\u6ce8\u7acb\u7ba1\u4f4d\u7f6e\uff0c"
                f"\u6682\u6309\u53a8\u623f\u53ca\u536b\u751f\u95f4\u6bcf\u4e2a\u7a7a\u95f4\u7efc\u5408\u957f\u5ea6"
                f"{factor_label}\u500d\u5c42\u9ad8\u4f30\u7b97\uff0c\u8bbe\u8ba1\u5e08\u53ef\u4fee\u6539"
            ),
        )

    if zero_note is not None:
        return _zero_aggregate(basis, zero_note)
    return None


def _building_area_percent_count_aggregate(building_area: float | None, percent: float) -> QuoteAggregateQuantity | None:
    if building_area is None or building_area <= 0:
        return None
    quantity = int(math.floor(building_area * percent + 0.5))
    if quantity <= 0:
        return None
    return QuoteAggregateQuantity(
        quantity=quantity,
        basis=f"\u5efa\u7b51\u9762\u79ef\u7684{percent * 100:g}%\u53d6\u6574",
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _building_area_aggregate(building_area: float | None) -> QuoteAggregateQuantity | None:
    if building_area is None or building_area <= 0:
        return None
    return QuoteAggregateQuantity(
        quantity=_round_quantity(building_area),
        basis="QUOTE_EXT_WALL\u56f4\u5408\u5efa\u7b51\u9762\u79ef",
        rooms=[],
        review_status="\u81ea\u52a8\u751f\u6210",
    )


def _zero_aggregate(basis: str, note: str | None = None, *, status: str | None = None) -> QuoteAggregateQuantity:
    return QuoteAggregateQuantity(
        quantity=0,
        basis=basis,
        rooms=[],
        review_status=status or ("\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad" if note else "\u81ea\u52a8\u751f\u6210"),
        review_note=note,
    )


def _construction_review_status(details: list[ConstructionQuantityDetail]) -> str:
    if any(detail.height_defaulted for detail in details):
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    return "\u81ea\u52a8\u751f\u6210"


def _construction_review_note(details: list[ConstructionQuantityDetail]) -> str | None:
    if any(detail.height_defaulted for detail in details):
        return "\u5899\u4f53\u6807\u8bc6\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u9879\u76ee/\u697c\u5c42\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97\uff0c\u9700\u590d\u6838"
    return None


def _custom_projected_area_aggregate(rooms: list[QuantityRow], rules: ResidentialQuoteRules) -> QuoteAggregateQuantity | None:
    detail_pairs = [(room, detail) for room in rooms for detail in room.custom_details]
    if not detail_pairs:
        return None
    projected_area_details = [
        detail
        for _, detail in detail_pairs
        if detail.pricing_mode == FixturePricingMode.PROJECTED_AREA
        and (
            detail.effective_height is None
            or detail.effective_height >= rules.low_custom_height_threshold
        )
    ]
    projected_area = sum(
        detail.projected_area for detail in projected_area_details
    )
    if not projected_area_details or projected_area <= 0:
        return None
    low_height_details = [
        detail
        for _, detail in detail_pairs
        if detail.pricing_mode == FixturePricingMode.LENGTH
        or (detail.effective_height is not None and detail.effective_height < rules.low_custom_height_threshold)
    ]
    notes: list[str] = []
    if any(detail.height_defaulted for _, detail in detail_pairs):
        notes.append(f"\u7f3a\u5c11\u9ad8\u5ea6\u7684\u5b9a\u5236\u9879\u6309\u9ed8\u8ba4{rules.default_custom_height:g}m\u8ba1\u7b97")
    if low_height_details:
        low_length = _round_quantity(sum(detail.length for detail in low_height_details))
        notes.append(
            f"\u9ad8\u5ea6\u5c0f\u4e8e{rules.low_custom_height_threshold:g}m\u7684\u5b9a\u5236\u9879\u6309\u957f\u5ea6\u590d\u6838\uff0c"
            f"\u5df2\u4ece\u6295\u5f71\u9762\u79ef\u6392\u9664{low_length:g}m"
        )
    if any(detail.fixture_type is None for _, detail in detail_pairs):
        notes.append("\u90e8\u5206\u5b9a\u5236\u9879\u7f3a\u5c11\u7c7b\u578b\uff0c\u9700\u590d\u6838")
    if any(detail.room_id is None and detail.room_name is None for _, detail in detail_pairs):
        notes.append("\u90e8\u5206\u5b9a\u5236\u9879\u7f3a\u5c11\u7a7a\u95f4\u5f52\u5c5e\uff0c\u9700\u590d\u6838")
    detail_rooms = list({room.room_id: room for room, _ in detail_pairs}.values())
    status = (
        "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
        if any(detail.height_defaulted for _, detail in detail_pairs) or low_height_details
        else "\u81ea\u52a8\u751f\u6210"
    )
    return QuoteAggregateQuantity(
        quantity=_round_quantity(projected_area),
        basis="\u5168\u5c4b\u5b9a\u5236\u6295\u5f71\u9762\u79ef\u6c47\u603b",
        rooms=detail_rooms,
        review_status=status,
        review_note="\uff1b".join(notes) if notes else None,
    )


def _cabinet_length_aggregate(
    rooms: list[QuantityRow],
    fixture_type: str | None = None,
    basis: str = "\u6a71\u67dc\u957f\u5ea6\u6c47\u603b",
) -> QuoteAggregateQuantity | None:
    detail_pairs = [(room, detail) for room in rooms for detail in room.cabinet_details]
    if fixture_type is not None:
        detail_pairs = [(room, detail) for room, detail in detail_pairs if detail.fixture_type == fixture_type]
    if not detail_pairs:
        return None
    notes = [] if fixture_type is not None else ["\u5730\u67dc/\u540a\u67dc\u9700\u786e\u8ba4"]
    if any(detail.fixture_type is None for _, detail in detail_pairs):
        notes.append("\u90e8\u5206\u6a71\u67dc\u7f3a\u5c11\u7c7b\u578b\uff0c\u9700\u590d\u6838")
    detail_rooms = list({room.room_id: room for room, _ in detail_pairs}.values())
    return QuoteAggregateQuantity(
        quantity=_round_quantity(sum(detail.length for _, detail in detail_pairs)),
        basis=basis,
        rooms=detail_rooms,
        review_status="\u81ea\u52a8\u751f\u6210",
        review_note="\uff1b".join(notes) if notes else None,
    )


def _review_values_for_item(
    item_name: str,
    room: QuantityRow | None,
    aggregate: QuoteAggregateQuantity | None = None,
    rules: ResidentialQuoteRules | None = None,
) -> list[str | None]:
    if aggregate is not None:
        return [
            "\u81ea\u52a8\u6c47\u603b",
            "\u5168\u5c4b",
            None,
            aggregate.basis,
            aggregate.review_status or _review_status_for_rooms(aggregate.rooms),
            aggregate.review_note
            if aggregate.review_note is not None or aggregate.review_status is not None
            else _review_note_for_rooms(aggregate.rooms, rules),
        ]
    if room is None:
        return [
            "\u6a21\u677f\u9ed8\u8ba4",
            None,
            None,
            "\u6a21\u677f\u9ed8\u8ba4\u6570\u91cf",
            "\u6309\u6a21\u677f\u751f\u6210",
            None,
        ]
    return [
        "\u81ea\u52a8\u7b97\u91cf",
        room.room_name,
        room.room_id,
        _measure_basis_for_item(item_name, room),
        _review_status_for_room(item_name, room),
        _review_note_for_room(item_name, room),
    ]


def _review_status_for_rooms(rooms: list[QuantityRow]) -> str:
    if any(room.status == DataStatus.NEEDS_REVIEW for room in rooms):
        return "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    if any(
        room.status == DataStatus.DEFAULT_INFERRED
        or any(window.height_defaulted for window in room.window_details)
        or any(door.height_defaulted for door in room.door_details)
        for room in rooms
    ):
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    return "\u81ea\u52a8\u751f\u6210"


def _review_note_for_rooms(rooms: list[QuantityRow], rules: ResidentialQuoteRules | None = None) -> str | None:
    notes: list[str] = []
    for room in rooms:
        if room.status in {DataStatus.DEFAULT_INFERRED, DataStatus.NEEDS_REVIEW}:
            notes.extend(_translated_exception_notes_for_room(room))
        if any(window.height_defaulted for window in room.window_details):
            notes.append("\u7a97\u5e18\u6309\u540c\u623f\u95f4\u540c\u5899\u6bb5\u53bb\u91cd\uff0cL\u5f62\u7a97\u9700\u4eba\u5de5\u786e\u8ba4")
        if any(door.height_defaulted for door in room.door_details):
            default_height = rules.default_door_height if rules is not None else 2.2
            notes.append(f"\u95e8\u6d1e\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u9ed8\u8ba4\u95e8\u6d1e\u9ad8\u5ea6{default_height:g}m")
    return "\uff1b".join(notes) if notes else None


def _wall_tile_review_status(
    rooms: list[QuantityRow],
    explicit_details: list[ConstructionQuantityDetail] | None = None,
) -> str:
    explicit_details = explicit_details or []
    if any(room.status == DataStatus.NEEDS_REVIEW for room in rooms):
        return "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    if any(any(window.height_defaulted for window in room.window_details) for room in rooms) or any(
        detail.height_defaulted for detail in explicit_details
    ):
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    return "\u81ea\u52a8\u751f\u6210"


def _wall_tile_review_note(
    rooms: list[QuantityRow],
    explicit_details: list[ConstructionQuantityDetail],
) -> str | None:
    notes: list[str] = []
    window_note = _window_default_note_for_rooms(rooms)
    if window_note is not None:
        notes.append(window_note)
    if any(detail.height_defaulted for detail in explicit_details):
        notes.append("QUOTE_WALL_TILE\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u6309\u5f53\u524d\u5c42\u9ad8/\u9879\u76ee\u9ed8\u8ba4\u9ad8\u5ea6\u8ba1\u7b97")
    return "\uff1b".join(notes) if notes else None


def _explicit_wall_tile_details(
    construction_details: list[ConstructionQuantityDetail],
    rooms: list[QuantityRow],
) -> list[ConstructionQuantityDetail]:
    rooms_by_id = {room.room_id: room for room in rooms}
    explicit_details: list[ConstructionQuantityDetail] = []
    for detail in construction_details:
        if detail.kind is not ConstructionKind.WALL_TILE or detail.area <= 0:
            continue
        room = rooms_by_id.get(detail.room_id or "")
        room_name = detail.room_name or (room.room_name if room is not None else "")
        if room is not None and _is_wet_room(room):
            continue
        if room is None and any(keyword in room_name for keyword in ["\u53a8\u623f", "\u536b", "\u6d17\u624b\u95f4", "\u536b\u751f\u95f4"]):
            continue
        explicit_details.append(detail)
    return explicit_details


def _explicit_wall_tile_area(details: list[ConstructionQuantityDetail]) -> float:
    return _round_quantity(sum(detail.area for detail in details))


def _wall_tile_rooms_with_explicit_details(
    wall_tile_rooms: list[QuantityRow],
    explicit_details: list[ConstructionQuantityDetail],
    rooms: list[QuantityRow],
) -> list[QuantityRow]:
    rooms_by_id = {room.room_id: room for room in rooms}
    selected = {room.room_id: room for room in wall_tile_rooms}
    for detail in explicit_details:
        if detail.room_id and detail.room_id in rooms_by_id:
            selected[detail.room_id] = rooms_by_id[detail.room_id]
    return list(selected.values())


def _window_default_note_for_rooms(rooms: list[QuantityRow]) -> str | None:
    if any(any(window.height_defaulted for window in room.window_details) for room in rooms):
        return "\u5899\u7816\u9762\u79ef\u6263\u7a97\u4f7f\u7528\u9ed8\u8ba4\u7a97\u9ad8\uff0c\u9700\u590d\u6838\u7a97\u9ad8"
    if any(any(_is_window_height_default_note(note) for note in room.exception_notes) for room in rooms):
        return "\u5899\u7816\u9762\u79ef\u6263\u7a97\u4f7f\u7528\u9ed8\u8ba4\u7a97\u9ad8\uff0c\u9700\u590d\u6838\u7a97\u9ad8"
    return None


def _window_area_review_status(rooms: list[QuantityRow]) -> str:
    if any(room.status == DataStatus.NEEDS_REVIEW for room in rooms):
        return "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    if _has_defaulted_window_height(rooms):
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    return "\u81ea\u52a8\u751f\u6210"


def _window_area_default_note_for_rooms(rooms: list[QuantityRow]) -> str | None:
    if not _has_defaulted_window_height(rooms):
        return None
    return "\u7a97\u9762\u79ef\u4f7f\u7528\u9ed8\u8ba4\u7a97\u9ad8\uff0c\u9700\u590d\u6838\u7a97\u9ad8"


def _has_defaulted_window_height(rooms: list[QuantityRow]) -> bool:
    return any(
        any(window.height_defaulted for window in room.window_details)
        or any(_is_window_height_default_note(note) for note in room.exception_notes)
        for room in rooms
    )


def _door_area_review_status(rooms: list[QuantityRow]) -> str:
    if any(room.status == DataStatus.NEEDS_REVIEW for room in rooms):
        return "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    if any(any(door.height_defaulted for door in room.door_details) for room in rooms):
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    return "\u81ea\u52a8\u751f\u6210"


def _door_default_note_for_rooms(
    rooms: list[QuantityRow],
    rules: ResidentialQuoteRules,
    *,
    default_kind: str,
) -> str | None:
    if not any(any(door.height_defaulted for door in room.door_details) for room in rooms):
        return None
    if default_kind == "sliding":
        return f"\u63a8\u62c9\u95e8\u95e8\u9ad8\u7f3a\u5c11\u65f6\u9ed8\u8ba4\u63a8\u62c9\u95e8\u9ad8\u5ea6{rules.default_sliding_door_height:g}m"
    return f"\u95e8\u6d1e\u7f3a\u5c11\u9ad8\u5ea6\u65f6\u9ed8\u8ba4\u95e8\u6d1e\u9ad8\u5ea6{rules.default_door_height:g}m"


def _review_status_for_room(item_name: str, room: QuantityRow) -> str:
    if item_name in _CURTAIN_BOX_ITEMS and room.window_details:
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    if room.status == DataStatus.DEFAULT_INFERRED:
        if _only_window_height_defaulted(room) and not _room_item_uses_window_height(item_name):
            return "\u81ea\u52a8\u751f\u6210"
        return "\u81ea\u52a8\u751f\u6210-\u9ed8\u8ba4\u63a8\u65ad"
    if room.status == DataStatus.NEEDS_REVIEW:
        return "\u81ea\u52a8\u751f\u6210-\u5f02\u5e38\u63d0\u793a"
    return "\u81ea\u52a8\u751f\u6210"


def _review_note_for_room(item_name: str, room: QuantityRow) -> str | None:
    if item_name in _CURTAIN_BOX_ITEMS and room.window_details:
        return "\u7a97\u5e18\u7bb1\u6309\u540c\u623f\u95f4\u540c\u5899\u6bb5\u53bb\u91cd\uff0cL\u5f62\u7a97\u9700\u4eba\u5de5\u786e\u8ba4"
    if room.status not in {DataStatus.DEFAULT_INFERRED, DataStatus.NEEDS_REVIEW}:
        return None
    notes = _translated_exception_notes_for_room(
        room,
        include_window_height_note=_room_item_uses_window_height(item_name),
    )
    if not notes:
        if room.exception_notes:
            return None
        return f"\u7b97\u91cf\u72b6\u6001\uff1a{room.status.value}"
    return "\uff1b".join(notes)


def _translated_exception_notes_for_room(
    room: QuantityRow,
    *,
    include_window_height_note: bool = True,
) -> list[str]:
    notes = [note for note in room.exception_notes if not _is_window_height_default_note(note)]
    window_note = _window_height_default_note_for_room(room) if include_window_height_note else None
    if window_note is not None:
        notes.insert(0, window_note)
    return notes


def _room_item_uses_window_height(item_name: str) -> bool:
    return item_name in {
        "\u5899\u9762\u754c\u9762\u5242\u5904\u7406",
        "\u5899\u9762\u6279\u5d4c",
        "\u5899\u9762\u4e73\u80f6\u6f06",
        "\u5899\u9762\u8d34\u74f7\u7816(600x1200)",
    }


def _only_window_height_defaulted(room: QuantityRow) -> bool:
    if not room.exception_notes:
        return False
    return all(_is_window_height_default_note(note) for note in room.exception_notes)


def _window_height_default_note_for_room(room: QuantityRow) -> str | None:
    raw_notes = [note for note in room.exception_notes if _is_window_height_default_note(note)]
    defaulted_windows = [window for window in room.window_details if window.height_defaulted]
    count = len(defaulted_windows) or len(raw_notes)
    if count == 0:
        return None

    heights = [window.height for window in defaulted_windows if window.height is not None]
    for note in raw_notes:
        match = re.search(r"default height ([0-9]+(?:\.[0-9]+)?)", note)
        if match:
            heights.append(float(match.group(1)))
    unique_heights = sorted({round(height, 6) for height in heights})
    if len(unique_heights) == 1:
        height_label = f"{unique_heights[0]:g}m"
    elif unique_heights:
        height_label = "/".join(f"{height:g}m" for height in unique_heights)
    else:
        height_label = "\u9ed8\u8ba4\u503c"
    return f"\u7a97\u9ad8\u7f3a\u5931{count}\u4e2a\uff0c\u5df2\u6309\u9ed8\u8ba4\u7a97\u9ad8{height_label}\u8ba1\u7b97"


def _is_window_height_default_note(note: str) -> bool:
    return "window_height_defaulted" in note or re.search(r"\bWindow\b.+\bused default height\b", note) is not None


def _measure_basis_for_item(item_name: str, room: QuantityRow) -> str:
    if item_name == "\u5899\u5730\u9762\u9632\u6f0f\u5904\u7406":
        if "\u53a8\u623f" in room.room_name:
            return "\u5730\u9762\u9762\u79ef+0.3m\u4ee5\u4e0b\u5899\u9762\u9762\u79ef"
        return "\u5730\u9762\u9762\u79ef+1.8m\u4ee5\u4e0b\u5899\u9762\u9762\u79ef"
    if item_name in {
        "\u8f7b\u94a2\u9f99\u9aa8\u5e73\u9876",
        "\u9876\u9762\u6279\u5d4c",
        "\u9876\u9762\u4e73\u80f6\u6f06",
        "\u5730\u9762\u7816\u94fa\u8d34(750X1500)",
        "\u5730\u9762\u627e\u5e73",
    }:
        return "\u5730\u9762\u9762\u79ef"
    if item_name in {
        "\u5899\u9762\u754c\u9762\u5242\u5904\u7406",
        "\u5899\u9762\u6279\u5d4c",
        "\u5899\u9762\u4e73\u80f6\u6f06",
    }:
        return "\u5899\u9762\u51c0\u9762\u79ef"
    if item_name in {
        "\u5899\u9762\u8d34\u74f7\u7816(600x1200)",
        "\u5899\u9762\u8d34\u74f7\u7816(600X1200)",
    }:
        if _is_wet_room(room):
            return "2.5m\u4ee5\u4e0b\u5899\u9762\u8d34\u7816\u9762\u79ef"
        return "QUOTE_WALL_TILE\u663e\u5f0f\u5899\u7816\u9762\u79ef"
    if item_name in _CURTAIN_BOX_ITEMS:
        return "\u7a97\u6240\u5728\u5899\u9762\u957f\u5ea6"
    return "\u81ea\u52a8\u7b97\u91cf"


def _wet_wall_tile_area(room: QuantityRow, rules: ResidentialQuoteRules) -> float:
    large_window_area = sum(
        detail.area for detail in room.window_details if detail.area > _WALL_TILE_WINDOW_DEDUCTION_THRESHOLD
    )
    if not room.window_details and room.window_area > _WALL_TILE_WINDOW_DEDUCTION_THRESHOLD:
        large_window_area = room.window_area
    sliding_door_area = sum(
        _door_quote_area(door, rules)
        for door in room.door_details
        if door.width is not None and door.width >= rules.wide_door_width_threshold
    )
    return _round_quantity(max(0, room.wall_measure_perimeter * rules.wall_tile_height - large_window_area - sliding_door_area))


def _explicit_wall_tile_area_for_room(
    room: QuantityRow,
    construction_details: list[ConstructionQuantityDetail],
) -> float:
    return _round_quantity(
        sum(
            detail.area
            for detail in construction_details
            if detail.kind is ConstructionKind.WALL_TILE and detail.room_id == room.room_id and detail.area > 0
        )
    )


def _curtain_wall_length(rooms: list[QuantityRow]) -> float:
    segment_lengths: dict[str, float] = {}
    fallback_total = 0.0
    for room in rooms:
        for detail in room.window_details:
            if "+" in detail.id:
                fallback_total += detail.width
            elif detail.wall_segment_key and detail.wall_segment_length is not None:
                segment_lengths[f"{room.room_id}:{detail.wall_segment_key}"] = detail.wall_segment_length
            else:
                fallback_total += detail.width
    return _round_quantity(sum(segment_lengths.values()) + fallback_total)


def _tile_spec_from_text(*values: str | None) -> tuple[str, float, float]:
    text = " ".join(value or "" for value in values)
    match = re.search(r"(\d+(?:\.\d+)?)\s*([xX*×])\s*(\d+(?:\.\d+)?)", text)
    if match is None:
        raise ValueError(f"Cannot parse tile specification from quote item '{text.strip()}'")
    width = float(match.group(1))
    height = float(match.group(3))
    spec_label = f"{match.group(1)}{match.group(2)}{match.group(3)}"
    if width > 20:
        width /= 1000
    if height > 20:
        height /= 1000
    if width <= 0 or height <= 0:
        raise ValueError(f"Tile specification must be positive in quote item '{text.strip()}'")
    return spec_label, width, height


def _tile_piece_count(area: float, spec: tuple[str, float, float], loss_rate: float) -> int:
    piece_area = spec[1] * spec[2]
    if piece_area <= 0:
        return 0
    return int(math.ceil(area * (1 + loss_rate) / piece_area - 1e-9))


def _tile_piece_basis(area_label: str, spec: tuple[str, float, float], loss_rate: float) -> str:
    loss_label = f"{loss_rate * 100:g}%"
    return f"{area_label}\u6309{spec[0]}\u89c4\u683c+{loss_label}\u635f\u8017\u6298\u7b97\u7247\u6570"


def _ordinary_interior_door_count(rooms: list[QuantityRow], rules: ResidentialQuoteRules) -> tuple[int, list[QuantityRow]]:
    bathroom_door_ids = {
        door.id
        for room in rooms
        if _is_bathroom(room)
        for door in room.door_details
    }
    seen_door_ids: set[str] = set()
    source_rooms: dict[str, QuantityRow] = {}
    for room in rooms:
        if _is_bathroom(room) or not any(keyword in room.room_name for keyword in _INTERIOR_DOOR_SOURCE_ROOM_KEYWORDS):
            continue
        for door in room.door_details:
            if (
                door.width is None
                or door.width >= rules.wide_door_width_threshold
                or door.id in seen_door_ids
                or door.id in bathroom_door_ids
            ):
                continue
            seen_door_ids.add(door.id)
            source_rooms[room.room_id] = room
    return len(seen_door_ids), list(source_rooms.values())


def _sliding_door_area(rooms: list[QuantityRow], rules: ResidentialQuoteRules, item_name: str) -> tuple[float, list[QuantityRow]]:
    seen_door_ids: set[str] = set()
    source_rooms: dict[str, QuantityRow] = {}
    area = 0.0
    keywords = _sliding_door_keywords_for_item(rules.sliding_door_room_keywords_by_item, item_name, rules)
    for room in rooms:
        if not any(keyword in room.room_name for keyword in keywords):
            continue
        for door in room.door_details:
            if door.width is None or door.width < rules.wide_door_width_threshold or door.id in seen_door_ids:
                continue
            seen_door_ids.add(door.id)
            source_rooms[room.room_id] = room
            area += _door_quote_area(door, rules)
    return _round_quantity(area), list(source_rooms.values())


def _sliding_door_trim_length(rooms: list[QuantityRow], rules: ResidentialQuoteRules, item_name: str) -> tuple[float, list[QuantityRow]]:
    seen_door_ids: set[str] = set()
    source_rooms: dict[str, QuantityRow] = {}
    length = 0.0
    keywords = _sliding_door_keywords_for_item(rules.sliding_door_trim_room_keywords_by_item, item_name, rules)
    for room in rooms:
        if not any(keyword in room.room_name for keyword in keywords):
            continue
        for door in room.door_details:
            if door.width is None or door.width < rules.wide_door_width_threshold or door.id in seen_door_ids:
                continue
            seen_door_ids.add(door.id)
            source_rooms[room.room_id] = room
            length += door.width + 2 * _door_quote_height(door, rules)
    return _round_quantity(length), list(source_rooms.values())


def _sliding_door_keywords_for_item(
    keywords_by_item: dict[str, set[str]],
    item_name: str,
    rules: ResidentialQuoteRules,
) -> set[str]:
    return keywords_by_item.get(item_name) or rules.sliding_door_room_keywords


def _is_optional_zero_sliding_item(item_name: str, keywords_by_item: dict[str, set[str]]) -> bool:
    keywords = keywords_by_item.get(item_name, set())
    return any(keyword in {"\u9633\u53f0", "\u9732\u53f0"} for keyword in keywords)


def _door_quote_area(door, rules: ResidentialQuoteRules) -> float:
    if door.width is not None and door.height_defaulted:
        return door.width * rules.default_sliding_door_height
    return door.area


def _door_quote_height(door, rules: ResidentialQuoteRules) -> float:
    if door.height_defaulted:
        return rules.default_door_height
    if door.effective_height is not None:
        return door.effective_height
    if door.height is not None:
        return door.height
    return rules.default_door_height


def _should_generate_room_section(room: QuantityRow) -> bool:
    if room.status == DataStatus.EXCLUDED:
        return False
    return room.include_in_floor_quantity or room.include_in_wall_paint_quantity


def _room_has_wall_tile(room: QuantityRow) -> bool:
    return _is_wet_room(room)


def _is_wet_room(room: QuantityRow) -> bool:
    return _is_kitchen(room) or _is_bathroom(room)


def _is_kitchen(room: QuantityRow) -> bool:
    return "\u53a8\u623f" in room.room_name


def _is_bathroom(room: QuantityRow) -> bool:
    return any(keyword in room.room_name for keyword in ["\u536b", "\u6d17\u624b\u95f4", "\u536b\u751f\u95f4"])


def _is_section_row(number: Any, sheet, row_index: int) -> bool:
    return isinstance(number, str) and _as_text(sheet.cell(row=row_index, column=2).value) is not None


def _is_skip_row(name: str) -> bool:
    return name.strip().replace(" ", "") in {"\u5c0f\u8ba1"} or name.startswith("\u7f16\u5236\u8bf4\u660e")


def _is_space_template_section(name: str) -> bool:
    if any(keyword in name for keyword in _NON_SPACE_SECTION_KEYWORDS):
        return False
    return any(keyword in name for keyword in _SPACE_TEMPLATE_KEYWORDS)


def _section_label(index: int) -> str:
    if index <= len(_SECTION_NUMERALS):
        return _SECTION_NUMERALS[index - 1]
    return str(index)


def _round_quantity(value: float) -> float:
    return round(value, 6)


def _number_or_zero(value: Any) -> float | int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    return float(value)


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
