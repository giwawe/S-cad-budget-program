from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cad_budget.models import QuantityResult


HEADERS = [
    "楼层",
    "空间名称",
    "空间类型",
    "层高",
    "地面面积",
    "地面周长",
    "墙面计量周长",
    "开放边界长度",
    "墙面毛面积",
    "窗数量",
    "窗面积",
    "门洞数量",
    "门洞面积",
    "墙面净面积",
    "是否室外空间",
    "是否计入室内地面",
    "是否计入室内墙面乳胶漆",
    "识别状态",
    "异常说明",
]
_HIDDEN_HEADERS = ["空间ID", "报价明细JSON"]

EXTERIOR_HEADERS = [
    "楼层",
    "外墙编号",
    "高度",
    "外墙计量长度",
    "外墙洞口长度",
    "外墙毛面积",
    "外墙净面积",
]

CONSTRUCTION_HEADERS = [
    "楼层",
    "标识编号",
    "类型",
    "长度",
    "高度",
    "有效高度",
    "高度是否默认",
    "厚度",
    "面积",
    "数量",
]

_COLUMN_WIDTHS = {
    "A": 10,
    "B": 16,
    "C": 14,
    "D": 10,
    "E": 12,
    "F": 12,
    "G": 16,
    "H": 16,
    "I": 14,
    "J": 10,
    "K": 12,
    "L": 12,
    "M": 12,
    "N": 14,
    "O": 14,
    "P": 16,
    "Q": 20,
    "R": 14,
    "S": 36,
}

_EXTERIOR_COLUMN_WIDTHS = {
    "A": 10,
    "B": 18,
    "C": 10,
    "D": 16,
    "E": 16,
    "F": 14,
    "G": 14,
}

_EDITABLE_COLUMNS = {"A", "B", "C", "D", "G", "H", "J", "K", "L", "M", "O", "P", "Q", "R", "S"}
_FORMULA_COLUMNS = {"I", "N"}
_NUMERIC_COLUMNS = {"D", "E", "F", "G", "H", "I", "K", "M", "N"}
_EXTERIOR_NUMERIC_COLUMNS = {"C", "D", "E", "F", "G"}
_CONSTRUCTION_NUMERIC_COLUMNS = {"D", "E", "F", "H", "I", "J"}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_FORMULA_FILL = PatternFill("solid", fgColor="D9EAD3")
_STATIC_FILL = PatternFill("solid", fgColor="E7E6E6")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def export_quantity_result(result: QuantityResult, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "算量表"
    sheet["A1"] = "项目名称"
    sheet["B1"] = result.project_name
    sheet["A1"].font = _BOLD_FONT
    sheet.append([])
    if result.building_area is not None:
        sheet["A2"] = "建筑面积"
        sheet["B2"] = result.building_area
    sheet.append(HEADERS)

    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(result.rows, start=4):
        sheet.append(
            [
                row.floor,
                row.room_name,
                row.space_type.value,
                row.height,
                row.floor_area,
                row.floor_perimeter,
                row.wall_measure_perimeter,
                row.open_boundary_length,
                f"=G{row_index}*D{row_index}",
                row.window_count,
                row.window_area,
                row.door_opening_count,
                row.door_opening_area,
                f"=I{row_index}-K{row_index}",
                row.is_outdoor,
                row.include_in_floor_quantity,
                row.include_in_wall_paint_quantity,
                row.status.value,
                "；".join(row.exception_notes),
                row.room_id,
                _quote_details_json(row),
            ]
        )

    _style_data_rows(sheet)
    _configure_sheet(sheet)
    if result.exterior_rows:
        _create_exterior_sheet(workbook, result)
    if result.construction_details:
        _create_construction_sheet(workbook, result)
    workbook.save(output_path)


def _style_data_rows(sheet) -> None:
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(HEADERS) + len(_HIDDEN_HEADERS)):
        for cell in row:
            column = get_column_letter(cell.column)
            if column in _FORMULA_COLUMNS:
                cell.fill = _FORMULA_FILL
            elif column in _EDITABLE_COLUMNS:
                cell.fill = _EDITABLE_FILL
            else:
                cell.fill = _STATIC_FILL

            if column in _NUMERIC_COLUMNS:
                cell.number_format = "0.###"

            cell.alignment = Alignment(vertical="top", wrap_text=column == "S")


def _configure_sheet(sheet) -> None:
    sheet.freeze_panes = "A4"
    hidden_start = len(HEADERS) + 1
    for index, header in enumerate(_HIDDEN_HEADERS, start=hidden_start):
        cell = sheet.cell(row=3, column=index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    max_column = len(HEADERS) + len(_HIDDEN_HEADERS)
    sheet.auto_filter.ref = f"A3:{get_column_letter(max_column)}{sheet.max_row}"
    sheet.row_dimensions[3].height = 24

    for column, width in _COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    for index in range(hidden_start, max_column + 1):
        sheet.column_dimensions[get_column_letter(index)].hidden = True


def _quote_details_json(row) -> str:
    details = {
        "window_details": [detail.model_dump(mode="json") for detail in row.window_details],
        "door_details": [detail.model_dump(mode="json") for detail in row.door_details],
        "custom_details": [detail.model_dump(mode="json") for detail in row.custom_details],
        "cabinet_details": [detail.model_dump(mode="json") for detail in row.cabinet_details],
    }
    if not any(details.values()):
        return ""
    return json.dumps(details, ensure_ascii=False, separators=(",", ":"))


def _create_exterior_sheet(workbook: Workbook, result: QuantityResult) -> None:
    sheet = workbook.create_sheet("外墙表")
    sheet["A1"] = "项目名称"
    sheet["B1"] = result.project_name
    sheet["A1"].font = _BOLD_FONT
    sheet.append([])
    sheet.append(EXTERIOR_HEADERS)

    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in result.exterior_rows:
        sheet.append(
            [
                row.floor,
                row.exterior_wall_id,
                row.height,
                row.measure_length,
                row.opening_length,
                row.gross_area,
                row.net_area,
            ]
        )

    for data_row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(EXTERIOR_HEADERS)):
        for cell in data_row:
            column = get_column_letter(cell.column)
            cell.fill = _STATIC_FILL
            if column in _EXTERIOR_NUMERIC_COLUMNS:
                cell.number_format = "0.###"
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(EXTERIOR_HEADERS))}{sheet.max_row}"
    sheet.row_dimensions[3].height = 24
    for column, width in _EXTERIOR_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width


def _create_construction_sheet(workbook: Workbook, result: QuantityResult) -> None:
    sheet = workbook.create_sheet("施工标识表")
    sheet["A1"] = "项目名称"
    sheet["B1"] = result.project_name
    sheet["A1"].font = _BOLD_FONT
    sheet.append([])
    sheet.append(CONSTRUCTION_HEADERS)

    for cell in sheet[3]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for detail in result.construction_details:
        sheet.append(
            [
                detail.floor,
                detail.id,
                detail.kind.value,
                detail.length,
                detail.height,
                detail.effective_height,
                detail.height_defaulted,
                detail.thickness,
                detail.area,
                detail.count,
            ]
        )

    for data_row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(CONSTRUCTION_HEADERS)):
        for cell in data_row:
            column = get_column_letter(cell.column)
            cell.fill = _STATIC_FILL
            if column in _CONSTRUCTION_NUMERIC_COLUMNS:
                cell.number_format = "0.###"
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(CONSTRUCTION_HEADERS))}{sheet.max_row}"
    sheet.row_dimensions[3].height = 24
