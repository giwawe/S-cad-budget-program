from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_OUTPUTS = (
    "quote-priced.xlsx",
    "quote-priced-review.md",
    "quote-priced-review.json",
    "quote-priced-review-checklist.xlsx",
)
DEFAULT_OUTPUT_DIR = Path("scratch") / "cad-import-10-real-template-priced-command"
DEFAULT_AUTOMATION_COUNTS = {"自动算量": 53, "自动汇总": 46, "模板默认": 0}
DEFAULT_STATUS_COUNTS = {"自动生成-默认推断": 38, "自动生成-异常提示": 0, "按模板生成": 0}


class PricedQuoteOutputError(Exception):
    pass


def check_priced_quote_outputs(
    output_dir: Path,
    *,
    unit_prices_path: Path | None = None,
    expected_automation_counts: dict[str, int] | None = None,
    expected_status_counts: dict[str, int] | None = None,
) -> dict[str, Any]:
    paths = _required_paths(output_dir)
    for name, path in paths.items():
        if not path.exists():
            raise PricedQuoteOutputError(f"Missing priced quote output: {name} ({path})")

    automation_counts = _automation_counts(paths["quote-priced.xlsx"])
    status_counts = _status_counts(paths["quote-priced-review.json"])

    _assert_counts("Automation count", automation_counts, expected_automation_counts)
    _assert_counts("Review status count", status_counts, expected_status_counts)

    matched_unit_price_rows = 0
    if unit_prices_path is not None:
        if not unit_prices_path.exists():
            raise PricedQuoteOutputError(f"Unit price workbook does not exist: {unit_prices_path}")
        matched_unit_price_rows = _assert_unit_prices_applied(paths["quote-priced.xlsx"], unit_prices_path)

    return {
        "output_dir": str(output_dir),
        "files": len(REQUIRED_OUTPUTS),
        "automation_counts": automation_counts,
        "status_counts": status_counts,
        "matched_unit_price_rows": matched_unit_price_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Check generated priced quote outputs.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory containing quote-priced outputs.")
    parser.add_argument("--unit-prices", type=Path, default=None, help="Optional unit price workbook to verify against quote rows.")
    parser.add_argument(
        "--skip-default-counts",
        action="store_true",
        help="Only check files and unit prices; do not require the real-template default counts.",
    )
    args = parser.parse_args()

    expected_automation = None if args.skip_default_counts else DEFAULT_AUTOMATION_COUNTS
    expected_status = None if args.skip_default_counts else DEFAULT_STATUS_COUNTS
    try:
        summary = check_priced_quote_outputs(
            args.output_dir,
            unit_prices_path=args.unit_prices,
            expected_automation_counts=expected_automation,
            expected_status_counts=expected_status,
        )
    except (OSError, ValueError, PricedQuoteOutputError) as exc:
        raise SystemExit(str(exc))

    print("Priced quote output checks passed")
    print(f"Output directory: {summary['output_dir']}")
    print("Automation counts: " + _format_counts(summary["automation_counts"]))
    print("Review status counts: " + _format_counts(summary["status_counts"]))
    if args.unit_prices is not None:
        print(f"Matched unit price rows: {summary['matched_unit_price_rows']}")


def _required_paths(output_dir: Path) -> dict[str, Path]:
    return {name: output_dir / name for name in REQUIRED_OUTPUTS}


def _automation_counts(quote_path: Path) -> dict[str, int]:
    sheet = load_workbook(quote_path, data_only=True).active
    labels = {"自动算量", "自动汇总", "模板默认"}
    counts = {label: 0 for label in labels}
    for row_index in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row_index, column=17).value
        if label in counts:
            counts[label] = int(sheet.cell(row=row_index, column=18).value or 0)
    return {label: counts[label] for label in ("自动算量", "自动汇总", "模板默认")}


def _status_counts(review_json_path: Path) -> dict[str, int]:
    data = json.loads(review_json_path.read_text(encoding="utf-8"))
    raw_counts = data.get("status_counts", {})
    return {
        "自动生成-默认推断": int(raw_counts.get("自动生成-默认推断", 0)),
        "自动生成-异常提示": int(raw_counts.get("自动生成-异常提示", 0)),
        "按模板生成": int(raw_counts.get("按模板生成", 0)),
    }


def _assert_counts(label: str, actual: dict[str, int], expected: dict[str, int] | None) -> None:
    if expected is None:
        return
    for key, expected_value in expected.items():
        actual_value = actual.get(key, 0)
        if actual_value != expected_value:
            raise PricedQuoteOutputError(f"{label} mismatch for {key}: expected {expected_value}, got {actual_value}")


def _assert_unit_prices_applied(quote_path: Path, unit_prices_path: Path) -> int:
    expected_prices = _load_unit_prices(unit_prices_path)
    matched = 0
    sheet = load_workbook(quote_path, data_only=True).active
    for row_index in range(1, sheet.max_row + 1):
        item_name = _as_text(sheet.cell(row=row_index, column=2).value)
        unit = _as_text(sheet.cell(row=row_index, column=3).value)
        key = (item_name, unit)
        if key not in expected_prices:
            continue
        actual = (
            _number(sheet.cell(row=row_index, column=5).value),
            _number(sheet.cell(row=row_index, column=6).value),
            _number(sheet.cell(row=row_index, column=7).value),
        )
        expected = expected_prices[key]
        if actual != expected:
            raise PricedQuoteOutputError(
                f"Unit price mismatch at quote row {row_index} for {item_name}({unit}): "
                f"expected {expected}, got {actual}"
            )
        matched += 1
    if matched == 0:
        raise PricedQuoteOutputError(f"No quote rows matched unit price workbook: {unit_prices_path}")
    return matched


def _load_unit_prices(unit_prices_path: Path) -> dict[tuple[str, str], tuple[float | int, float | int, float | int]]:
    sheet = load_workbook(unit_prices_path, data_only=True).active
    header = [_as_text(sheet.cell(row=1, column=column).value) for column in range(1, sheet.max_column + 1)]
    required = ["项目名称", "单位", "主材单价", "辅材单价", "人工单价"]
    columns = {name: header.index(name) + 1 for name in required}
    prices: dict[tuple[str, str], tuple[float | int, float | int, float | int]] = {}
    for row_index in range(2, sheet.max_row + 1):
        item_name = _as_text(sheet.cell(row=row_index, column=columns["项目名称"]).value)
        if not item_name:
            continue
        unit = _as_text(sheet.cell(row=row_index, column=columns["单位"]).value)
        prices[(item_name, unit)] = (
            _number(sheet.cell(row=row_index, column=columns["主材单价"]).value),
            _number(sheet.cell(row=row_index, column=columns["辅材单价"]).value),
            _number(sheet.cell(row=row_index, column=columns["人工单价"]).value),
        )
    return prices


def _as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> float | int:
    if value is None:
        raise PricedQuoteOutputError("Unit price value is empty")
    number = float(value)
    if number.is_integer():
        return int(number)
    return number


def _format_counts(counts: dict[str, int]) -> str:
    return ", ".join(f"{label}={count}" for label, count in counts.items())


if __name__ == "__main__":
    main()
