from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_OUTPUT_DIR = Path("scratch") / "cad-import-10-real-template-current"
TOLERANCE = 0.000001


def assert_real_template_key_results(output_dir: Path) -> None:
    result_path = output_dir / "result.json"
    quote_path = output_dir / "quote.xlsx"
    if not result_path.exists():
        raise AssertionError(f"Missing quantity result JSON: {result_path}")
    if not quote_path.exists():
        raise AssertionError(f"Missing quote workbook: {quote_path}")

    result = json.loads(result_path.read_text(encoding="utf-8"))
    quote_rows = list(load_workbook(quote_path, data_only=False).active.iter_rows(values_only=True))

    _assert_close(result.get("building_area"), 136.237652, "building_area")
    _assert_close(_quote_quantity(quote_rows, "瓷砖加工费"), 136.237652, "瓷砖加工费 quantity")
    _assert_close(_quote_quantity(quote_rows, "地面砖现场维护费"), 116.615998, "地面砖现场维护费 quantity")

    main_bedroom = _room(result, "主卧")
    main_bedroom_windows = main_bedroom.get("window_details") or []
    assert len(main_bedroom_windows) == 1, "主卧 should have one merged L-shaped window"
    _assert_close(main_bedroom_windows[0].get("width"), 3.467, "主卧 merged L-shaped window width")

    kitchen = _room(result, "厨房")
    assert kitchen.get("window_area", 0) < 3, "厨房窗洞面积 should stay below the no-deduct threshold"
    kitchen_wall_tile = _quote_row(quote_rows, "墙面贴瓷砖(600x1200)", source_space="厨房")
    _assert_close(kitchen_wall_tile[3], 14.954467, "厨房墙砖 quantity")
    assert kitchen_wall_tile[12] == "2.5m以下墙面贴砖面积"

    _assert_dark_curtain_boxes(quote_rows)

    assert _quote_quantity(quote_rows, "室内门") == 3
    _assert_close(_quote_quantity(quote_rows, "厨房推拉门"), 3.843228, "厨房推拉门 quantity")
    _assert_close(_quote_quantity(quote_rows, "厨房推拉门双包套"), 6.146922, "厨房推拉门双包套 quantity")
    assert "2.2m" in (_quote_row(quote_rows, "厨房推拉门")[14] or "")
    assert "2.2m" in (_quote_row(quote_rows, "厨房推拉门双包套")[14] or "")


def main() -> None:
    parser = argparse.ArgumentParser(description="Assert key business results for the real CAD/template quote sample.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    try:
        assert_real_template_key_results(args.output_dir)
    except (OSError, ValueError, AssertionError) as exc:
        raise SystemExit(str(exc))

    print("Real template key result assertions passed")
    print(f"Output directory: {args.output_dir}")


def _assert_dark_curtain_boxes(quote_rows: list[tuple[Any, ...]]) -> None:
    rows = [row for row in quote_rows if row[1] == "暗窗帘箱"]
    assert len(rows) == 5, "暗窗帘箱 should only be generated for five ordinary dry window rooms"

    spaces = Counter(row[10] for row in rows)
    assert spaces == Counter({"餐厅": 1, "卧室": 2, "主卧": 1, "客厅": 1})
    assert not any(row[10] in {"厨房", "卫生间", "阳台", "露台"} for row in rows)

    quantities = sorted(float(row[3]) for row in rows)
    assert quantities == [1.774, 1.972, 3.467, 4.195, 6.628]
    main_bedroom_box = _quote_row(quote_rows, "暗窗帘箱", source_space="主卧")
    _assert_close(main_bedroom_box[3], 3.467, "主卧暗窗帘箱 quantity")


def _room(result: dict[str, Any], room_name: str) -> dict[str, Any]:
    matches = [row for row in result.get("rows", []) if row.get("room_name") == room_name]
    if not matches:
        raise AssertionError(f"Missing quantity room: {room_name}")
    if len(matches) > 1 and room_name != "卧室" and room_name != "卫生间":
        raise AssertionError(f"Expected one room named {room_name}, found {len(matches)}")
    return matches[0]


def _quote_quantity(quote_rows: list[tuple[Any, ...]], item_name: str) -> Any:
    return _quote_row(quote_rows, item_name)[3]


def _quote_row(quote_rows: list[tuple[Any, ...]], item_name: str, *, source_space: str | None = None) -> tuple[Any, ...]:
    matches = [
        row
        for row in quote_rows
        if row[1] == item_name and (source_space is None or row[10] == source_space)
    ]
    if not matches:
        suffix = f" for source space {source_space}" if source_space is not None else ""
        raise AssertionError(f"Missing quote row: {item_name}{suffix}")
    if len(matches) > 1:
        suffix = f" for source space {source_space}" if source_space is not None else ""
        raise AssertionError(f"Expected one quote row for {item_name}{suffix}, found {len(matches)}")
    return matches[0]


def _assert_close(actual: Any, expected: float, label: str) -> None:
    if actual is None:
        raise AssertionError(f"{label}: expected {expected}, got None")
    if abs(float(actual) - expected) > TOLERANCE:
        raise AssertionError(f"{label}: expected {expected}, got {actual}")


if __name__ == "__main__":
    main()
