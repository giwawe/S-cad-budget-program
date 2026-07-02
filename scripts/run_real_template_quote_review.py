from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from cad_budget.cad_adapter_models import CadImportOptions, CadUnit
from cad_budget.dxf_adapter import import_dxf
from cad_budget.export_excel import export_quantity_result
from cad_budget.quantity import calculate_quantities
from cad_budget.quote_excel import export_residential_quote
from cad_budget.quote_report import build_quote_review_data, generate_quote_review_report
try:
    from scripts.check_priced_quote_outputs import check_priced_quote_outputs
except ModuleNotFoundError:
    from check_priced_quote_outputs import check_priced_quote_outputs


DEFAULT_DXF = Path(r"D:\Desktop\10.dxf")
DEFAULT_TEMPLATE = Path(r"D:\Desktop\清单式报价表（商品房）-修正版.xlsx")
DEFAULT_OUTPUT_DIR = Path("scratch") / "cad-import-10-real-template-current"
PRIORITY_ORDER = {"high": 0, "medium": 1, "low": 2}


class PipelineError(Exception):
    pass


@dataclass(frozen=True)
class QuoteReviewPipelineSummary:
    output_dir: Path
    automation_counts: dict[str, int]
    review_status_counts: dict[str, int]
    action_priority_counts: dict[str, int]
    failed_actions: list[dict[str, Any]]
    gate_failed: bool
    priced_output_dir: Path | None = None
    priced_output_check: dict[str, Any] | None = None

    def to_json_data(self) -> dict[str, Any]:
        return {
            "output_dir": str(self.output_dir),
            "automation_counts": self.automation_counts,
            "review_status_counts": self.review_status_counts,
            "action_priority_counts": self.action_priority_counts,
            "failed_actions": self.failed_actions,
            "gate_failed": self.gate_failed,
            "priced_output_dir": str(self.priced_output_dir) if self.priced_output_dir is not None else None,
            "priced_output_check": self.priced_output_check,
        }


def run_quote_review_pipeline(
    *,
    dxf_path: Path,
    template_path: Path,
    output_dir: Path,
    unit: CadUnit = CadUnit.MILLIMETER,
    rules_path: Path | None = None,
    unit_prices_path: Path | None = None,
    fail_on: str | None = None,
    priced_output_dir: Path | None = None,
    check_priced_output: bool = False,
) -> QuoteReviewPipelineSummary:
    if not dxf_path.exists():
        raise PipelineError(f"Input DXF does not exist: {dxf_path}")
    if not template_path.exists():
        raise PipelineError(f"Quote template does not exist: {template_path}")
    if rules_path is not None and not rules_path.exists():
        raise PipelineError(f"Quote rules file does not exist: {rules_path}")
    if unit_prices_path is not None and not unit_prices_path.exists():
        raise PipelineError(f"Unit price file does not exist: {unit_prices_path}")
    if priced_output_dir is not None and unit_prices_path is None:
        raise PipelineError("Unit price file is required when generating priced quote outputs.")
    if check_priced_output and priced_output_dir is None:
        raise PipelineError("--check-priced-output requires --priced-output-dir.")

    output_dir.mkdir(parents=True, exist_ok=True)
    project_path = output_dir / "project.json"
    result_path = output_dir / "result.json"
    result_excel_path = output_dir / "result.xlsx"
    quote_path = output_dir / "quote.xlsx"
    markdown_path = output_dir / "quote-review.md"
    review_json_path = output_dir / "quote-review.json"
    checklist_path = output_dir / "quote-review-checklist.xlsx"
    summary_path = output_dir / "summary.json"

    import_result = import_dxf(CadImportOptions(source_path=dxf_path, confirmed_unit=unit))
    if import_result.has_blockers or import_result.project is None:
        issues = "\n".join(f"- {issue.severity.value}: {issue.code}: {issue.message}" for issue in import_result.issues)
        raise PipelineError(f"DXF import failed:\n{issues}")

    quantity_result = calculate_quantities(import_result.project)
    export_residential_quote(quantity_result, template_path, quote_path, rules_path=rules_path, unit_prices_path=unit_prices_path)
    export_quantity_result(quantity_result, result_excel_path)
    project_path.write_text(import_result.project.model_dump_json(indent=2), encoding="utf-8")
    result_path.write_text(quantity_result.model_dump_json(indent=2), encoding="utf-8")
    generate_quote_review_report(
        quote_path,
        markdown_path,
        quantity_result=quantity_result,
        json_output=review_json_path,
        checklist_output=checklist_path,
    )

    report_data = build_quote_review_data(quote_path, quantity_result=quantity_result)
    automation_counts = _automation_counts(quote_path)
    review_status_counts = report_data["status_counts"]
    action_priority_counts = _action_priority_counts(report_data["actions"])
    priced_output_check: dict[str, Any] | None = None
    if priced_output_dir is not None:
        _generate_priced_outputs(
            quantity_result=quantity_result,
            template_path=template_path,
            output_dir=priced_output_dir,
            rules_path=rules_path,
            unit_prices_path=unit_prices_path,
        )
        if check_priced_output:
            priced_output_check = check_priced_quote_outputs(
                priced_output_dir,
                unit_prices_path=unit_prices_path,
                expected_automation_counts=automation_counts,
                expected_status_counts=review_status_counts,
            )

    failed_actions = _failed_actions(report_data["actions"], fail_on)
    summary = QuoteReviewPipelineSummary(
        output_dir=output_dir,
        automation_counts=automation_counts,
        review_status_counts=review_status_counts,
        action_priority_counts=action_priority_counts,
        failed_actions=failed_actions,
        gate_failed=bool(failed_actions),
        priced_output_dir=priced_output_dir,
        priced_output_check=priced_output_check,
    )
    summary_path.write_text(json.dumps(summary.to_json_data(), ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the real-template CAD quote review regression pipeline.")
    parser.add_argument("--dxf", type=Path, default=DEFAULT_DXF, help="DXF file to import.")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="Residential quote template workbook.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated outputs.")
    parser.add_argument("--unit", choices=[unit.value for unit in CadUnit], default=CadUnit.MILLIMETER.value)
    parser.add_argument("--rules", type=Path, default=None, help="Optional residential quote rules JSON.")
    parser.add_argument("--unit-prices", type=Path, default=None, help="Optional global unit price workbook.")
    parser.add_argument("--fail-on", choices=["high", "medium"], default=None, help="Return exit code 1 at this review priority threshold.")
    parser.add_argument("--priced-output-dir", type=Path, default=None, help="Optional directory for quote-priced outputs.")
    parser.add_argument("--check-priced-output", action="store_true", help="Check quote-priced outputs after generation.")
    args = parser.parse_args()

    try:
        summary = run_quote_review_pipeline(
            dxf_path=args.dxf,
            template_path=args.template,
            output_dir=args.output_dir,
            unit=CadUnit(args.unit),
            rules_path=args.rules,
            unit_prices_path=args.unit_prices,
            fail_on=args.fail_on,
            priced_output_dir=args.priced_output_dir,
            check_priced_output=args.check_priced_output,
        )
    except (OSError, ValueError, PipelineError) as exc:
        raise SystemExit(str(exc))

    print("Quote review pipeline complete")
    print(f"Output directory: {summary.output_dir}")
    print(f"Automation counts: {_format_counts(summary.automation_counts)}")
    print(f"Review status counts: {_format_counts(summary.review_status_counts)}")
    print(f"Action priority counts: {_format_counts(summary.action_priority_counts)}")
    if summary.priced_output_dir is not None:
        print(f"Priced output directory: {summary.priced_output_dir}")
    if summary.priced_output_check is not None:
        print(f"Priced output matched unit price rows: {summary.priced_output_check['matched_unit_price_rows']}")
    if summary.gate_failed:
        print("Gate failed actions: " + ", ".join(action["label"] for action in summary.failed_actions))
        raise SystemExit(1)


def _generate_priced_outputs(
    *,
    quantity_result,
    template_path: Path,
    output_dir: Path,
    rules_path: Path | None,
    unit_prices_path: Path | None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    quote_path = output_dir / "quote-priced.xlsx"
    markdown_path = output_dir / "quote-priced-review.md"
    review_json_path = output_dir / "quote-priced-review.json"
    checklist_path = output_dir / "quote-priced-review-checklist.xlsx"
    export_residential_quote(
        quantity_result,
        template_path,
        quote_path,
        rules_path=rules_path,
        unit_prices_path=unit_prices_path,
    )
    generate_quote_review_report(
        quote_path,
        markdown_path,
        quantity_result=quantity_result,
        json_output=review_json_path,
        checklist_output=checklist_path,
    )


def _automation_counts(quote_path: Path) -> dict[str, int]:
    sheet = load_workbook(quote_path, data_only=False).active
    labels = ["自动算量", "自动汇总", "模板默认"]
    counts: dict[str, int] = {label: 0 for label in labels}
    for row_index in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row_index, column=17).value
        if label in counts:
            counts[label] = int(sheet.cell(row=row_index, column=18).value or 0)
    return counts


def _action_priority_counts(actions: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"high": 0, "medium": 0, "low": 0}
    for action in actions:
        priority = action["priority"]
        counts[priority] = counts.get(priority, 0) + 1
    return counts


def _failed_actions(actions: list[dict[str, Any]], fail_on: str | None) -> list[dict[str, Any]]:
    if fail_on is None:
        return []
    threshold = PRIORITY_ORDER[fail_on]
    return [action for action in actions if PRIORITY_ORDER.get(action["priority"], 99) <= threshold]


def _format_counts(counts: dict[str, int]) -> str:
    return ", ".join(f"{label}={count}" for label, count in counts.items())


if __name__ == "__main__":
    main()
