from __future__ import annotations

import argparse
from pathlib import Path

import ezdxf
from openpyxl import Workbook, load_workbook

from cad_budget.cad_adapter_models import CadImportOptions, CadUnit
from cad_budget.dxf_adapter import import_dxf
from cad_budget.quantity import calculate_quantities
from cad_budget.quote_excel import export_residential_quote


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a small CAD-to-quote sample with quote marker layers."
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("scratch") / "marker-rich-quote-sample",
        help="Directory for generated DXF, JSON, template, quote workbook, and README.",
    )
    args = parser.parse_args()

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    dxf_path = output_dir / "marker-rich-plan.dxf"
    template_path = output_dir / "marker-rich-template.xlsx"
    project_path = output_dir / "project.json"
    result_path = output_dir / "result.json"
    quote_path = output_dir / "quote.xlsx"
    readme_path = output_dir / "README.md"

    build_marker_rich_dxf(dxf_path)
    create_marker_quote_template(template_path)

    import_result = import_dxf(CadImportOptions(source_path=dxf_path, confirmed_unit=CadUnit.MILLIMETER))
    if import_result.has_blockers or import_result.project is None:
        issues = "\n".join(f"- {issue.code}: {issue.message}" for issue in import_result.issues)
        raise SystemExit(f"DXF import failed:\n{issues}")

    quantity = calculate_quantities(import_result.project)
    export_residential_quote(quantity, template_path, quote_path)
    project_path.write_text(import_result.project.model_dump_json(indent=2), encoding="utf-8")
    result_path.write_text(quantity.model_dump_json(indent=2), encoding="utf-8")
    readme_path.write_text(_sample_readme(quote_path), encoding="utf-8")

    print(f"Wrote marker-rich quote sample to {output_dir}")


def build_marker_rich_dxf(path: Path) -> None:
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4
    doc.appids.new("CAD_BUDGET")
    doc.blocks.new("pipe_block")
    modelspace = doc.modelspace()
    for layer in [
        "QUOTE_ROOM",
        "QUOTE_TEXT",
        "QUOTE_DOOR",
        "QUOTE_EXT_WALL",
        "QUOTE_EXT_OPENING",
        "QUOTE_EXT_REPAIR",
        "QUOTE_DEMO_WALL",
        "QUOTE_NEW_WALL",
        "QUOTE_PIPE_INSULATION",
        "QUOTE_PIPE_WRAP",
        "QUOTE_CUSTOM",
        "QUOTE_CABINET",
    ]:
        doc.layers.add(layer)

    modelspace.add_lwpolyline(
        [(0, 0), (4000, 0), (4000, 3000), (0, 3000), (0, 0)],
        dxfattribs={"layer": "QUOTE_ROOM", "closed": True},
    )
    modelspace.add_text("厨房", dxfattribs={"layer": "QUOTE_TEXT", "height": 250}).set_placement((1500, 1200))
    modelspace.add_lwpolyline(
        [(5000, 0), (8000, 0), (8000, 2000), (5000, 2000), (5000, 0)],
        dxfattribs={"layer": "QUOTE_ROOM", "closed": True},
    )
    modelspace.add_text("阳台", dxfattribs={"layer": "QUOTE_TEXT", "height": 250}).set_placement((6000, 1000))

    modelspace.add_lwpolyline(
        [(-500, -500), (8500, -500), (8500, 3500), (-500, 3500), (-500, -500)],
        dxfattribs={"layer": "QUOTE_EXT_WALL", "closed": True},
    )
    modelspace.add_lwpolyline([(1000, -500), (2000, -500)], dxfattribs={"layer": "QUOTE_EXT_OPENING"})
    modelspace.add_lwpolyline(
        [(9000, 0), (11000, 0), (11000, 3000), (9000, 3000), (9000, 0)],
        dxfattribs={"layer": "QUOTE_EXT_REPAIR", "closed": True},
    )

    demo = modelspace.add_lwpolyline([(0, 500), (3000, 500)], dxfattribs={"layer": "QUOTE_DEMO_WALL"})
    demo.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2600")])
    new_120 = modelspace.add_lwpolyline([(0, 1000), (2000, 1000)], dxfattribs={"layer": "QUOTE_NEW_WALL"})
    new_120.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2800"), (1000, "THICKNESS=120")])
    new_240 = modelspace.add_lwpolyline([(0, 1500), (1500, 1500)], dxfattribs={"layer": "QUOTE_NEW_WALL"})
    new_240.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=3000"), (1000, "THICKNESS=240")])

    insulation = modelspace.add_point((3500, 1000), dxfattribs={"layer": "QUOTE_PIPE_INSULATION"})
    insulation.set_xdata("CAD_BUDGET", [(1000, "HEIGHT=2400")])
    pipe_wrap = modelspace.add_blockref("pipe_block", (3500, 1500), dxfattribs={"layer": "QUOTE_PIPE_WRAP"})
    pipe_wrap.add_attrib("HEIGHT", "2.1")

    custom = modelspace.add_line((500, 2200), (2500, 2200), dxfattribs={"layer": "QUOTE_CUSTOM"})
    custom.set_xdata("CAD_BUDGET", [(1000, "TYPE=衣柜")])
    cabinet = modelspace.add_line((500, 2600), (3500, 2600), dxfattribs={"layer": "QUOTE_CABINET"})
    cabinet.set_xdata("CAD_BUDGET", [(1000, "TYPE=地柜")])

    modelspace.add_lwpolyline([(5000, 500), (7000, 500)], dxfattribs={"layer": "QUOTE_DOOR"})

    doc.saveas(path)


def create_marker_quote_template(path: Path) -> None:
    workbook = Workbook()
    half = workbook.active
    half.title = "半包"
    _write_quote_header(half)
    half.append(["一", "半包工程"])
    half.append([1, "半包项目", "M2", 1, 2, 3, 4, None, "不应读取"])

    fitout = workbook.create_sheet("整装")
    _write_quote_header(fitout)
    fitout.append(["一", "标识自动化工程"])
    items = [
        ("外墙批嵌", "M2", 77, "外墙"),
        ("外墙批嵌以及修补", "M2", 88, "外墙修补"),
        ("拆改及拆墙", "M2", 93, "拆墙"),
        ("砌120厚砖墙", "M2", 44.8, "120墙"),
        ("砌240厚砖墙", "M2", 64.56, "240墙"),
        ("砖墙门窗洞过梁", "支", 15, "人工填写"),
        ("打混凝土过梁孔", "个", 108, "过梁孔"),
        ("厨房、卫生间排污管包隔音棉", "M", 50, "隔音棉"),
        ("包上/下水管道(单管)", "M", 36, "包管"),
        ("全屋定制", "M2", 66, "全屋定制"),
        ("橱柜", "M", 22, "橱柜"),
        ("阳台推拉门", "M2", 66, "阳台推拉门"),
        ("阳台推拉门双包套", "M", 55, "阳台推拉门双包套"),
    ]
    for index, (name, unit, quantity, note) in enumerate(items, start=1):
        fitout.append([index, name, unit, quantity, 0, 0, 10, None, note])
    fitout.append([None, "小 计", None, None, None, None, None, "=SUM(H6:H18)"])
    fitout.append(["A", "直接费合计", None, None, None, None, None, "=H19"])
    workbook.save(path)


def _write_quote_header(sheet) -> None:
    sheet.append(["工程(预) 算表"])
    sheet.append(["名称：Marker Demo"])
    sheet.append([
        "编号",
        "项目名称",
        "单位",
        "数量",
        "材料费(元)",
        None,
        "人工费\n(元)",
        "总价(元)",
        "材  料  及  工  艺  说  明",
    ])
    sheet.append([None, None, None, None, "主材\n单价", "辅材\n单价"])


def _sample_readme(quote_path: Path) -> str:
    rows = list(load_workbook(quote_path, data_only=False).active.iter_rows(values_only=True))
    auto_summary = _summary_value(rows, "自动汇总")
    template_default = _summary_value(rows, "模板默认")
    return "\n".join(
        [
            "# Marker-Rich Quote Sample",
            "",
            "This directory is generated by `scripts/generate_marker_rich_quote_sample.py`.",
            "",
            "Generated files:",
            "- `marker-rich-plan.dxf`: small CAD sample with QUOTE_* marker layers.",
            "- `marker-rich-template.xlsx`: minimal residential quote template.",
            "- `project.json`: imported ProjectInput JSON.",
            "- `result.json`: calculated QuantityResult JSON.",
            "- `quote.xlsx`: generated residential quote workbook.",
            "",
            "Covered CAD layers:",
            "- `QUOTE_EXT_WALL` / `QUOTE_EXT_OPENING`",
            "- `QUOTE_EXT_REPAIR`",
            "- `QUOTE_DEMO_WALL`",
            "- `QUOTE_NEW_WALL`",
            "- `QUOTE_PIPE_INSULATION` / `QUOTE_PIPE_WRAP`",
            "- `QUOTE_CUSTOM` / `QUOTE_CABINET`",
            "- `QUOTE_DOOR`",
            "",
            f"自动汇总: {auto_summary}",
            f"模板默认: {template_default}",
            "",
        ]
    )


def _summary_value(rows, label: str):
    for row in rows:
        if row[16] == label:
            return row[17]
    raise ValueError(f"Missing summary row for {label}")


if __name__ == "__main__":
    main()
